import streamlit as st
import pandas as pd
import calendar
import datetime
import random
import requests
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

APP_VERSION = "1.021"

st.set_page_config(page_title=f"層級式護理排班系統 v{APP_VERSION}", layout="wide")

# ── 行政職稱常數 ──────────────────────────────────────────────
# 只排白班（D）的職稱
ADMIN_TITLES  = {"組長", "護理長", "副護理長", "助理", "傷兵"}
# 不排假日班（週六/週日/國定）、不排夜班、不排加班線；且不計入單位人力
NO_HOL_ADMIN  = {"護理長", "副護理長", "助理", "傷兵"}
# 不排假日班（含組長）；組長仍計入單位人力，故使用獨立常數
NO_HOL_SET    = {"護理長", "副護理長", "助理", "傷兵", "組長"}
# 包班人員包班班別下限（達到後，人力額滿時可改排其他班別）
PACK_MIN_SHIFTS = 15

# --- 狀態初始化 ---
if "step" not in st.session_state:
    st.session_state.step = 1
for key in ["base_sched", "pack_sched", "night_sched", "d_sched", "twelve_sched", "final_sched", "classified_sched", "ai_df", "custom_targets"]:
    if key not in st.session_state:
        st.session_state[key] = None

# ── 存檔 / 載入進度 ─────────────────────────────────────────────────────────
_CHECKPOINT_KEYS = [
    "step", "sel_year", "sel_month",
    "month_days", "first_wday",
    "saturdays_list", "sundays_list", "nat_holidays_list",
    "holiday_list", "target_off",
    "ai_df", "default_quota", "edited_quota_df", "edited_weekly_df",
    "skill_cols", "custom_targets", "personal_targets",
    "base_sched", "pack_sched", "night_sched", "d_sched",
    "twelve_sched", "final_sched", "classified_sched",
    "pack_warnings", "target_warnings", "shortages_export",
    "explanation_df", "s6_deficits",
    "extra_staffing_df",
    "prev_month_buffer",
]

def _make_checkpoint() -> bytes:
    import pickle, hashlib, hmac
    data = {k: st.session_state[k] for k in _CHECKPOINT_KEYS if k in st.session_state}
    payload = pickle.dumps(data, protocol=4)
    sig = hmac.new(b"nurse-scheduling-checkpoint-key", payload, hashlib.sha256).digest()
    return sig + payload

def _restore_checkpoint(raw: bytes):
    import pickle, hashlib, hmac
    if len(raw) < 32:
        st.error("❌ 進度檔案格式錯誤或已損壞，無法載入。")
        return
    sig, payload = raw[:32], raw[32:]
    expected_sig = hmac.new(b"nurse-scheduling-checkpoint-key", payload, hashlib.sha256).digest()
    if not hmac.compare_digest(sig, expected_sig):
        st.error("❌ 進度檔案簽章驗證失敗，檔案可能已被竄改，拒絕載入。")
        return
    data = pickle.loads(payload)
    for k, v in data.items():
        st.session_state[k] = v

# ── 顯示用縮寫對照表 ─────────────────────────────────────────────
_ABBREV_MAP = {
    "預假":  "預",
    "預長假": "預",
    "預白":  "Dx",
    "例假":  "例",
    "休假":  "休",
    "國定":  "國",
    "公差":  "公",   # 公差縮寫
    "特休":  "V",    # 特休縮寫
}

def abbrev_display(val):
    """
    將班表顯示值縮短，節省窄欄空間：
      預假/預長假 → 預；預白 → Dx；例假 → 例；休假 → 休；國定 → 國
      其他非工作、非空白的特殊假別 → 取第一個字
      工作班別（D/E/N/12-8/上課/公差）與 O/休/空白 不處理
    """
    v = str(val).strip()
    if not v or v in ("nan", "O", "休"):
        return val
    mapped = _ABBREV_MAP.get(v)
    if mapped is not None:
        return mapped
    # 特殊假別（非工作班別且長度 > 1）→ 取第一個字
    if not is_work(v) and len(v) > 1:
        return v[0]
    return val

def build_schedule_with_counts(disp_df, src_df, day_cols, ai_df_local):
    """
    在顯示用班表下方附加 D/E/N/12-8 每日人數統計列，
    回傳已套色的 Styler 物件，可直接傳入 st.dataframe()。
    disp_df : 已套 abbrev_display / apply_prewhite_dx 的顯示副本
    src_df  : 原始班表（用於計算人數，未經顯示轉換）
    day_cols: 日期欄位名稱清單，如 ["1","2",..."31"]
    ai_df_local: 護理師參數 DataFrame
    """
    _count_rows = []
    for _lbl, _key in [("D班人數", "D"), ("E班人數", "E"), ("N班人數", "N"), ("12-8人數", "12-8")]:
        _row = {"姓名": _lbl}
        for _dc in day_cols:
            try:
                if _key == "D":
                    _cnt = sum(
                        1 for i in ai_df_local.index
                        if isinstance(src_df.at[i, _dc], str)
                        and src_df.at[i, _dc].upper().startswith("D")
                        and str(ai_df_local.at[i, "職稱"]).strip() not in NO_HOL_ADMIN
                    )
                else:
                    # E★/N★（控台班別）與 E/N 同等計入班別人數
                    _e_variants = {"E", "E★", "E*"}
                    _n_variants = {"N", "N★", "N*"}
                    if _key == "E":
                        _cnt = sum(1 for i in ai_df_local.index if str(src_df.at[i, _dc]).strip() in _e_variants)
                    elif _key == "N":
                        _cnt = sum(1 for i in ai_df_local.index if str(src_df.at[i, _dc]).strip() in _n_variants)
                    else:
                        _cnt = sum(1 for i in ai_df_local.index if str(src_df.at[i, _dc]).strip() == _key)
            except Exception:
                _cnt = 0
            _row[_dc] = str(_cnt) if _cnt > 0 else "0"
        _count_rows.append(_row)

    _count_labels = {"D班人數", "E班人數", "N班人數", "12-8人數"}
    _combined = pd.concat([disp_df, pd.DataFrame(_count_rows)], ignore_index=True)

    _color_map = {
        "D班人數": ("#dbeafe", "#1e40af"),
        "E班人數": ("#fef9c3", "#92400e"),
        "N班人數": ("#f3e8ff", "#6b21a8"),
        "12-8人數": ("#d1fae5", "#065f46"),
    }

    def _style_row(row):
        """套用全列：護理師列只套班別色；統計列套彩色底色。"""
        lbl = str(row.iloc[0])
        if lbl not in _color_map:
            # 護理師資料列：日期欄套班別色，其餘不動
            styles = []
            for col_name in row.index:
                if col_name in day_cols:
                    styles.append(color_shifts(row[col_name]))
                else:
                    styles.append("")
            return styles
        # 統計列：整列套彩色底色
        bg, fg = _color_map[lbl]
        styles = []
        for j, col_name in enumerate(row.index):
            val = str(row.iloc[j])
            border = "border-top:2px solid #999;" if j == 0 else "border-top:2px solid #ccc;"
            zero = "color:#bbb;" if val == "0" else f"color:{fg};font-weight:600;"
            if col_name == "姓名":
                styles.append(f"background-color:{bg};color:{fg};font-weight:700;font-size:10px;{border}")
            else:
                styles.append(f"background-color:{bg};{zero}font-size:11px;{border}")
        return styles

    return _combined.style.apply(_style_row, axis=1)


def _xmonth_shift(n_idx, d, sched_dict, ai_df_local, month_days_local):
    """
    跨月班別查詢：d >= 1 查本月；d < 1 查前月緩衝（prev_month_buffer）。
    用於 14 日窗口計算，讓排班時即時考慮前月尾段的出勤情況。
    """
    if d >= 1:
        return sched_dict[n_idx][d] if 1 <= d <= month_days_local else ""
    try:
        _nm = str(ai_df_local.at[n_idx, "姓名"]).strip()
    except Exception:
        return ""
    return (st.session_state.get("prev_month_buffer") or {}).get(_nm, {}).get(d, "")


def apply_prewhite_dx(disp_df, ai_df_local, month_days_local):
    """
    將顯示用 DataFrame 中，預白日期對應的 D 格標為 Dx。
    disp_df 已套用 abbrev_display()，此函數在其之後呼叫。
    使用 iat（位置索引）須確保 disp_df 與 ai_df_local 列順序一致。
    """
    for _dx_ri, (_, _dx_row) in enumerate(ai_df_local.iterrows()):
        for _dx_ds in str(_dx_row.get("預白日期", "")).split(","):
            if _dx_ds.strip().isdigit():
                _dx_d = int(_dx_ds.strip())
                if 1 <= _dx_d <= month_days_local:
                    _dx_col = str(_dx_d)
                    if _dx_col in disp_df.columns:
                        _dx_curr = str(disp_df.iat[_dx_ri, disp_df.columns.get_loc(_dx_col)]).strip()
                        if _dx_curr == "D":
                            disp_df.iat[_dx_ri, disp_df.columns.get_loc(_dx_col)] = "Dx"
    return disp_df

# 定義全域顏色渲染函數
def color_shifts(val):
    v = str(val).upper().strip()
    if v.startswith('D'): return 'background-color: #d4edda; color: #155724;'
    if v == 'E': return 'background-color: #fff3cd; color: #856404;'
    if v in ('E★', 'E*'): return 'background-color: #ffe08a; color: #6b4c00; font-weight: bold;'  # E加班
    if v == 'N': return 'background-color: #e2d9f3; color: #4a148c;'
    if v in ('N★', 'N*'): return 'background-color: #c9b8f0; color: #3a006f; font-weight: bold;'  # N加班
    if v == '12-8': return 'background-color: #cce5ff; color: #004085;'
    if v == 'O': return 'background-color: #dee2e6; color: #495057;'
    if v == '休': return 'background-color: #f1f3f5; color: #adb5bd;'
    if v == '上課': return 'background-color: #ffeeba; color: #856404; font-weight: bold;'
    if v == '公差': return 'background-color: #fad7a0; color: #784212; font-weight: bold;'
    if v == '公':   return 'background-color: #fad7a0; color: #784212; font-weight: bold;'  # 公差縮寫
    # 預覽用標示（帶入班表時特別顯示，實際資料仍儲存為 O / D）
    v_raw = str(val).strip()
    if v_raw == '預假':   return 'background-color: #cff4fc; color: #055160; font-style: italic;'
    if v_raw == '預長假': return 'background-color: #ffe5b4; color: #7a3e00; font-weight: bold;'
    if v_raw == '預白':   return 'background-color: #a8d5b5; color: #0a3622; font-weight: bold;'
    # 縮寫顯示對應色彩
    if v_raw == '預': return 'background-color: #cff4fc; color: #055160; font-style: italic;'
    if v_raw == '例': return 'background-color: #f8d7da; color: #721c24; font-weight: bold;'
    if v_raw == '國': return 'background-color: #ffecd2; color: #8a4519; font-weight: bold;'
    if v_raw == 'V':  return 'background-color: #f3e8ff; color: #6b21a8; font-weight: bold;'  # 特休
    if v != '': return 'background-color: #ffe4e8; color: #c0392b;'  # 特殊假別（公假、喪假等）
    return ''

def color_classified(val):
    """Step 7 假別分類視圖專用色彩"""
    v = str(val).strip()
    vu = v.upper()
    if vu.startswith('D'): return 'background-color:#d4edda;color:#155724;'
    if vu == 'E':          return 'background-color:#fff3cd;color:#856404;'
    if vu in ('E★','E*'): return 'background-color:#ffe08a;color:#6b4c00;font-weight:bold;'
    if vu == 'N':          return 'background-color:#e2d9f3;color:#4a148c;'
    if vu in ('N★','N*'): return 'background-color:#c9b8f0;color:#3a006f;font-weight:bold;'
    if v == '12-8':  return 'background-color:#cce5ff;color:#004085;'
    if v == '上課':  return 'background-color:#ffeeba;color:#856404;font-weight:bold;'
    if v == '公差':  return 'background-color:#fad7a0;color:#784212;font-weight:bold;'
    if v == '公':    return 'background-color:#fad7a0;color:#784212;font-weight:bold;'  # 公差縮寫
    if v == '例假':  return 'background-color:#f8d7da;color:#721c24;font-weight:bold;'
    if v == '休假':  return 'background-color:#e2e3e5;color:#383d41;'
    if v == '國定':  return 'background-color:#ffecd2;color:#8a4519;font-weight:bold;'
    if v == '特休':  return 'background-color:#f3e8ff;color:#6b21a8;font-weight:bold;'  # 特休
    if v == '預假':  return 'background-color:#cff4fc;color:#055160;font-style:italic;'
    if v == '預長假': return 'background-color:#ffe5b4;color:#7a3e00;font-weight:bold;'
    # 縮寫顯示對應色彩
    if v == '預': return 'background-color:#cff4fc;color:#055160;font-style:italic;'
    if v == '例': return 'background-color:#f8d7da;color:#721c24;font-weight:bold;'
    if v == '休': return 'background-color:#e2e3e5;color:#383d41;'
    if v == '國': return 'background-color:#ffecd2;color:#8a4519;font-weight:bold;'
    if v == 'V':  return 'background-color:#f3e8ff;color:#6b21a8;font-weight:bold;'  # 特休縮寫
    if v == 'O':  return 'background-color:#d1ecf1;color:#0c5460;'
    if v != '':   return 'background-color:#ffe4e8;color:#c0392b;'  # 其他特殊假別
    return ''

def is_work(val):
    v = str(val).strip()
    vu = v.upper()
    # E★/N★/D★ 為加班班別，仍算出勤；D值由 startswith("D") 覆蓋
    # 公差 / 公（縮寫）算出勤；特休/V、國/例/休 不算出勤
    if vu in ("E★", "E*"): return True
    if vu in ("N★", "N*"): return True
    return vu in ["D", "E", "N", "12-8", "上課", "公差", "公"] or vu.startswith("D")

_CORE_SHIFTS = {"D", "E", "N", "12-8"}

def _shift_norm(v):
    """將班別值正規化為核心班別鍵（D/E/N/12-8）或 None"""
    vs = str(v).strip()
    vu = vs.upper()
    if not vs or vu in ("O", "NAN", ""): return None
    # E★/N★：加班班別，歸回 E/N 計算換班間距
    if vu in ("E★", "E*"): return "E"
    if vu in ("N★", "N*"): return "N"
    # D值（行政假日白班）、上課、公差、公（公差縮寫）均視為 D 班
    if vu.startswith("D") or vs in ("上課", "公差", "公"): return "D"
    if vs in _CORE_SHIFTS: return vs
    return None

# ── A/B 組別排班條件 ─────────────────────────────────────────────
# 硬性上限：12-8 每日 A≤3 B≤3；E 每日 A≤2 B≤2
# 軟性最低保障（評分加分）：週六 D 班 A≥3；週日 D 班 A≥2 B≥2
# A/B 互排規則：A 組可排入 B 組（A 超額後借用 B 剩餘空間），B 組不可排入 A 組
_GROUP_SHIFT_CAP = {"12-8": {"A": 3, "B": 3}, "E": {"A": 2, "B": 2}}
_GROUP_D_SAT_MIN = {"A": 3}
_GROUP_D_SUN_MIN = {"A": 2, "B": 2}

def group_cap_ok(n_idx, s, d_int, sched_dict, cache_group_local):
    """
    回傳 True = 該護理師可在 d_int 排 s 班（組別人數未超上限）。
    僅對 12-8 / E 班做硬性組別人數管制；D 班最低保障由評分機制處理。
    無組別設定（空白）者不受限制。

    A/B 互排規則：
      A 組：自身配額未滿時直接允許；自身配額已滿時可借用 B 組剩餘空間（A 可排入 B）。
      B 組：只能使用自身配額，不可借用 A 組空間（B 不可排入 A）。
    """
    grp = cache_group_local.get(n_idx, "")
    if grp not in ("A", "B"):
        return True
    caps = _GROUP_SHIFT_CAP.get(s)
    if not caps:
        return True

    cap_a = caps.get("A", 999)
    cap_b = caps.get("B", 999)
    curr_a = sum(1 for i, sv in sched_dict.items()
                 if cache_group_local.get(i, "") == "A" and sv[d_int] == s)
    curr_b = sum(1 for i, sv in sched_dict.items()
                 if cache_group_local.get(i, "") == "B" and sv[d_int] == s)

    if grp == "A":
        # A 組：先用自身配額，滿了可借用 B 組剩餘空間
        if curr_a < cap_a:
            return True
        return curr_b < cap_b   # A 超額後借 B 組空間
    else:
        # B 組：只能使用自身配額，不可排入 A 組空間
        return curr_b < cap_b

def group_d_score(n_idx, d_int, sat_set, sun_set, sched_dict, cache_group_local):
    """
    D 班組別最低保障評分：
      週六 A 組未達 3 人時，A 組護理師加 20_000_000 分
      週日 A/B 組未達 2 人時，各自加 20_000_000 分
    無組別設定者回傳 0。
    """
    grp = cache_group_local.get(n_idx, "")
    if grp not in ("A", "B"):
        return 0
    if d_int in sat_set:
        mn = _GROUP_D_SAT_MIN.get(grp, 0)
    elif d_int in sun_set:
        mn = _GROUP_D_SUN_MIN.get(grp, 0)
    else:
        return 0
    if mn <= 0:
        return 0
    curr = sum(
        1 for i, sv in sched_dict.items()
        if cache_group_local.get(i, "") == grp
        and isinstance(sv[d_int], str) and sv[d_int].startswith("D")
    )
    return 20_000_000 if curr < mn else 0

def week_variety_ok(sched_dict, n_idx, s, d_int, first_wday, month_days):
    """
    Rule 1：週日至隔週週六（Sun-Sat 窗格）內，同一人最多出現 2 種核心班別（D/E/N/12-8）。
    回傳 True = 可排；False = 已有 2 種且本次要加第 3 種。
    """
    wday = (first_wday + d_int - 1) % 7   # 0=Mon … 6=Sun
    days_since_sun = (wday + 1) % 7         # Sun=0, Mon=1, …, Sat=6
    week_start = max(1, d_int - days_since_sun)
    week_end   = min(month_days, week_start + 6)

    s_norm = _shift_norm(s)
    if s_norm is None:
        return True   # 非核心班別，不受限制

    types_in_week = set()
    types_in_week.add(s_norm)
    for dd in range(week_start, week_end + 1):
        if dd == d_int: continue
        v_norm = _shift_norm(sched_dict[n_idx][dd])
        if v_norm:
            types_in_week.add(v_norm)
    return len(types_in_week) <= 2

def can_work_holiday_check(n_idx, d_int, cache_can_sat, cache_can_sun, cache_can_nat, sat_list, sun_list, nat_list):
    """
    假日出勤能力判斷（依護理師屬性設定）：
      - 能上週日=是 AND 能上國定假日=是 → 可上所有假日（含週六）
      - 僅能上週六=是（週日/國定皆否）  → 只能上週六
      - 都未標示                         → 不可上任何假日班
    """
    is_sat = d_int in sat_list
    is_sun = d_int in sun_list
    is_nat = d_int in nat_list

    if not (is_sat or is_sun or is_nat):
        return True  # 平日，無假日限制

    can_sat = cache_can_sat.get(n_idx, False)
    can_sun = cache_can_sun.get(n_idx, False)
    can_nat = cache_can_nat.get(n_idx, False)

    # 全假日能力：能上週日 且 能上國定假日 → 三種假日都可上
    if can_sun and can_nat:
        return True

    # 僅週六能力：只能上週六
    if can_sat:
        return is_sat

    # 無假日能力
    return False

def calc_extra_leaves(row, month_days, sat_set=None, sun_set=None, nat_set=None, target_off=None):
    """
    計算個人額外扣除的休假天數。

    ── target_off 傳入時（新邏輯）─────────────────────────────────────────────
      ① 預約長假日期 → 只有「平日」才扣一天應上班天數（假日跳過）。
      ② 特殊假別（病假、事假、公假、喪假等）→ 不論平假日一律扣一天應上班天數。
      ③ 預休日期（O）→ 一律不扣（已包含於 target_off 應休預算，對一般護理師無影響）。
      ④ 例外（target_off == 0，即 NO_HOL_SET：護理長/組長等）：
         這類人員無固定假日休假預算，其平日預休仍需扣除；
         且假日公差可獲補休一天，抵消等量的平日O日扣減。

    ── target_off=None 時（舊邏輯，向下相容）──────────────────────────────────
      - 預休日期 (O) → 不扣
      - 預約長假 + 特殊假別 → 只有平日才扣工作天數

    sat_set / sun_set / nat_set: 該月的週六、週日、國定假日日期集合（int）
    """
    hol_set = set()
    if sat_set: hol_set |= set(sat_set)
    if sun_set: hol_set |= set(sun_set)
    if nat_set: hol_set |= set(nat_set)

    def is_weekday(d_int):
        """True = 平日（非六日非國定）"""
        return d_int not in hol_set if hol_set else True

    if target_off is not None:
        # ── 新邏輯 ────────────────────────────────────────────────────────────
        # ① 特殊假別：不論平假日一律扣
        special_leave_days = 0
        sp_str = str(row.get("特殊假別", "")).strip()
        if sp_str:
            for item in sp_str.split(","):
                sep = ":" if ":" in item else "-"
                parts = item.split(sep, 1)
                if len(parts) >= 1 and parts[0].strip().isdigit():
                    d_int_sp = int(parts[0].strip())
                    if 1 <= d_int_sp <= month_days:   # 不過濾平假日
                        special_leave_days += 1

        # ② 長假：只有平日才扣
        long_leave_days = 0
        long_leave_str = str(row.get("預約長假日期", "")).strip()
        if long_leave_str:
            long_leave_days = sum(
                1 for d in long_leave_str.split(",")
                if d.strip().isdigit()
                and 1 <= int(d.strip()) <= month_days
                and is_weekday(int(d.strip()))
            )

        if target_off == 0:
            # ── NO_HOL_SET（護理長/組長等，無固定假日預算）───────────────────
            # 平日預休（O）仍需扣減
            pre_o_days = 0
            pre_o_str = str(row.get("預休日期", "")).strip()
            if pre_o_str:
                pre_o_days = sum(
                    1 for d in pre_o_str.split(",")
                    if d.strip().isdigit()
                    and 1 <= int(d.strip()) <= month_days
                    and is_weekday(int(d.strip()))
                )
            excess_rest = pre_o_days + long_leave_days

            # 假日公差補休抵扣：假日出公差屬額外出勤，每一假日公差補休一天
            gongcha_str = str(row.get("公差日期", "")).strip()
            if gongcha_str:
                gongcha_hol_credit = sum(
                    1 for d in gongcha_str.split(",")
                    if d.strip().isdigit()
                    and 1 <= int(d.strip()) <= month_days
                    and not is_weekday(int(d.strip()))   # 落在假日的公差才計入
                )
                excess_rest = max(0, excess_rest - gongcha_hol_credit)

            return excess_rest + special_leave_days
        else:
            # ── 一般護理師：O 一律不扣，長假平日直接扣 ─────────────────────
            return long_leave_days + special_leave_days
    else:
        # ── 舊邏輯（向下相容）：O 不扣，長假 + 特殊假別只有平日才扣 ──────────
        special_leave_days = 0
        sp_str = str(row.get("特殊假別", "")).strip()
        if sp_str:
            for item in sp_str.split(","):
                sep = ":" if ":" in item else "-"
                parts = item.split(sep, 1)
                if len(parts) >= 1 and parts[0].strip().isdigit():
                    d_int_sp = int(parts[0].strip())
                    if 1 <= d_int_sp <= month_days and is_weekday(d_int_sp):
                        special_leave_days += 1

        long_leave_days = 0
        long_leave_str = str(row.get("預約長假日期", "")).strip()
        if long_leave_str:
            long_leave_days = sum(
                1 for d in long_leave_str.split(",")
                if d.strip().isdigit()
                and 1 <= int(d.strip()) <= month_days
                and is_weekday(int(d.strip()))
            )
        return long_leave_days + special_leave_days

# ============================================================
# 🎨 直接用 openpyxl 寫入顏色的彩色排班表匯出函數
# ============================================================
SHIFT_FILL = {
    'D':    PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'),
    'E':    PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'),
    'N':    PatternFill(start_color='E2D9F3', end_color='E2D9F3', fill_type='solid'),
    '12-8': PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid'),
    'O':    PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid'),  # 預約長假：藍
    '休':   PatternFill(start_color='F1F3F5', end_color='F1F3F5', fill_type='solid'),  # 休息日：淺灰
    '上課': PatternFill(start_color='FFEEBA', end_color='FFEEBA', fill_type='solid'),
    '公差': PatternFill(start_color='FAD7A0', end_color='FAD7A0', fill_type='solid'),  # 公差：淺橙
    '例假': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),  # 例假：淡紅
    '休假': PatternFill(start_color='E2E3E5', end_color='E2E3E5', fill_type='solid'),  # 休假日：中灰
    '國定': PatternFill(start_color='FFECD2', end_color='FFECD2', fill_type='solid'),  # 國定假日：橘金
    '預假': PatternFill(start_color='CFF4FC', end_color='CFF4FC', fill_type='solid'),  # 預休假：淺藍
    '預長假':PatternFill(start_color='FFE5B4', end_color='FFE5B4', fill_type='solid'),  # 預長假：淺橙
    'OTHER':PatternFill(start_color='FFE4E8', end_color='FFE4E8', fill_type='solid'),  # 特殊假別：淡粉
}
SHIFT_FONT_COLOR = {
    'D': '155724', 'E': '856404', 'N': '4A148C',
    '12-8': '004085', 'O': '0C5460', '休': 'ADB5BD', '上課': '856404', '公差': '784212',
    '例假': '721C24', '休假': '383D41', '國定': '8A4519',
    '預假': '055160', '預長假': '7A3E00', 'OTHER': 'C0392B',
}
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

def get_shift_key(val):
    """根據班別字串回傳顏色分類鍵"""
    v = str(val).strip()
    vu = v.upper()
    if not v or vu == 'NAN': return None
    if vu.startswith('D'): return 'D'
    if vu == 'E': return 'E'
    if vu == 'N': return 'N'
    if v == '12-8': return '12-8'
    if v == 'O': return 'O'
    if v == '上課': return '上課'
    if v == '公差': return '公差'
    if v == '例假': return '例假'
    if v == '休假': return '休假'
    if v == '國定': return '國定'
    if v == '預假': return '預假'
    if v == '預長假': return '預長假'
    # 縮寫顯示對應原始鍵
    if v == '預': return '預假'
    if v == '例': return '例假'
    if v == '休': return '休'
    if v == '國': return '國定'
    return 'OTHER'

def write_df_to_sheet(ws, df, day_cols=None, freeze_col=2):
    """將 DataFrame 寫入工作表，day_cols 欄位會套用班別顏色"""
    day_col_set = set(day_cols) if day_cols else set()
    headers = list(df.columns)

    # 表頭
    header_fill = PatternFill(start_color='343A40', end_color='343A40', fill_type='solid')
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

    # 資料列
    for row_num, (_, row_data) in enumerate(df.iterrows(), 2):
        for col_idx, h in enumerate(headers, 1):
            val = row_data[h]
            display_val = '' if pd.isna(val) else str(val)
            cell = ws.cell(row=row_num, column=col_idx, value=display_val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER
            cell.font = Font(size=10)

            if h in day_col_set and display_val:
                key = get_shift_key(display_val)
                if key:
                    cell.fill = SHIFT_FILL[key]
                    is_bold = (key == '上課')
                    cell.font = Font(color=SHIFT_FONT_COLOR[key], bold=is_bold, size=10)

        # 隔行底色（非班別欄）
        if row_num % 2 == 0:
            for col_idx, h in enumerate(headers, 1):
                if h not in day_col_set:
                    ws.cell(row=row_num, column=col_idx).fill = PatternFill(
                        start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')

    # 欄寬自動調整
    for col_idx, h in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        if h in day_col_set:
            ws.column_dimensions[col_letter].width = 5
        else:
            max_len = max(
                (len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, ws.max_row + 1)),
                default=4
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

    # 凍結第一列 + 前幾欄
    ws.freeze_panes = ws.cell(row=2, column=freeze_col + 1)
    ws.row_dimensions[1].height = 25

def build_colored_excel(final_sched_df, stats_df, explanation_df, shortages_export, month_days, per_week_df=None, violation_df=None):
    """建立全彩 Excel 並回傳 BytesIO 物件"""
    wb = Workbook()
    day_cols = [str(d) for d in range(1, month_days + 1)]

    # 空白格與預休(O)替換為「休」，特殊假別保留原代碼
    _REST_VALS = {"", "O"}
    display_df = final_sched_df.copy()
    for c in day_cols:
        display_df[c] = display_df[c].apply(
            lambda v: "休" if str(v).strip() in _REST_VALS else abbrev_display(v)
        )

    # === 工作表 1：全彩排班表（含每日人數統計列）===
    ws1 = wb.active
    ws1.title = "🗓️ 全彩排班表"

    _count_rows_xl = []
    for _lbl_xl, _key_xl in [("D班人數", "D"), ("E班人數", "E"), ("N班人數", "N"), ("12-8人數", "12-8")]:
        _row_xl = {"姓名": _lbl_xl}
        for _dc_xl in day_cols:
            try:
                if _key_xl == "D":
                    _cnt_xl = sum(
                        1 for _ri in range(len(final_sched_df))
                        if isinstance(final_sched_df.iat[_ri, final_sched_df.columns.get_loc(_dc_xl)], str)
                        and final_sched_df.iat[_ri, final_sched_df.columns.get_loc(_dc_xl)].upper().startswith("D")
                    )
                else:
                    _cnt_xl = sum(
                        1 for _ri in range(len(final_sched_df))
                        if str(final_sched_df.iat[_ri, final_sched_df.columns.get_loc(_dc_xl)]).strip() == _key_xl
                    )
            except Exception:
                _cnt_xl = 0
            _row_xl[_dc_xl] = str(_cnt_xl) if _cnt_xl > 0 else "0"
        _count_rows_xl.append(_row_xl)

    _display_with_counts = pd.concat([display_df, pd.DataFrame(_count_rows_xl)], ignore_index=True)
    write_df_to_sheet(ws1, _display_with_counts, day_cols=day_cols, freeze_col=1)

    # 加入圖例說明列
    legend_row = ws1.max_row + 2
    legend_items = [
        ('D 白班', 'D'), ('E 小夜', 'E'), ('N 大夜', 'N'),
        ('12-8 中班', '12-8'), ('公差', '公差'), ('O 預休/長假', 'O'), ('休 休息日', '休'),
        ('上課', '上課'), ('特殊假別', 'OTHER'),
    ]
    ws1.cell(row=legend_row, column=1, value='圖例：').font = Font(bold=True)
    for i, (label, key) in enumerate(legend_items, 2):
        cell = ws1.cell(row=legend_row, column=i, value=label)
        cell.fill = SHIFT_FILL[key]
        cell.font = Font(color=SHIFT_FONT_COLOR[key], bold=True, size=9)
        cell.alignment = Alignment(horizontal='center')

    # === 工作表 2：公平性結算 ===
    ws2 = wb.create_sheet("⚖️ 公平性結算")
    write_df_to_sheet(ws2, stats_df, freeze_col=2)

    # === 工作表 3：差異診斷 ===
    ws3 = wb.create_sheet("🔍 差異診斷")
    write_df_to_sheet(ws3, explanation_df, freeze_col=1)

    # === 工作表 4：警示清單（若有） ===
    if shortages_export is not None and (isinstance(shortages_export, list) and len(shortages_export) > 0 or (hasattr(shortages_export, 'empty') and not shortages_export.empty)):
        ws4 = wb.create_sheet("🚨 安全警示")
        ws4.cell(row=1, column=1, value='系統警示清單').font = Font(bold=True, size=11)
        for r, msg in enumerate(shortages_export, 2):
            cell = ws4.cell(row=r, column=1, value=msg)
            if '🚨' in msg:
                cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
                cell.font = Font(color='721C24')
            else:
                cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
                cell.font = Font(color='856404')
        ws4.column_dimensions['A'].width = 80

    # === 工作表 5：四周變形工時審查 ===
    if per_week_df is not None and violation_df is not None:
        ws5 = wb.create_sheet("📋 四周變形工時審查")
        header_fill_dark = PatternFill(start_color='343A40', end_color='343A40', fill_type='solid')
        header_fill_viol = PatternFill(start_color='495057', end_color='495057', fill_type='solid')

        # 標題與說明
        title_cell = ws5.cell(row=1, column=1, value="勞基法 §30-1 四周變形工時合法性審查")
        title_cell.font = Font(bold=True, size=12)
        cap_cell = ws5.cell(row=2, column=1, value="審查依據：① 任意7天窗格工作天 ≤ 6；② 任意14天窗格工作天 ≤ 12（14休2）；③ 任意28天窗格工作天 ≤ 24（四週制）")
        cap_cell.font = Font(size=9, italic=True)

        # ── 每週統計表 ──────────────────────────────
        ws5.cell(row=4, column=1, value="【每週工時統計】").font = Font(bold=True)
        pw_headers = list(per_week_df.columns)
        for c_idx, h in enumerate(pw_headers, 1):
            hcell = ws5.cell(row=5, column=c_idx, value=h)
            hcell.font = Font(bold=True, color="FFFFFF", size=10)
            hcell.fill = header_fill_dark
            hcell.alignment = Alignment(horizontal="center")
        for r_off, (_, pw_row) in enumerate(per_week_df.iterrows(), 6):
            for c_idx, h in enumerate(pw_headers, 1):
                val = pw_row[h]
                cell = ws5.cell(row=r_off, column=c_idx, value=str(val) if not pd.isna(val) else "")
                cell.alignment = Alignment(horizontal="center")
                if h == "合法判斷" and "🚨" in str(val):
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    cell.font = Font(color="721C24", bold=True)
                elif h == "合法判斷" and "✅" in str(val):
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    cell.font = Font(color="155724")
                if r_off % 2 == 0 and h not in ("合法判斷",):
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

        # ── 違規明細表 ──────────────────────────────
        viol_start = 5 + len(per_week_df) + 3
        ws5.cell(row=viol_start, column=1, value="【違規明細】").font = Font(bold=True)
        viol_headers = list(violation_df.columns)
        for c_idx, h in enumerate(viol_headers, 1):
            hcell = ws5.cell(row=viol_start + 1, column=c_idx, value=h)
            hcell.font = Font(bold=True, color="FFFFFF", size=10)
            hcell.fill = header_fill_viol
            hcell.alignment = Alignment(horizontal="center")
        for r_off, (_, vrow) in enumerate(violation_df.iterrows(), viol_start + 2):
            for c_idx, val in enumerate(vrow, 1):
                vcell = ws5.cell(row=r_off, column=c_idx, value=str(val))
                vcell.alignment = Alignment(horizontal="left" if c_idx == 2 else "center", wrap_text=True)
                if "🚨" in str(val):
                    vcell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    vcell.font = Font(color="721C24")
                elif "✅" in str(val):
                    vcell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    vcell.font = Font(color="155724")

        # 欄寬
        ws5.column_dimensions["A"].width = 18
        ws5.column_dimensions["B"].width = 65
        for col_idx in range(3, len(pw_headers) + 1):
            ws5.column_dimensions[get_column_letter(col_idx)].width = 12
        ws5.freeze_panes = ws5.cell(row=6, column=3)

    # === 工作表 6：欠班歷史 ===
    ws6 = wb.create_sheet("📊 欠班歷史")

    # 備注行（第 1 列）
    note_cell = ws6.cell(row=1, column=1,
        value="本工作表可作為次月排班參考，欠班天數 > 0 者建議次月優先保留平日排班空間。")
    note_cell.font = Font(italic=True, size=10, color="555555")
    ws6.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws6.row_dimensions[1].height = 20

    # 標題列（第 2 列）
    _h6_cols = ["姓名", "本月應上班", "本月實際上班", "欠班天數", "欠班原因提示"]
    _h6_fill = PatternFill(start_color="4A4A4A", end_color="4A4A4A", fill_type="solid")
    for _ci6, _hname in enumerate(_h6_cols, 1):
        _hc = ws6.cell(row=2, column=_ci6, value=_hname)
        _hc.font = Font(bold=True, color="FFFFFF", size=10)
        _hc.fill = _h6_fill
        _hc.alignment = Alignment(horizontal="center")

    # 準備資料
    _red_fill6  = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    _red_font6  = Font(color="721C24")
    _ok_font6   = Font(color="155724")

    _stats_for6 = stats_df[["姓名", "目標上班", "實際上班"]].copy() if "目標上班" in stats_df.columns else pd.DataFrame()

    for _ri6, (_, _sr) in enumerate(_stats_for6.iterrows(), 3):
        _name6    = _sr.get("姓名", "")
        _target6  = int(_sr.get("目標上班", 0))
        _actual6  = int(_sr.get("實際上班", 0))
        _deficit6 = max(0, _target6 - _actual6)
        _hint6    = "請確認假日資格 / 特殊假別數量 / 配額設定" if _deficit6 > 0 else "正常"

        _row_data6 = [_name6, _target6, _actual6, _deficit6, _hint6]
        for _ci6, _val6 in enumerate(_row_data6, 1):
            _c6 = ws6.cell(row=_ri6, column=_ci6, value=_val6)
            _c6.alignment = Alignment(horizontal="left" if _ci6 in (1, 5) else "center")
            if _deficit6 > 0:
                _c6.fill = _red_fill6
                _c6.font = _red_font6
            else:
                _c6.font = _ok_font6

    # 欄寬
    ws6.column_dimensions["A"].width = 14
    ws6.column_dimensions["B"].width = 14
    ws6.column_dimensions["C"].width = 14
    ws6.column_dimensions["D"].width = 12
    ws6.column_dimensions["E"].width = 42
    ws6.freeze_panes = ws6.cell(row=3, column=1)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ============================================================
# ⚖️ 四周變形工時審查引擎（台灣勞基法 §30-1）
# ============================================================
def build_four_week_review(final_sched_df, ai_df, month_days, prev_buffer=None):
    """
    四周變形工時合法性審查。
    法規依據：
      §30-1 規則1：任意 7 天窗格內，工作天數 ≤ 6（每週至少1天例假）
      §30-1 規則2：任意 14 天窗格內，工作天數 ≤ 12（14休2）
      §30-1 規則3：任意 28 天窗格內，工作天數 ≤ 24（四週制）
      §34   規則4：連續班次之間休息時間 ≥ 11 小時
    回傳：(per_week_df, violation_df)
    """
    # ── 班次時間表（用於 §34 換班間距計算） ─────────────────
    # D班：08:00–16:00  E班：16:00–24:00
    # N班：00:00–08:00（大夜班於當日 00:00 開始、08:00 結束）
    # 12-8班：12:00–20:00
    # 換班間距計算：rest = (24 + 次日班開始時) - 當日班結束時
    # 例：N(4號)→D(5號)：(24+8)-8 = 24h ≥ 11h → 合法
    _SHIFT_END_H   = {"D": 16, "E": 24, "N": 8,  "12-8": 20}
    _SHIFT_START_H = {"D": 8,  "E": 16, "N": 0,  "12-8": 12}

    def _timing_key(val):
        """將班別值對應到換班時間 key（D/E/N/12-8），非上班格回傳 None"""
        v = str(val).strip()
        vu = v.upper()
        # D 系列（D/D1-D14/D★/D值 等）
        if vu.startswith("D"):
            return "D"
        # E 及 E★（控台小夜，時間同 E 班）
        if vu in ("E", "E★", "E*"):
            return "E"
        # N 及 N★（控台大夜，時間同 N 班）
        if vu in ("N", "N★", "N*"):
            return "N"
        if v == "12-8":
            return "12-8"
        # 公差 / 公 視同白班時間（08-16）
        if v in ("公差", "公"):
            return "D"
        return None
    # ── 每週固定區塊統計（週1~4 及尾餘） ─────────────────────
    week_blocks = []
    d = 1
    while d <= month_days:
        end = min(d + 6, month_days)
        week_blocks.append((d, end))
        d = end + 1

    per_week_rows = []
    violations = []

    for nurse_idx, row in ai_df.iterrows():
        name  = row["姓名"]
        title = str(row.get("職稱", "")).strip()
        s_vals = list(final_sched_df.iloc[nurse_idx].values[1:])   # index 0 = 姓名

        # ── 建立含前月緩衝的班表查詢函數 ──────────────────────────
        # prev_buffer: {姓名: {-13: 'D', ..., -1: 'N'}}
        # 正數日期查本月 s_vals；負數日期查前月緩衝
        _pname = str(row["姓名"]).strip()
        _pbuf  = (prev_buffer or {}).get(_pname, {})
        def _get_shift(d_idx):
            """d_idx: 正數=本月第d天(1-based)，負數=前月倒數"""
            if d_idx >= 1:
                return s_vals[d_idx - 1] if 1 <= d_idx <= month_days else ""
            return _pbuf.get(d_idx, "")

        week_counts = []
        for (wb_start, wb_end) in week_blocks:
            cnt = sum(1 for d in range(wb_start, wb_end + 1) if is_work(s_vals[d - 1]))
            days_in_week = wb_end - wb_start + 1
            week_counts.append((cnt, days_in_week))

        # ── 滑動窗格違規掃描 ──────────────────────────────────
        nurse_viols = []

        # ── 本月內規則（改用 _get_shift 查詢）──────────

        # 規則 1：護理人員例假可移動，但兩次例假間距不可超過 12 天
        # 等同：連續出勤不能超過 12 天（若超過，必有某段 13 天以上無例假）
        # 另外，手動班表中連續 6 天出勤屬警示（自動排班已限制≤5天）
        _li_days_full = [d for d in range(1, month_days + 1)
                         if str(_get_shift(d)).strip() in ("例", "例假")]
        # 加入前月最後一個例假（若有緩衝）
        if _pbuf:
            _prev_li = [d for d in range(-13, 0) if str(_pbuf.get(d,"")).strip() in ("例","例假")]
            _li_anchor = max(_prev_li) if _prev_li else None
        else:
            _li_anchor = None

        _all_li = (([_li_anchor] if _li_anchor else []) + _li_days_full)

        # 檢查月初到第一個例假的間距
        _first_li = _li_days_full[0] if _li_days_full else None
        if _first_li:
            _gap_before = _first_li - (_li_anchor if _li_anchor else -13) - 1
            if _gap_before > 12:
                nurse_viols.append(
                    f"月初前 13 天內無例假（最近例假在前月第{_li_anchor or '?'}日），"
                    f"例假間距 {_gap_before + 1} 天 > 12 天"
                )
        elif not _li_days_full:
            nurse_viols.append("本月整月無例假，違反護理人員例假規定")

        # 檢查相鄰例假間距
        for _j in range(1, len(_li_days_full)):
            _gap = _li_days_full[_j] - _li_days_full[_j - 1]
            if _gap > 13:  # 兩例假之間超過 13 天（即中間超過 12 天無例假）
                nurse_viols.append(
                    f"第 {_li_days_full[_j-1]} 日到第 {_li_days_full[_j]} 日例假，"
                    f"間距 {_gap} 天 > 13 天（違反例假間距 ≤ 12 天規定）"
                )

        # 檢查最後一個例假到月底的間距
        if _li_days_full:
            _gap_after = month_days - _li_days_full[-1]
            if _gap_after > 12:
                nurse_viols.append(
                    f"最後例假在第 {_li_days_full[-1]} 日，月底前尚有 {_gap_after} 天無例假 > 12 天"
                )

        # 連續出勤超過 5 天：警示（手動調整可能發生，但自動排班不應產生）
        _consec = 0
        for d in range(1, month_days + 1):
            if is_work(_get_shift(d)):
                _consec += 1
                if _consec > 5:
                    nurse_viols.append(
                        f"第 {d - _consec + 1}~{d} 日連續出勤 {_consec} 天，"
                        f"超過 5 天（⚠️ 手動調整所致，請確認是否符合院方規定）"
                    )
                    break  # 只報告第一個違規段，避免重複
            else:
                _consec = 0

        # 規則 2：任意 14 天窗格（含跨月延伸）
        _r2_start = max(1, 1 - 13) if _pbuf else 1  # 有前月緩衝時從本月第1天往前延伸
        for start in range(_r2_start, month_days - 12):
            end = start + 13
            if end < 1: continue  # 窗格完全在前月，跳過
            work_cnt = sum(1 for d in range(start, end + 1) if is_work(_get_shift(d)))
            if work_cnt > 12:
                loc = f"前月末至本月第{end}日" if start < 1 else f"第{start}~{end}日"
                nurse_viols.append(f"14日內工作{work_cnt}天（{loc}），應 ≤ 12 天（14休2）")

        # 規則 3：任意 28 天窗格，工作天 ≤ 24
        for start in range(1, month_days - 26):
            end = start + 27
            work_cnt = sum(1 for d in range(start, end + 1) if is_work(_get_shift(d)))
            if work_cnt > 24:
                nurse_viols.append(f"28日內工作{work_cnt}天（第{start}~{end}日），應 ≤ 24 天（四週制）")

        # 規則 5（8週制）：任意 56 天窗格，工作天 ≤ 48
        for start in range(1, month_days - 54):
            end = start + 55
            if end > month_days: break
            work_cnt = sum(1 for d in range(start, end + 1) if is_work(_get_shift(d)))
            if work_cnt > 48:
                nurse_viols.append(f"56日內工作{work_cnt}天（第{start}~{end}日），應 ≤ 48 天（8週制）")

        # 規則 4：§34 換班間距（含跨月：前月最後1天→本月第1天）
        _prev_last = _get_shift(-1)
        k_prev = _timing_key(_prev_last)
        k_first = _timing_key(_get_shift(1))
        if k_prev and k_first:
            rest_h0 = (24 + _SHIFT_START_H[k_first]) - _SHIFT_END_H[k_prev]
            if rest_h0 < 11:
                nurse_viols.append(
                    f"前月最後1日（{k_prev}班）→ 本月第1日（{k_first}班）：換班間距僅 {rest_h0} 小時，"
                    f"違反 §34 應 ≥ 11 小時（跨月交接）"
                )
        for d in range(1, month_days):
            k1 = _timing_key(_get_shift(d))
            k2 = _timing_key(_get_shift(d + 1))
            if k1 and k2:
                rest_h = (24 + _SHIFT_START_H[k2]) - _SHIFT_END_H[k1]
                if rest_h < 11:
                    nurse_viols.append(
                        f"第{d}日（{k1}班）→ 第{d+1}日（{k2}班）：換班間距僅 {rest_h} 小時，"
                        f"違反 §34 應 ≥ 11 小時規定"
                    )

        # 去除重複違規提示
        nurse_viols = list(dict.fromkeys(nurse_viols))

        total_work = sum(1 for v in s_vals if is_work(v))
        status = "✅ 合法" if not nurse_viols else f"🚨 違規 ({len(nurse_viols)} 項)"

        row_data = {
            "姓名":   name,
            "職稱":   title if title in ADMIN_TITLES else "護理師",
            "全月工作天": total_work,
        }
        for i, (cnt, days_in_wk) in enumerate(week_counts):
            label = f"W{i+1}({days_in_wk}天)"
            flag  = "🚨" if cnt > 6 else ("⚠️" if cnt == 6 else "")
            row_data[label] = f"{cnt}{flag}"
        row_data["合法判斷"] = status

        per_week_rows.append(row_data)

        if nurse_viols:
            for v in nurse_viols:
                violations.append({"姓名": name, "違規說明": v})

    per_week_df  = pd.DataFrame(per_week_rows)
    violation_df = pd.DataFrame(violations) if violations else pd.DataFrame({"姓名": ["（無違規）"], "違規說明": ["所有人員皆符合四週變形工時規定 ✅"]})
    return per_week_df, violation_df

# 💡 定義共用的全視角安全雷達引擎
def display_safety_radar(sched_df, quota_df, ai_df_local):
    shortages = []
    month_days_local = len([c for c in sched_df.columns if str(c).isdigit()])
    
    for d in range(1, month_days_local + 1):
        day_str = str(d)
        day_q = quota_df[quota_df["日期"] == day_str].iloc[0]
        
        for s_c, s_n in [("N", "N班"), ("E", "E班"), ("12-8", "12-8"), ("D", "D班")]:
            req_c = int(day_q[s_n])
            if req_c > 0:
                if s_c == "D":
                    act_nurses = [i for i in ai_df_local.index
                                  if str(sched_df.at[i, day_str]).startswith("D")
                                  and str(ai_df_local.at[i, "職稱"]).strip() not in NO_HOL_ADMIN]
                else:
                    # E★/N★（控台班別）視同 E/N，納入人力計算
                    _s_variants = {s_c}
                    if s_c == "E": _s_variants = {"E", "E★", "E*"}
                    elif s_c == "N": _s_variants = {"N", "N★", "N*"}
                    act_nurses = [i for i in ai_df_local.index
                                  if str(sched_df.at[i, day_str]).strip() in _s_variants
                                  and str(ai_df_local.at[i, "職稱"]).strip() not in NO_HOL_ADMIN]
                    
                act_c = len(act_nurses)
                if act_c < req_c:
                    shortages.append(f"⚠️ {day_str}號：{s_c} 班缺 {req_c - act_c} 人")
                
                if act_c > 0:
                    target_circ = (req_c + 1) // 2  
                    act_circ = sum(1 for i in act_nurses if str(ai_df_local.at[i, "流動資格"]).strip() == "是")
                    if act_circ < target_circ: 
                        shortages.append(f"🚨 {day_str}號：{s_c} 班 流動短缺，需 {target_circ} 人，僅排 {act_circ} 人")
                    
                    # 💡 階梯式控台授權檢測 (排除 12-8 班，因為通常已有白班或小夜控台在場)
                    if s_c != "12-8":
                        has_leader = False
                        for i in act_nurses:
                            l_str = str(ai_df_local.at[i, "控台資格"]).strip()
                            if not l_str: continue
                            
                            if "白" in l_str: 
                                has_leader = True # 白班資格：無敵星星，所有班都能控
                            elif "小" in l_str and s_c in ["E", "N"]: 
                                has_leader = True # 小夜資格：可控小夜、大夜
                            elif "大" in l_str and s_c == "N": 
                                has_leader = True # 大夜資格：只能控大夜
                        
                        if not has_leader: 
                            shortages.append(f"🚨 {day_str}號：{s_c} 班 缺乏專屬控台指揮官！")
                        
    # ── A/B 組別平衡警示 ─────────────────────────────────────────────────────
    try:
        cache_group_radar = {}
        for i in ai_df_local.index:
            try:
                cache_group_radar[i] = str(ai_df_local.at[i, "組別"]).strip().upper()
            except Exception:
                cache_group_radar[i] = ""

        sat_set_r = set(st.session_state.get("saturdays_list", []))
        sun_set_r = set(st.session_state.get("sundays_list", []))

        for d in range(1, month_days_local + 1):
            day_str = str(d)

            # 12-8 班：A組需3人、B組需3人（僅在當日 12-8 配額 > 0 時才做組別平衡檢查，避免節假日假陽性）
            _req_128_r = int(quota_df[quota_df["日期"] == day_str].iloc[0]["12-8"]) if not quota_df[quota_df["日期"] == day_str].empty else 0
            nurses_128 = [i for i in ai_df_local.index if sched_df.at[i, day_str] == "12-8"]
            if _req_128_r > 0 and nurses_128:
                a128 = sum(1 for i in nurses_128 if cache_group_radar.get(i) == "A")
                b128 = sum(1 for i in nurses_128 if cache_group_radar.get(i) == "B")
                if a128 < 3:
                    shortages.append(f"🚨 {day_str}號：12-8 班 A組僅 {a128} 人（需≥3）")
                if b128 < 3 and (a128 + b128) < 6:  # A組已填補B組空缺時不警示
                    shortages.append(f"🚨 {day_str}號：12-8 班 B組僅 {b128} 人（需≥3，A組補位後合計 {a128+b128} 人）")

            # E 小夜班：A組需2人、B組需2人（同上，僅在 E 班配額 > 0 時才檢查）
            _req_e_r = int(quota_df[quota_df["日期"] == day_str].iloc[0]["E班"]) if not quota_df[quota_df["日期"] == day_str].empty else 0
            nurses_e = [i for i in ai_df_local.index if str(sched_df.at[i, day_str]).strip() in {"E", "E★", "E*"}]
            if _req_e_r > 0 and nurses_e:
                ae = sum(1 for i in nurses_e if cache_group_radar.get(i) == "A")
                be = sum(1 for i in nurses_e if cache_group_radar.get(i) == "B")
                if ae < 2:
                    shortages.append(f"🚨 {day_str}號：E 班 A組僅 {ae} 人（需≥2）")
                if be < 2 and (ae + be) < 4:  # A組已填補B組空缺時不警示
                    shortages.append(f"🚨 {day_str}號：E 班 B組僅 {be} 人（需≥2，A組補位後合計 {ae+be} 人）")

            # 週六白班：A組至少3人
            if d in sat_set_r:
                nurses_d_sat = [i for i in ai_df_local.index
                                if isinstance(sched_df.at[i, day_str], str)
                                and sched_df.at[i, day_str].startswith("D")]
                ad_sat = sum(1 for i in nurses_d_sat if cache_group_radar.get(i) == "A")
                if ad_sat < 3:
                    shortages.append(f"⚠️ {day_str}號（週六）：D 班 A組僅 {ad_sat} 人（需≥3）")

            # 週日白班：A組至少2人、B組至少2人
            if d in sun_set_r:
                nurses_d_sun = [i for i in ai_df_local.index
                                if isinstance(sched_df.at[i, day_str], str)
                                and sched_df.at[i, day_str].startswith("D")]
                ad_sun = sum(1 for i in nurses_d_sun if cache_group_radar.get(i) == "A")
                bd_sun = sum(1 for i in nurses_d_sun if cache_group_radar.get(i) == "B")
                if ad_sun < 2:
                    shortages.append(f"⚠️ {day_str}號（週日）：D 班 A組僅 {ad_sun} 人（需≥2）")
                if bd_sun < 2 and (ad_sun + bd_sun) < 4:  # A組已填補B組空缺時不警示
                    shortages.append(f"⚠️ {day_str}號（週日）：D 班 B組僅 {bd_sun} 人（需≥2，A組補位後合計 {ad_sun+bd_sun} 人）")
    except Exception:
        pass

    if shortages:
        with st.expander(f"🚨 結算警示：共 {len(shortages)} 項（點擊展開）", expanded=False):
            for msg in shortages:
                if "🚨" in msg: st.error(msg)
                else: st.warning(msg)
        return shortages
    else:
        st.success("🎉 完美！目前班別的總人數、流動比例、及控台配置皆已完全達標，無任何安全死角！")
        return []

def make_sched_col_config(month_days):
    """建立排班表 data_editor 的欄位設定：姓名欄鎖定，日期欄可自由輸入班別代碼"""
    cfg = {"姓名": st.column_config.TextColumn("姓名", disabled=True, width="medium")}
    for d in range(1, month_days + 1):
        cfg[str(d)] = st.column_config.TextColumn(
            str(d), max_chars=6, width="small"
        )
    return cfg

# ── 側欄：存檔 / 載入進度 ────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 💾 存檔 / 載入進度")

    # ── 存檔區 ──
    _cur_step = st.session_state.get("step", 1)
    if _cur_step > 1:
        _yr  = st.session_state.get("sel_year",  "")
        _mo  = st.session_state.get("sel_month", "")
        _fname = f"排班進度_{_yr}_{str(_mo).zfill(2)}.pkl" if (_yr and _mo) else "排班進度存檔.pkl"
        st.download_button(
            label=f"💾 下載目前進度（第 {_cur_step} 步）",
            data=_make_checkpoint(),
            file_name=_fname,
            mime="application/octet-stream",
            use_container_width=True,
            help="將目前排班進度下載為 .pkl 檔案，稍後可重新載入繼續作業",
        )
    else:
        st.info("完成第一步後即可存檔", icon="ℹ️")

    st.divider()

    # ── 載入區 ──
    st.markdown("**📂 載入已存進度**")
    _uploaded_ckpt = st.file_uploader(
        "選擇 .pkl 進度檔",
        type=["pkl"],
        key="checkpoint_uploader",
        label_visibility="collapsed",
        help="上傳之前下載的 .pkl 存檔，還原所有排班步驟與資料",
    )
    if _uploaded_ckpt is not None:
        if st.button("✅ 套用進度", use_container_width=True, type="primary", key="btn_restore_ckpt"):
            _restore_checkpoint(_uploaded_ckpt.read())
            st.rerun()

    st.divider()
    st.caption("💡 建議每完成一個步驟就存一次，避免重新整理後資料遺失。")

    st.divider()
    st.markdown("## 🔍 班表驗證站")
    st.caption("上傳手動微調後的班表 Excel，秒速檢查勞基法與人力安全")
    _val_file = st.file_uploader(
        "上傳微調後的班表 Excel", type=["xlsx"], key="validator_upload"
    )
    if _val_file:
        try:
            _ai = st.session_state.get("ai_df")          # None → 獨立驗證模式
            _quota = st.session_state.get("edited_quota_df")
            if _quota is None:
                _quota = st.session_state.get("quota_df")
            _md = st.session_state.get("month_days")     # None → 從班表推算
            # 讀取上傳的班表（取第一個工作表）
            _val_xl_raw = pd.read_excel(_val_file, sheet_name=0, header=None)
            # ── 自動偵測 header 格式：支援系統輸出格式 及 外部班表格式 ──
            # 掃描前5行，找包含「姓名」的 header 行
            _hdr_row_idx = None
            for _ri in range(min(5, len(_val_xl_raw))):
                _row_vals = [str(v).strip() for v in _val_xl_raw.iloc[_ri].values]
                if "姓名" in _row_vals:
                    _hdr_row_idx = _ri
                    break
            if _hdr_row_idx is None:
                st.error("找不到「姓名」欄，請確認班表格式正確")
            else:
                # 以找到的行為 header 重新讀取
                _hdr = [str(v).strip() for v in _val_xl_raw.iloc[_hdr_row_idx].values]
                _name_col_idx_v = _hdr.index("姓名")
                # 找所有純數字欄（日期）
                _date_cols_raw = []
                for _ci, _cv in enumerate(_hdr):
                    try:
                        _date_cols_raw.append((_ci, int(float(_cv))))
                    except:
                        pass
                # 偵測跨月分界：日期從大跳到1
                _dates_seq = [d for _, d in _date_cols_raw]
                _split = None
                for _si in range(1, len(_dates_seq)):
                    if _dates_seq[_si] < _dates_seq[_si - 1] and _dates_seq[_si] <= 5:
                        _split = _si
                        break
                # 本月原始欄位
                _raw_this_month = _date_cols_raw[_split:] if _split else _date_cols_raw
                # 若未載入月份天數，從班表欄位推算
                if _md is None:
                    _raw_days_tmp = [d for _, d in _raw_this_month]
                    _md = max(_raw_days_tmp) if _raw_days_tmp else 31
                # 本月欄位映射：col_index → day
                _this_month_cols = [(ci, d) for ci, d in _raw_this_month if 1 <= d <= _md]
                _col_to_day = {ci: d for ci, d in _this_month_cols}
                # ── 若未載入人員資料，從班表自動萃取姓名與職稱（獨立驗證模式）──
                if _ai is None:
                    _title_ci_sa = _hdr.index("職稱") if "職稱" in _hdr else None
                    _sa_names, _sa_titles = [], {}
                    for _ri_sa in range(_hdr_row_idx + 1, len(_val_xl_raw)):
                        _vn_sa = str(_val_xl_raw.iloc[_ri_sa, _name_col_idx_v]).strip()
                        if not _vn_sa or _vn_sa in ("nan","NaN","None","姓名"): continue
                        try: float(_vn_sa); continue
                        except: pass
                        if _vn_sa not in _sa_names:
                            _sa_names.append(_vn_sa)
                            if _title_ci_sa is not None:
                                _sa_titles[_vn_sa] = str(_val_xl_raw.iloc[_ri_sa, _title_ci_sa]).strip()
                    _ai = pd.DataFrame({
                        "姓名": _sa_names,
                        "職稱": [_sa_titles.get(n, "") for n in _sa_names],
                    }).reset_index(drop=True)
                    st.info(f"ℹ️ 獨立驗證模式：從班表自動偵測到 {len(_sa_names)} 位人員（共 {_md} 天）。未載入配額資料時，配額相關檢查將略過。")
                # 資料行：姓名非空、非統計行
                _ai_names = _ai["姓名"].str.strip().tolist()
                _val_aligned = _ai[["姓名"]].copy().reset_index(drop=True)
                for _dv in range(1, _md + 1):
                    _val_aligned[str(_dv)] = ""
                for _ri in range(_hdr_row_idx + 1, len(_val_xl_raw)):
                    _vname = str(_val_xl_raw.iloc[_ri, _name_col_idx_v]).strip()
                    if not _vname or _vname in ("nan", "NaN", "None", "姓名"):
                        continue
                    try:
                        float(_vname)
                        continue  # 跳過純數字（統計列）
                    except:
                        pass
                    if _vname not in _ai_names:
                        continue
                    _ridx = _ai_names.index(_vname)
                    for _ci, _d in _col_to_day.items():
                        _raw_v = str(_val_xl_raw.iloc[_ri, _ci]).strip()
                        _val_aligned.at[_ridx, str(_d)] = "" if _raw_v in ("nan", "NaN", "", "V") else _raw_v
                st.success("班表解析成功，開始驗證...")
                # 安全警示雷達
                with st.expander("🚨 安全警示雷達", expanded=True):
                    if _quota is not None:
                        display_safety_radar(_val_aligned, _quota, _ai)
                    else:
                        st.info("ℹ️ 未載入配額資料（quota），安全警示雷達需要配額資訊，本次略過此項檢查。")
                # 勞基法四週變形工時審查
                with st.expander("⚖️ 勞基法審查", expanded=True):
                    _pw_df, _viol_df = build_four_week_review(
                        _val_aligned, _ai, _md,
                        prev_buffer=st.session_state.get("prev_month_buffer")
                    )
                    if "🚨" in " ".join(_viol_df["合法判斷"].tolist() if "合法判斷" in _viol_df.columns else []):
                        st.error("發現勞基法違規")
                    st.dataframe(_viol_df, use_container_width=True)
                # ── ⚖️ 公平性結算 ──────────────────────────────────────────
                with st.expander("⚖️ 公平性結算", expanded=False):
                    _vhol_set = set(st.session_state.get("holiday_list", []))
                    _illegal_next = {"D": ["N"], "E": ["D", "N", "12-8"], "12-8": ["N"], "N": []}
                    _v_stats = []
                    for _vi, _vr in _ai.iterrows():
                        _vname2 = str(_vr["姓名"]).strip()
                        _vtitle = str(_vr.get("職稱", "")).strip()
                        _vis_no_hol = _vtitle in NO_HOL_SET
                        _vrow_al = _val_aligned[_val_aligned["姓名"] == _vname2]
                        if _vrow_al.empty: continue
                        _vs = {d: str(_vrow_al.iloc[0].get(str(d), "")).strip() for d in range(1, _md + 1)}
                        _v_worked = sum(1 for v in _vs.values() if is_work(v))
                        _v_hol = 0 if _vis_no_hol else sum(1 for d, v in _vs.items() if d in _vhol_set and is_work(v))
                        _v_night = sum(1 for v in _vs.values() if v in ("E", "N", "12-8") or v in ("E★", "N★"))
                        # 目標天數
                        _vpt = st.session_state.get("personal_targets", {}).get(_vi, _md - 10)
                        _vstatus = "✅ 達標" if _v_worked == _vpt else (f"⚠️ 欠班 {_vpt - _v_worked} 天" if _v_worked < _vpt else f"🟢 超班 {_v_worked - _vpt} 天")
                        _v_stats.append({
                            "姓名": _vname2, "職稱": _vtitle,
                            "目標": _vpt, "實際": _v_worked, "狀態": _vstatus,
                            "假日出勤": "-" if _vis_no_hol else _v_hol,
                            "夜班數": "-" if _vis_no_hol else _v_night,
                        })
                    if _v_stats:
                        st.dataframe(pd.DataFrame(_v_stats), use_container_width=True, hide_index=True)
                # ── 🌸 假日出勤分布 ────────────────────────────────────────
                with st.expander("🌸 假日出勤分布", expanded=False):
                    _vhol_set2 = set(st.session_state.get("holiday_list", []))
                    _hol_dist = []
                    for _vi, _vr in _ai.iterrows():
                        _vname3 = str(_vr["姓名"]).strip()
                        _vtitle3 = str(_vr.get("職稱", "")).strip()
                        if _vtitle3 in NO_HOL_ADMIN: continue
                        _vrow_al3 = _val_aligned[_val_aligned["姓名"] == _vname3]
                        if _vrow_al3.empty: continue
                        _vs3 = {d: str(_vrow_al3.iloc[0].get(str(d), "")).strip() for d in range(1, _md + 1)}
                        _vh3 = sum(1 for d, v in _vs3.items() if d in _vhol_set2 and is_work(v))
                        _hol_dist.append({"姓名": _vname3, "假日出勤": _vh3})
                    if _hol_dist:
                        _hol_df3 = pd.DataFrame(_hol_dist).sort_values("假日出勤", ascending=False).reset_index(drop=True)
                        _h_max3 = _hol_df3["假日出勤"].max()
                        _h_min3 = _hol_df3["假日出勤"].min()
                        _h_diff3 = _h_max3 - _h_min3
                        _msg3 = f"最多 {_h_max3} 天 / 最少 {_h_min3} 天 / 差距 {_h_diff3} 天"
                        if _h_diff3 <= 1:
                            st.success(f"✅ 假日出勤均衡：{_msg3}")
                        elif _h_diff3 <= 3:
                            st.warning(f"⚠️ 假日出勤略不均：{_msg3}")
                        else:
                            st.error(f"🚨 假日出勤差距過大：{_msg3}")
                        st.dataframe(_hol_df3, use_container_width=True, hide_index=True)
                # ── 🌙 week_variety 違規掃描 ──────────────────────────────
                with st.expander("🌙 週多樣性（花花班）掃描", expanded=False):
                    _first_wd_v = st.session_state.get("first_wday", 0)
                    _wv_violations = []
                    for _vi, _vr in _ai.iterrows():
                        _vname4 = str(_vr["姓名"]).strip()
                        _vrow_al4 = _val_aligned[_val_aligned["姓名"] == _vname4]
                        if _vrow_al4.empty: continue
                        _vs4 = {d: str(_vrow_al4.iloc[0].get(str(d), "")).strip() for d in range(1, _md + 1)}
                        _checked4 = set()
                        for _d4 in range(1, _md + 1):
                            _wd4 = (_first_wd_v + _d4 - 1) % 7
                            _dsun4 = (_wd4 + 1) % 7
                            _ws4 = max(1, _d4 - _dsun4)
                            if _ws4 in _checked4: continue
                            _checked4.add(_ws4)
                            _we4 = min(_md, _ws4 + 6)
                            _wtypes4 = set()
                            for _dd4 in range(_ws4, _we4 + 1):
                                _v4 = _vs4.get(_dd4, "")
                                _vu4 = _v4.upper()
                                if _vu4 in ("D","公","DX") or (_vu4.startswith("D") and _vu4 != ""): _wtypes4.add("D")
                                elif _vu4 in ("E","E★","E*"): _wtypes4.add("E")
                                elif _vu4 in ("N","N★","N*"): _wtypes4.add("N")
                                elif _vu4 == "12-8": _wtypes4.add("12-8")
                            if len(_wtypes4) >= 3:
                                _week_vals4 = [_vs4.get(_dd4, "休") for _dd4 in range(_ws4, _we4 + 1)]
                                _wv_violations.append({
                                    "姓名": _vname4,
                                    "週次": f"第{_ws4}~{_we4}日",
                                    "班別種類": " / ".join(sorted(_wtypes4)),
                                    "班表": " ".join(_week_vals4),
                                })
                    if _wv_violations:
                        st.error(f"🚨 共 {len(_wv_violations)} 人次出現一週三種班別（違反院內 week_variety 規定）")
                        st.dataframe(pd.DataFrame(_wv_violations), use_container_width=True, hide_index=True)
                    else:
                        st.success("✅ 全體人員週多樣性符合規定（每週最多 2 種班別）")
                # ── 📋 欠班清單 ────────────────────────────────────────────
                with st.expander("📋 欠班清單", expanded=False):
                    _deficit_list5 = []
                    _vpts5 = st.session_state.get("personal_targets", {})
                    for _vi5, _vr5 in _ai.iterrows():
                        _vname5 = str(_vr5["姓名"]).strip()
                        _vrow5 = _val_aligned[_val_aligned["姓名"] == _vname5]
                        if _vrow5.empty: continue
                        _vs5 = {d: str(_vrow5.iloc[0].get(str(d), "")).strip() for d in range(1, _md + 1)}
                        _vw5 = sum(1 for v in _vs5.values() if is_work(v))
                        _vt5 = _vpts5.get(_vi5, _md - 10)
                        if _vw5 < _vt5:
                            _deficit_list5.append({"姓名": _vname5, "目標": _vt5, "實際": _vw5, "欠班天數": _vt5 - _vw5})
                    if _deficit_list5:
                        st.warning(f"⚠️ 共 {len(_deficit_list5)} 人欠班")
                        st.dataframe(pd.DataFrame(_deficit_list5).sort_values("欠班天數", ascending=False), use_container_width=True, hide_index=True)
                    else:
                        st.success("✅ 所有人員均達到目標出勤天數")
                # ── 🔗 §34 班別接續違規掃描 ───────────────────────────────
                with st.expander("🔗 §34 班別接續違規（換班間距）", expanded=False):
                    _illegal_next6 = {"D": ["N"], "E": ["D", "N", "12-8"], "12-8": ["N"], "N": []}
                    _seq_viols6 = []
                    for _vi6, _vr6 in _ai.iterrows():
                        _vname6 = str(_vr6["姓名"]).strip()
                        _vrow6 = _val_aligned[_val_aligned["姓名"] == _vname6]
                        if _vrow6.empty: continue
                        _vs6 = {d: str(_vrow6.iloc[0].get(str(d), "")).strip() for d in range(1, _md + 1)}
                        for _d6 in range(1, _md):
                            _v6a = _vs6.get(_d6, "")
                            _v6b = _vs6.get(_d6 + 1, "")
                            # 正規化
                            def _norm6(v):
                                vu = v.upper()
                                if vu in ("D","公","DX") or (vu.startswith("D") and vu != ""): return "D"
                                if vu in ("E","E★","E*"): return "E"
                                if vu in ("N","N★","N*"): return "N"
                                if vu == "12-8": return "12-8"
                                return None
                            _t6a = _norm6(_v6a)
                            _t6b = _norm6(_v6b)
                            if _t6a and _t6b and _t6b in _illegal_next6.get(_t6a, []):
                                _seq_viols6.append({
                                    "姓名": _vname6,
                                    "日期": f"第{_d6}→{_d6+1}日",
                                    "班別接續": f"{_v6a} → {_v6b}",
                                    "違規原因": f"{_t6a}→{_t6b}（§34 換班間距不足 11h）",
                                })
                    if _seq_viols6:
                        st.error(f"🚨 共 {len(_seq_viols6)} 處班別接續違規")
                        st.dataframe(pd.DataFrame(_seq_viols6), use_container_width=True, hide_index=True)
                    else:
                        st.success("✅ 全月無班別接續違規（§34 換班間距符合規定）")
                # ── 📥 下載驗證報告 ────────────────────────────────────────
                st.divider()
                st.markdown("#### 📥 下載驗證報告")
                try:
                    import io as _io_val
                    _val_buf = _io_val.BytesIO()
                    with pd.ExcelWriter(_val_buf, engine="openpyxl") as _val_wr:
                        # 工作表1：公平性結算
                        if _v_stats:
                            pd.DataFrame(_v_stats).to_excel(_val_wr, sheet_name="公平性結算", index=False)
                        # 工作表2：假日出勤分布
                        if _hol_dist:
                            _hol_df3.to_excel(_val_wr, sheet_name="假日出勤分布", index=False)
                        # 工作表3：週多樣性違規
                        _wv_df_out = pd.DataFrame(_wv_violations) if _wv_violations else pd.DataFrame({"結果": ["✅ 無違規"]})
                        _wv_df_out.to_excel(_val_wr, sheet_name="週多樣性掃描", index=False)
                        # 工作表4：欠班清單
                        _def_df_out = pd.DataFrame(_deficit_list5) if _deficit_list5 else pd.DataFrame({"結果": ["✅ 全員達標"]})
                        _def_df_out.to_excel(_val_wr, sheet_name="欠班清單", index=False)
                        # 工作表5：班別接續違規
                        _seq_df_out = pd.DataFrame(_seq_viols6) if _seq_viols6 else pd.DataFrame({"結果": ["✅ 無違規"]})
                        _seq_df_out.to_excel(_val_wr, sheet_name="班別接續違規", index=False)
                        # 工作表6：原始對齊班表
                        _val_aligned.to_excel(_val_wr, sheet_name="對齊班表", index=False)
                    _val_buf.seek(0)
                    st.download_button(
                        label="⬇️ 下載驗證報告 Excel",
                        data=_val_buf.getvalue(),
                        file_name="驗證報告.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_val_report",
                    )
                except Exception as _e_dl:
                    st.warning(f"報告產生失敗：{_e_dl}")
        except Exception as _e:
            st.error(f"解析失敗：{_e}，請確認上傳的班表格式與系統輸出的 Excel 格式一致")

    st.divider()
    st.markdown("## 📊 加班線分配 + 變形工時審查")
    st.caption("上傳已排好班別的班表，系統自動分配白班加班線並驗證四周變形工時，不更動任何班別。")
    _ot_file = st.file_uploader("上傳班表 Excel（D/E/N/12-8 已排好）", type=["xlsx"], key="ot_upload")
    if _ot_file:
        try:
            _ai_ot = st.session_state.get("ai_df")        # None → 獨立模式
            _md_ot = st.session_state.get("month_days")   # None → 從班表推算
            _hol_ot = set(st.session_state.get("holiday_list", []))
            _sat_ot = st.session_state.get("saturdays_list", [])
            _sun_ot = st.session_state.get("sundays_list", [])
            _nat_ot = st.session_state.get("nat_holidays_list", [])
            # 讀取上傳班表（支援系統輸出格式 及 外部多列標題格式）
            _ot_xl_raw = pd.read_excel(_ot_file, sheet_name=0, header=None)
            # 自動偵測 header 行（找含「姓名」的第一行）
            _ot_hdr_idx = None
            for _ri in range(min(5, len(_ot_xl_raw))):
                _rv = [str(v).strip() for v in _ot_xl_raw.iloc[_ri].values]
                if "姓名" in _rv:
                    _ot_hdr_idx = _ri; break
            if _ot_hdr_idx is None:
                st.error("找不到「姓名」欄，請確認格式正確")
                st.stop()
            _ot_hdr = [str(v).strip() for v in _ot_xl_raw.iloc[_ot_hdr_idx].values]
            _ot_name_ci = _ot_hdr.index("姓名")
            # 解析日期欄（支援跨月格式：前月尾+本月）
            _ot_date_cols = []
            for _ci, _cv in enumerate(_ot_hdr):
                try: _ot_date_cols.append((_ci, int(float(_cv))))
                except: pass
            _ot_dates = [d for _, d in _ot_date_cols]
            _ot_split = next((i for i in range(1, len(_ot_dates)) if _ot_dates[i] < _ot_dates[i-1] and _ot_dates[i] <= 5), None)
            _ot_raw_this = _ot_date_cols[_ot_split:] if _ot_split else _ot_date_cols
            # 若未載入月份天數，從班表欄位推算
            if _md_ot is None:
                _ot_raw_days_tmp = [d for _, d in _ot_raw_this]
                _md_ot = max(_ot_raw_days_tmp) if _ot_raw_days_tmp else 31
            _ot_this_month = [(ci, d) for ci, d in _ot_raw_this if 1 <= d <= _md_ot]
            _ot_ci_to_day = {ci: d for ci, d in _ot_this_month}
            # 若未載入人員資料，從班表自動萃取姓名與職稱（獨立模式）
            if _ai_ot is None:
                _ot_title_ci = _ot_hdr.index("職稱") if "職稱" in _ot_hdr else None
                _ot_sa_names, _ot_sa_titles = [], {}
                for _ri_ot in range(_ot_hdr_idx + 1, len(_ot_xl_raw)):
                    _vn_ot = str(_ot_xl_raw.iloc[_ri_ot, _ot_name_ci]).strip()
                    if not _vn_ot or _vn_ot in ("nan","NaN","None","姓名"): continue
                    try: float(_vn_ot); continue
                    except: pass
                    if _vn_ot not in _ot_sa_names:
                        _ot_sa_names.append(_vn_ot)
                        if _ot_title_ci is not None:
                            _ot_sa_titles[_vn_ot] = str(_ot_xl_raw.iloc[_ri_ot, _ot_title_ci]).strip()
                _ai_ot = pd.DataFrame({
                    "姓名": _ot_sa_names,
                    "職稱": [_ot_sa_titles.get(n, "") for n in _ot_sa_names],
                }).reset_index(drop=True)
                st.info(f"ℹ️ 獨立分析模式：從班表自動偵測到 {len(_ot_sa_names)} 位人員（共 {_md_ot} 天）。假日/週末資訊未載入，加班線分配結果僅供參考。")
            # 對齊姓名索引
            _day_cols_ot = [str(d) for d in range(1, _md_ot + 1)]
            _ai_names_ot = _ai_ot["姓名"].str.strip().tolist()
            _ot_aligned = _ai_ot[["姓名"]].copy().reset_index(drop=True)
            for _dc in _day_cols_ot:
                _ot_aligned[_dc] = ""
            for _ri in range(_ot_hdr_idx + 1, len(_ot_xl_raw)):
                _vname = str(_ot_xl_raw.iloc[_ri, _ot_name_ci]).strip()
                if not _vname or _vname in ("nan", "NaN", "None", "姓名"): continue
                try: float(_vname); continue
                except: pass
                if _vname not in _ai_names_ot: continue
                _ridx = _ai_names_ot.index(_vname)
                for _ci, _d in _ot_ci_to_day.items():
                    _v = str(_ot_xl_raw.iloc[_ri, _ci]).strip()
                    _ot_aligned.at[_ridx, str(_d)] = "" if _v in ("nan", "NaN", "", "V") else _v
            # 建立 sched dict（index 為 ai_df index）
            _sched_ot = {}
            for _i in _ai_ot.index:
                _sched_ot[_i] = [""] + [str(_ot_aligned.at[_i, _dc]).strip() for _dc in _day_cols_ot]
            # 快取
            _cache_pref_ot  = {i: str(row.get("包班意願","")).strip() for i, row in _ai_ot.iterrows()}
            _cache_title_ot = {i: str(row.get("職稱","")).strip() for i, row in _ai_ot.iterrows()}
            _class_map_ot   = {i: [s.strip() for s in str(row.get("上課日期","")).split(",") if s.strip().isdigit()] for i, row in _ai_ot.iterrows()}
            # ── 加班線均分（完整複製 Step 6 邏輯）────────────────────
            _ot_days_count = {i: 0 for i in _ai_ot.index}
            _ot_history    = {i: [] for i in _ai_ot.index}
            for _d_ot in range(1, _md_ot + 1):
                if _d_ot in _hol_ot: continue
                _dw_ot = [i for i in _ai_ot.index if _sched_ot[i][_d_ot] == "D"]
                if not _dw_ot: continue
                _elig_ot = []
                for _i in _dw_ot:
                    if _cache_pref_ot[_i] != "": continue
                    if _cache_title_ot[_i] in NO_HOL_ADMIN: continue
                    _y_ot = _sched_ot[_i][_d_ot - 1] if _d_ot > 1 else ""
                    if _y_ot == "12-8": continue
                    _elig_ot.append(_i)
                _elig_ot.sort(key=lambda x: (_ot_days_count[x], random.random()))
                _num_ot = min(14, len(_elig_ot))
                _sel_ot = _elig_ot[:_num_ot]
                _heavy_ot = {i for i in _sel_ot if (_sched_ot[i][_d_ot - 1] if _d_ot > 1 else "") in ["D1","D2","D3"]}
                def _avg_ot(i):
                    return sum(_ot_history[i]) / len(_ot_history[i]) if _ot_history[i] else 15.0
                _leaders_ot = [x for x in _sel_ot if _cache_title_ot[x] == "組長"]
                _class_ot   = [x for x in _sel_ot if x not in _leaders_ot and str(_d_ot) in _class_map_ot.get(x, [])]
                _regs_ot    = [x for x in _sel_ot if x not in _leaders_ot and x not in _class_ot]
                _leaders_ot.sort(key=_avg_ot, reverse=True)
                _regs_ot.sort(key=_avg_ot, reverse=True)
                _slots_ot = list(range(1, _num_ot + 1))
                _assign_ot = {}
                for _p in _leaders_ot:
                    _vs = [s for s in _slots_ot if s >= 6]
                    _ch = min(_vs) if _vs else (max(_slots_ot) if _slots_ot else None)
                    if _ch: _slots_ot.remove(_ch); _assign_ot[_p] = _ch
                for _p in _class_ot:
                    _vs = [s for s in _slots_ot if s >= 8]
                    if _vs:
                        _ch = min(_vs); _slots_ot.remove(_ch); _assign_ot[_p] = _ch
                for _p in _regs_ot:
                    if not _slots_ot: break
                    _valid = [s for s in _slots_ot if s >= 4] if _p in _heavy_ot else _slots_ot
                    _ch = min(_valid) if _valid else min(_slots_ot)
                    _slots_ot.remove(_ch); _assign_ot[_p] = _ch
                for _i in _dw_ot:
                    if _cache_pref_ot[_i] != "": continue
                    if _cache_title_ot[_i] in NO_HOL_ADMIN: continue
                    if _i in _assign_ot:
                        _sched_ot[_i][_d_ot] = f"D{_assign_ot[_i]}"
                        _ot_days_count[_i] += 1
                        _ot_history[_i].append(_assign_ot[_i])
                    else:
                        _ot_history[_i].append(15)
            # OT 天數均等後處理
            _elig_ot_set = {i for i in _ai_ot.index if _cache_pref_ot[i] == "" and _cache_title_ot[i] not in NO_HOL_ADMIN}
            for _ in range(500):
                _cts = {i: _ot_days_count[i] for i in _elig_ot_set}
                if not _cts or max(_cts.values()) - min(_cts.values()) <= 1: break
                _ov_l = [i for i, c in _cts.items() if c == max(_cts.values())]
                _un_l = [i for i, c in _cts.items() if c == min(_cts.values())]
                _sw = False
                for _ov in _ov_l:
                    if _sw: break
                    for _un in _un_l:
                        if _sw: break
                        for _d in range(1, _md_ot + 1):
                            if _d in _hol_ot: continue
                            _vo = _sched_ot[_ov][_d]; _vu = _sched_ot[_un][_d]
                            if not (isinstance(_vo, str) and _vo.startswith("D") and len(_vo) > 1): continue
                            if _vu != "D": continue
                            if (_sched_ot[_un][_d - 1] if _d > 1 else "") == "12-8": continue
                            _ln = int(_vo[1:])
                            _sched_ot[_un][_d] = _vo; _sched_ot[_ov][_d] = "D"
                            _ot_days_count[_un] += 1; _ot_days_count[_ov] -= 1
                            if _ln in _ot_history[_ov]: _ot_history[_ov].remove(_ln)
                            _ot_history[_un].append(_ln)
                            _sw = True; break
                if not _sw: break
            # 產出結果 DataFrame
            _result_ot = pd.DataFrame({"姓名": _ai_ot["姓名"]})
            for _d in range(1, _md_ot + 1):
                _result_ot[str(_d)] = [_sched_ot[i][_d] for i in _ai_ot.index]
            st.success("加班線分配完成！")
            # 顯示班表
            with st.expander("📋 含加班線班表預覽", expanded=True):
                _dcols_ot = [str(d) for d in range(1, _md_ot + 1)]
                _disp_ot = _result_ot.copy()
                for _c in _dcols_ot:
                    _disp_ot[_c] = _disp_ot[_c].apply(abbrev_display)
                st.dataframe(_disp_ot.style.map(color_shifts, subset=_dcols_ot), use_container_width=True)
            # OT 統計
            with st.expander("📊 加班天數統計", expanded=False):
                _ot_stat = [{"姓名": _ai_ot.at[i,"姓名"], "加班天數": _ot_days_count[i], "平均線位": round(sum(_ot_history[i])/len(_ot_history[i]),1) if _ot_history[i] else "-"}
                            for i in _ai_ot.index if _cache_pref_ot[i] == "" and _cache_title_ot[i] not in NO_HOL_ADMIN]
                st.dataframe(pd.DataFrame(_ot_stat).sort_values("加班天數"), use_container_width=True, hide_index=True)
            # 四周變形工時審查
            with st.expander("⚖️ 勞基法四周變形工時審查", expanded=True):
                _pw_ot, _viol_ot = build_four_week_review(_result_ot, _ai_ot, _md_ot)
                st.dataframe(_pw_ot, use_container_width=True, hide_index=True)
                _n_viol = len([r for _, r in _pw_ot.iterrows() if "🚨" in str(r.get("合法判斷",""))])
                if _n_viol == 0:
                    st.success("✅ 全體人員均符合四周變形工時規定，無違規！")
                else:
                    st.error(f"🚨 共 {_n_viol} 位人員有違規")
                    st.dataframe(_viol_ot, use_container_width=True, hide_index=True)
            # 下載含加班線的班表
            import io as _io_ot
            _dl_ot = _io_ot.BytesIO()
            with pd.ExcelWriter(_dl_ot, engine="openpyxl") as _wr_ot:
                _result_ot.to_excel(_wr_ot, sheet_name="含加班線班表", index=False)
                _pw_ot.to_excel(_wr_ot, sheet_name="變形工時審查", index=False)
                _viol_ot.to_excel(_wr_ot, sheet_name="違規明細", index=False)
            st.download_button("📥 下載含加班線班表", data=_dl_ot.getvalue(),
                               file_name="Schedule_With_OT.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="dl_ot")
        except Exception as _e_ot:
            st.error(f"處理失敗：{_e_ot}，請確認班表格式（第一欄為姓名，之後為1~{st.session_state.get('month_days',31)}日）")

st.title("🏥 層級式護理輔助排班工作站")
st.progress(min(st.session_state.step / 7, 1.0), text=f"目前進度：第 {st.session_state.step} 步 / 共 7 步")

weekday_names_list = ["週一", "週二", "週三", "週四", "週五", "週六", "週日"]

# ═══════════════════════════════════════════════════════════════════
# 🔒 前置步驟：前月班表上傳（上傳後即鎖定，不可替換；如需重置請重新整理頁面）
# ═══════════════════════════════════════════════════════════════════
st.write("#### 🔗 前月班表（選填）｜用於跨月連五 / 14日窗口 / 8週變形工時審查")
st.caption("上傳上個月的最終班表 Excel，系統會擷取最後 13 天的班別做為本月排班的接軌緩衝，審查時自動偵測跨月違規。若不上傳則僅審查本月內的合法性。上傳成功後將鎖定，如需重新上傳請重新整理頁面。")

_prev_locked = st.session_state.get("prev_month_buffer") is not None

_prev_file = st.file_uploader(
    "上傳前月班表 Excel（第一欄為姓名，之後為各日班別）",
    type=["xlsx"], key="prev_month_uploader",
    disabled=_prev_locked
)

if _prev_locked:
    st.success(f"🔒 前月班表已鎖定（{len(st.session_state.prev_month_buffer)} 人）｜如需重新上傳請重新整理頁面")
elif _prev_file is not None:
    try:
        # 自動偵測標題行：往前5行搜尋含「姓名」的那一行
        _prev_raw = pd.read_excel(_prev_file, sheet_name=0, header=None)
        _header_row = None
        for _ri in range(min(5, len(_prev_raw))):
            if any("姓名" in str(v) for v in _prev_raw.iloc[_ri].tolist()):
                _header_row = _ri
                break
        if _header_row is None:
            st.error("找不到「姓名」欄，請確認前月班表格式（標題行需包含『姓名』）")
            st.stop()
        _prev_xl = pd.read_excel(_prev_file, sheet_name=0, header=_header_row)
        # 統一欄位名稱為字串，找姓名欄
        _prev_xl.columns = [str(c).strip() for c in _prev_xl.columns]
        _pnc = [c for c in _prev_xl.columns if "姓名" in c]
        if _pnc:
            _prev_xl = _prev_xl.rename(columns={_pnc[0]: "姓名"})
            _prev_xl["姓名"] = _prev_xl["姓名"].astype(str).str.strip()
            # 自動偵測前月天數：欄位名稱為純數字，且取最大者推算月份天數
            # 注意：此班表格式可能同時包含上月尾巴和本月日期，只取 1~31 範圍的純數字
            _prev_day_cols = sorted(
                [c for c in _prev_xl.columns if str(c).isdigit() and 1 <= int(c) <= 31],
                key=lambda x: int(x)
            )
            _prev_md = max(int(c) for c in _prev_day_cols) if _prev_day_cols else 0
            if _prev_md >= 13:
                _buf_cols = [str(d) for d in range(_prev_md - 12, _prev_md + 1)]
                _buf = {}
                for _, _pr in _prev_xl.iterrows():
                    _pname = str(_pr["姓名"]).strip()
                    if not _pname or _pname == "nan": continue
                    _buf[_pname] = {}
                    for _offset, _bc in enumerate(_buf_cols):
                        _rel = _offset - 13
                        _v = str(_pr.get(_bc, "")).strip()
                        _buf[_pname][_rel] = "" if _v in ("nan", "NaN", "", "休假", "例假", "國定") else _v
                st.session_state.prev_month_buffer = _buf
                st.success(f"✅ 前月班表解析成功，擷取最後 13 天（第 {_prev_md-12}~{_prev_md} 日）作為跨月緩衝，上傳已鎖定")
                st.rerun()
            else:
                st.warning("前月班表天數不足，無法建立緩衝")
        else:
            st.error("找不到「姓名」欄")
    except Exception as _pe:
        st.error(f"前月班表解析失敗：{_pe}")

st.divider()

# ==========================================
# 📍 第一步：排班月行事曆
# ==========================================
st.header("1️⃣ 第一步：排班月行事曆")

# 自動計算下個月
_today = datetime.date.today()
_next = (_today.replace(day=1) + datetime.timedelta(days=32)).replace(day=1)
_def_year, _def_month = _next.year, _next.month
_year_list = list(range(_today.year - 1, _today.year + 3))
_def_year_idx = _year_list.index(_def_year) if _def_year in _year_list else 2

col1, col2, col3 = st.columns(3)
with col1:
    sel_year = st.selectbox("年份", _year_list, index=_def_year_idx, disabled=st.session_state.step > 1)
with col2:
    sel_month = st.selectbox("月份", list(range(1, 13)), index=_def_month - 1, disabled=st.session_state.step > 1)

temp_month_days = calendar.monthrange(sel_year, sel_month)[1]
temp_first_wday = calendar.monthrange(sel_year, sel_month)[0] 

all_saturdays = [d for d in range(1, temp_month_days + 1) if (temp_first_wday + d - 1) % 7 == 5]
all_sundays = [d for d in range(1, temp_month_days + 1) if (temp_first_wday + d - 1) % 7 == 6]

@st.cache_data(ttl=86400)
def fetch_gov_holidays(y, m):
    prefix = f"{y}{m:02d}"
    off_d = []
    try:
        res = requests.get(f"https://cdn.jsdelivr.net/gh/ruyut/TaiwanCalendar/data/{y}.json", timeout=3)
        if res.status_code == 200:
            for item in res.json():
                if str(item.get("date", "")).startswith(prefix):
                    if item.get("isHoliday") in [True, "是", "true"]:
                        off_d.append(int(str(item.get("date"))[-2:]))
    except Exception: pass
    return off_d

fetched_off_days = fetch_gov_holidays(sel_year, sel_month)
pure_nat_holidays = [d for d in fetched_off_days if d not in all_saturdays and d not in all_sundays]

with col3: 
    temp_target_off = st.number_input("📌 當月常規應休天數", min_value=0, max_value=20, value=len(all_saturdays)+len(all_sundays)+len(pure_nat_holidays), disabled=st.session_state.step > 1)

st.write("#### 🎈 假日日期確認 (將影響配額與排班判定)")
col_sat, col_sun, col_nat = st.columns(3)
with col_sat:
    temp_sats = st.multiselect("週六日期", [str(i) for i in range(1, temp_month_days + 1)], default=[str(d) for d in all_saturdays], disabled=st.session_state.step > 1)
with col_sun:
    temp_suns = st.multiselect("週日日期", [str(i) for i in range(1, temp_month_days + 1)], default=[str(d) for d in all_sundays], disabled=st.session_state.step > 1)
with col_nat:
    temp_nats = st.multiselect("國定假日日期", [str(i) for i in range(1, temp_month_days + 1)], default=[str(d) for d in pure_nat_holidays], disabled=st.session_state.step > 1)

if st.session_state.step == 1:
    if st.button("✅ 確認行事曆，進入下一步 (上傳名單與設定配額)", type="primary"):
        st.session_state.sel_year  = sel_year
        st.session_state.sel_month = sel_month
        st.session_state.month_days = temp_month_days
        st.session_state.first_wday = temp_first_wday
        st.session_state.saturdays_list = [int(x) for x in temp_sats]
        st.session_state.sundays_list = [int(x) for x in temp_suns]
        st.session_state.nat_holidays_list = [int(x) for x in temp_nats]
        st.session_state.holiday_list = list(set(st.session_state.saturdays_list + st.session_state.sundays_list + st.session_state.nat_holidays_list))
        st.session_state.target_off = temp_target_off

        default_quota_data = []
        for d in range(1, temp_month_days + 1):
            day_str = str(d)
            current_weekday = (temp_first_wday + d - 1) % 7
            weekday_name = weekday_names_list[current_weekday]
            is_sun_or_nat = d in st.session_state.sundays_list or d in st.session_state.nat_holidays_list
            is_sat = d in st.session_state.saturdays_list
            if is_sun_or_nat: q = {"日期": day_str, "星期": weekday_name, "D班": 4, "12-8": 0, "E班": 4, "N班": 2}
            elif is_sat: q = {"日期": day_str, "星期": weekday_name, "D班": 7, "12-8": 0, "E班": 4, "N班": 2}
            else: q = {"日期": day_str, "星期": weekday_name, "D班": 30, "12-8": 6, "E班": 4, "N班": 2}
            default_quota_data.append(q)
        st.session_state.default_quota = pd.DataFrame(default_quota_data)
        
        st.session_state.step = 2
        st.rerun()

# ==========================================
# 📍 第二步：帶入人員、動態次專科與配額預警
# ==========================================
if st.session_state.step >= 2:
    st.divider()
    st.header("2️⃣ 第二步：帶入預班表")
    
    if st.session_state.step == 2:
        uploaded_ai = st.file_uploader("📂 上傳護理師參數名單 (Excel 檔)", type=["xlsx", "xls"])
        
        if uploaded_ai is not None:
            ai_df = pd.read_excel(uploaded_ai, header=0).fillna("")
            # 若第一列為分類標頭（如「▌ 基本資料」），則改以第二列為欄位名
            if "姓名" not in ai_df.columns:
                ai_df = pd.read_excel(uploaded_ai, header=1).fillna("")
            if "姓名" not in ai_df.columns:
                st.error("❌ 找不到「姓名」欄位，請確認上傳的 Excel 格式正確（使用官方範本）。")
                st.stop()
            st.session_state.ai_df = ai_df
            
            all_skills = set()
            for idx, row in ai_df.iterrows():
                for s in str(row.get("次專科能力", "")).split(","):
                    if s.strip(): all_skills.add(s.strip())
            skill_cols = sorted(list(all_skills))
            
            st.write("### ⚙️ 平日次專科能力需求設定")
            if not skill_cols:
                st.info("💡 您的 Excel 名單中沒有任何人填寫「次專科能力」，將跳過此設定。")
                temp_skills = []
                edited_weekly_df = pd.DataFrame()
            else:
                st.success(f"✅ 成功從您的 Excel 名單中抓取到以下次專科：{', '.join(skill_cols)}")
                temp_skills = st.multiselect("請選擇本次排班需考量的次專科項目", skill_cols, default=skill_cols)
                
                weekly_template = []
                for wd in weekday_names_list[:5]:   # 僅週一～週五
                    q_week = {"星期": wd}
                    for s in temp_skills:
                        q_week[f"{s}需求"] = 0
                    weekly_template.append(q_week)
                st.info("👉 請在下方表格輸入每個平日各次專科「至少需要幾個人上班」：")
                edited_weekly_df = st.data_editor(pd.DataFrame(weekly_template), hide_index=True, use_container_width=True, disabled=["星期"])
            
            st.write("### ⚙️ 每日基礎人力配額設定 (修改配額將即時更新下方缺額警示)")
            edited_quota_df = st.data_editor(st.session_state.default_quota, hide_index=True, use_container_width=True)
            
            sched = {i: [""] * (st.session_state.month_days + 1) for i in ai_df.index}
            pre_type_map = {}  # (idx, d) -> "預假" | "預長假" | "預白"（預覽標示用，不影響實際排班資料）
            for idx, row in ai_df.iterrows():
                # 預休日期 → O（預覽標示為「預假」）
                for d_str in str(row.get("預休日期", "")).split(","):
                    if d_str.strip().isdigit() and 1 <= int(d_str.strip()) <= st.session_state.month_days:
                        d_int = int(d_str.strip())
                        sched[idx][d_int] = "O"
                        pre_type_map[(idx, d_int)] = "預假"

                # 預約長假日期 → O（預覽標示為「預長假」）
                for d_str in str(row.get("預約長假日期", "")).split(","):
                    if d_str.strip().isdigit() and 1 <= int(d_str.strip()) <= st.session_state.month_days:
                        d_int = int(d_str.strip())
                        sched[idx][d_int] = "O"
                        pre_type_map[(idx, d_int)] = "預長假"

                sp_leave_str = str(row.get("特殊假別", "")).strip()
                if sp_leave_str:
                    for item in sp_leave_str.split(","):
                        if ":" in item or "-" in item:
                            delim = ":" if ":" in item else "-"
                            d_v, l_v = item.split(delim)[0].strip(), item.split(delim)[1].strip()
                            if d_v.isdigit() and 1 <= int(d_v) <= st.session_state.month_days:
                                sched[idx][int(d_v)] = l_v

                # ── 公差日期（優先於預白日期填入）────────────────────────
                gongcha_str = str(row.get("公差日期", "")).strip()
                if gongcha_str:
                    for d_v in [s.strip() for s in gongcha_str.split(",") if s.strip().isdigit()]:
                        if 1 <= int(d_v) <= st.session_state.month_days and sched[idx][int(d_v)] == "":
                            sched[idx][int(d_v)] = "公差"

                rtd = str(row.get("預白日期", "")).strip()
                if rtd:
                    for d_v in [s.strip() for s in rtd.split(",") if s.strip().isdigit()]:
                        if 1 <= int(d_v) <= st.session_state.month_days and sched[idx][int(d_v)] == "":
                            sched[idx][int(d_v)] = "D"
                            pre_type_map[(idx, int(d_v))] = "預白"

                # ── 上課日期：以包班班別填入（大夜→N，小夜→E，其他→D）──────
                # D加班線僅可在第8天以後；初始化一律填基本班別
                class_str = str(row.get("上課日期", "")).strip()
                if class_str:
                    _pref_cls = str(row.get("包班意願", "")).strip()
                    if "大夜" in _pref_cls:
                        _class_shift = "N"
                    else:
                        # 上課日不可排E班或12-8班，小夜包班亦改填D
                        _class_shift = "D"
                    for d_v in [s.strip() for s in class_str.split(",") if s.strip().isdigit()]:
                        d_int_cls = int(d_v)
                        # 預休日期優先於上課日期：格子非空時不覆蓋
                        if 1 <= d_int_cls <= st.session_state.month_days and sched[idx][d_int_cls] == "":
                            sched[idx][d_int_cls] = _class_shift
                            
                mand_holiday_str = str(row.get("國定假日必上班別", "")).strip()
                if mand_holiday_str:
                    for item in mand_holiday_str.split(","):
                        if ":" in item or "-" in item:
                            delim = ":" if ":" in item else "-"
                            d_v, s_v = item.split(delim)[0].strip(), item.split(delim)[1].strip()
                            if d_v.isdigit() and 1 <= int(d_v) <= st.session_state.month_days: 
                                sched[idx][int(d_v)] = s_v.upper()
            
            # 僅在首次（base_sched 尚未建立）時計算底稿，避免覆蓋手動修改
            if st.session_state.base_sched is None:
                base_df = pd.DataFrame({"姓名": ai_df["姓名"]})
                for d in range(1, st.session_state.month_days + 1):
                    base_df[str(d)] = [sched[i][d] for i in ai_df.index]
                st.session_state.base_sched = base_df

            base_df = st.session_state.base_sched

            st.write("#### 🚨 本月預休與人力缺口初步掃描（逐班別估算）")

            # ── 準備假日出勤 & 夜班豁免快取 ──────────────────────
            _scan_can_sat = {idx: str(row.get("能上週六",       "")).strip() == "是" for idx, row in ai_df.iterrows()}
            _scan_can_sun = {idx: str(row.get("能上週日",       "")).strip() == "是" for idx, row in ai_df.iterrows()}
            _scan_can_nat = {idx: str(row.get("能上國定假日",   "")).strip() == "是" for idx, row in ai_df.iterrows()}
            _scan_night_exempt = {idx for idx, row in ai_df.iterrows()
                                  if str(row.get("孕/育嬰免夜班", "")).strip() == "是"}
            _scan_pref = {idx: str(row.get("包班意願", "")).strip() for idx, row in ai_df.iterrows()}
            _scan_sat = st.session_state.saturdays_list
            _scan_sun = st.session_state.sundays_list
            _scan_nat = st.session_state.nat_holidays_list
            _scan_hol_set = set(_scan_sat) | set(_scan_sun) | set(_scan_nat)
            _month_days_s = st.session_state.month_days

            # ── 全月總量檢核（最先顯示）────────────────────────────
            # 計算每人本月可用工作天數（personal_target 近似值）
            _cap_total   = 0   # 全員可用總工作人次
            _cap_en      = 0   # 可排 E/N 的護理師可用人次
            _cap_128     = 0   # 可排 12-8 的護理師可用人次
            for _ci, _cr in ai_df.iterrows():
                _ctitle = str(_cr.get("職稱", "")).strip()
                if _ctitle in NO_HOL_ADMIN:
                    continue   # 護理長/副護理長不計
                _cel = calc_extra_leaves(_cr, _month_days_s, _scan_sat, _scan_sun, _scan_nat,
                                         target_off=st.session_state.target_off)
                _ctgt = max(0, _month_days_s - st.session_state.target_off - _cel)
                _cap_total += _ctgt
                _is_exempt = _ci in _scan_night_exempt
                _is_admin  = _ctitle in ADMIN_TITLES
                _cn = str(_cr.get("夜班資格", "")).strip()          # 空白 = 無夜班資格
                _can_en  = _cn in ("大夜", "小夜")           # 可排 E/N
                _can_128 = _cn in ("大夜", "小夜", "中班")   # 可排 12-8
                if not _is_exempt and not _is_admin:
                    if _can_en:  _cap_en  += _ctgt
                    if _can_128: _cap_128 += _ctgt
                elif _is_exempt and not _is_admin:   # 母性保護：可排 12-8 但不排 E/N
                    _cap_128 += _ctgt

            _dem_total = sum(
                int(edited_quota_df.iloc[d-1]["D班"]) + int(edited_quota_df.iloc[d-1]["E班"])
                + int(edited_quota_df.iloc[d-1]["N班"]) + int(edited_quota_df.iloc[d-1]["12-8"])
                for d in range(1, _month_days_s + 1)
            )
            _dem_en  = sum(int(edited_quota_df.iloc[d-1]["E班"]) + int(edited_quota_df.iloc[d-1]["N班"])
                           for d in range(1, _month_days_s + 1))
            _dem_128 = sum(int(edited_quota_df.iloc[d-1]["12-8"])
                           for d in range(1, _month_days_s + 1))

            _cap_rows = [
                {"檢核項目": "全班別總人次",
                 "全月需求人次": _dem_total, "可用人次上限": _cap_total,
                 "餘裕／缺口": _cap_total - _dem_total},
                {"檢核項目": "E+N 夜班人次（夜班資格人員）",
                 "全月需求人次": _dem_en, "可用人次上限": _cap_en,
                 "餘裕／缺口": _cap_en - _dem_en},
                {"檢核項目": "12-8 中班人次",
                 "全月需求人次": _dem_128, "可用人次上限": _cap_128,
                 "餘裕／缺口": _cap_128 - _dem_128},
            ]
            _cap_df = pd.DataFrame(_cap_rows)
            _has_global_risk = any(r["餘裕／缺口"] < 0 for r in _cap_rows)

            if _has_global_risk:
                st.error("🔴 **全月總量檢核：人力總量不足！** 即使不考慮勞基法限制，全月人次需求已超過全員可工作上限，排班必然出現缺班。")
            else:
                st.warning("🟡 **全月總量檢核（樂觀上限）：** 總量勉強足夠，但勞基法連五/14日窗口/班別相容限制會進一步壓縮可用名額，實際排班仍可能出現缺額。")
            st.caption("⚠️ 可用人次上限 = 未含勞基法連班限制的樂觀估計，實際可排人次通常更低。若餘裕不足 10%，建議調降配額或增加人力。")
            st.dataframe(_cap_df.style.map(
                lambda v: "color:red;font-weight:bold" if isinstance(v, (int, float)) and v < 0 else "",
                subset=["餘裕／缺口"]
            ), use_container_width=False, hide_index=True)
            st.divider()

            with st.expander("📄 點擊展開初始排班底稿（預假/預長假/預白班 特別標示）", expanded=True):
                _edit_base = st.checkbox("🖊️ 開啟手動編輯模式", value=False, key="chk_edit_base_sched")
                if _edit_base:
                    st.caption("💡 直接點擊儲存格輸入班別代碼（D/E/N/12-8/O/上課/公假 等），修改即時儲存")
                    _edited_base = st.data_editor(
                        st.session_state.base_sched,
                        column_config=make_sched_col_config(st.session_state.month_days),
                        use_container_width=True, hide_index=True, key="edit_base_sched"
                    )
                    st.session_state.base_sched = _edited_base
                else:
                    _day_cols_b = [str(i) for i in range(1, st.session_state.month_days + 1)]
                    # 建立預覽用 DataFrame，將 pre_type_map 標記覆蓋於顯示欄位
                    _preview_df = st.session_state.base_sched.copy()
                    for (_p_idx, _p_d), _p_label in pre_type_map.items():
                        if _p_idx in _preview_df.index:
                            _preview_df.at[_p_idx, str(_p_d)] = _p_label
                    st.caption(
                        "🔵 **預**（預假）　"
                        "🟠 **預**（預長假）　"
                        "🟢 **Dx**（預白）　"
                        "🟡 上課　其餘空白＝待排"
                    )
                    for _c in _day_cols_b:
                        _preview_df[_c] = _preview_df[_c].apply(abbrev_display)
                    st.dataframe(
                        build_schedule_with_counts(_preview_df, st.session_state.base_sched, _day_cols_b, st.session_state.ai_df),
                        use_container_width=True
                    )

            display_safety_radar(st.session_state.base_sched, edited_quota_df, st.session_state.ai_df)

            # ── 應上班天數確認表（可手動修改）────────────────────────
            st.write("#### 📋 確認每人本月應上班天數（可直接修改）")
            st.caption(
                "計算公式：當月天數 − 計畫休假天數 − 額外扣減天數。"
                "額外扣減 = 預休(O)+長假超出計畫休假的部分 + 特殊假別（病假/事假/公假/喪假等）。"
                "如有特殊情況（如長期停班、特殊協議），可在「應上班天數」欄直接調整。"
            )
            _targets_data = []
            _scan_hol_set_days = set(_scan_sat) | set(_scan_sun) | set(_scan_nat)
            _scan_weekday_count = sum(1 for d in range(1, st.session_state.month_days + 1)
                                      if d not in _scan_hol_set_days)
            _toff_scan = st.session_state.target_off
            for idx, row in ai_df.iterrows():
                _title_scan = str(row.get("職稱", "")).strip()
                if _title_scan in NO_HOL_SET:
                    # 護理長/副護理長/助理/傷兵/組長：不排假日班，預休/長假/特假均直接扣平日
                    # target_off=0 → 所有平日假別（O+長假+特假）都算扣減
                    _el = calc_extra_leaves(row, st.session_state.month_days,
                                            _scan_sat, _scan_sun, _scan_nat, target_off=0)
                    _calc_t = max(0, _scan_weekday_count - _el)
                else:
                    _el = calc_extra_leaves(row, st.session_state.month_days,
                                            _scan_sat, _scan_sun, _scan_nat, target_off=_toff_scan)
                    _calc_t = max(0, st.session_state.month_days - _toff_scan - _el)

                # ── 理論可達上限：考慮 O + 長假佔用的平日，估算在勞基法限制下最多能上幾天 ──
                # 原理：月份所有平日 - O平日 - 長假平日 = 剩餘可工作平日
                #       加上護理師可上的假日天數（若具備假日出勤能力）
                #       再扣除勞基法估計折損（連五/孤立班，約每5天失去0.5天）
                _o_days_scan   = set()
                _ll_days_scan  = set()
                for _ds in str(row.get("預休日期",      "")).split(","):
                    if _ds.strip().isdigit() and 1 <= int(_ds.strip()) <= st.session_state.month_days:
                        _o_days_scan.add(int(_ds.strip()))
                for _ds in str(row.get("預約長假日期",  "")).split(","):
                    if _ds.strip().isdigit() and 1 <= int(_ds.strip()) <= st.session_state.month_days:
                        _ll_days_scan.add(int(_ds.strip()))
                _blocked_scan  = _o_days_scan | _ll_days_scan | _scan_hol_set_days
                _avail_weekday = sum(1 for d in range(1, st.session_state.month_days + 1)
                                     if d not in _blocked_scan)
                # 加上護理師可上的假日（假日出勤能力允許，且未被O/長假佔用）
                _can_sat_s  = str(row.get("能上週六",    "")).strip() == "是"
                _can_sun_s  = str(row.get("能上週日",    "")).strip() == "是"
                _can_nat_s  = str(row.get("能上國定假日","")).strip() == "是"
                _avail_hol  = sum(
                    1 for d in _scan_hol_set_days
                    if d not in (_o_days_scan | _ll_days_scan) and (
                        (d in set(_scan_sat) and _can_sat_s) or
                        (d in set(_scan_sun) and _can_sun_s) or
                        (d in set(_scan_nat) and _can_nat_s)
                    )
                ) if _title_scan not in NO_HOL_SET else 0
                _raw_avail  = _avail_weekday + _avail_hol
                # 勞基法折損估計：連五（每5天強制休1天）+孤立班約讓可用天數縮減 15%
                # 若月末長假導致最後工作日集中前段，折損可能更高（此為保守下限估計）
                # 傷兵/助理/護理長/副護理長（NO_HOL_ADMIN）：強制填滿所有平日，無勞基法連班限制，折損為 0
                if _title_scan in NO_HOL_ADMIN:
                    _law_deduct = 0
                else:
                    _law_deduct = max(0, round(_raw_avail * 0.15))
                _max_achiev = max(0, _raw_avail - _law_deduct)

                # 若已有手動調整，顯示調整後的值
                _prev = (st.session_state.custom_targets or {}).get(idx, _calc_t)
                _targets_data.append({
                    "姓名":           row["姓名"],
                    "職稱":           _title_scan,
                    "當月天數":       st.session_state.month_days,
                    "計畫休假":       _toff_scan,
                    "額外扣減":       _el,
                    "系統計算值":     _calc_t,
                    "應上班天數":     _prev,
                    "理論可達上限":   _max_achiev,
                })
            _targets_df_raw = pd.DataFrame(_targets_data)

            # ── 應上班天數確認表（與底稿表相同的 checkbox 開關模式）────────────────
            # 每次重繪都把最新計算值寫入 session_state（第一次建立；
            # 之後若使用者未開啟編輯模式，仍保持 custom_targets 的手動值）
            if "targets_df_ss" not in st.session_state:
                st.session_state.targets_df_ss = _targets_df_raw.copy()

            with st.expander("📋 確認每人本月應上班天數", expanded=True):
                _edit_tgt = st.checkbox("🖊️ 開啟手動編輯模式", value=False, key="chk_edit_targets")
                if _edit_tgt:
                    st.caption("💡 直接點擊「應上班天數」欄輸入目標天數，其餘欄位鎖定不可改")
                    _edited_tgt = st.data_editor(
                        st.session_state.targets_df_ss,
                        column_config={
                            "姓名":          st.column_config.TextColumn("姓名",         disabled=True),
                            "職稱":          st.column_config.TextColumn("職稱",         disabled=True),
                            "當月天數":      st.column_config.NumberColumn("當月天數",   disabled=True),
                            "計畫休假":      st.column_config.NumberColumn("計畫休假",   disabled=True),
                            "額外扣減":      st.column_config.NumberColumn("額外扣減",   disabled=True),
                            "系統計算值":    st.column_config.NumberColumn("系統計算值", disabled=True),
                            "理論可達上限":  st.column_config.NumberColumn("理論可達上限 📊", disabled=True),
                            "應上班天數":    st.column_config.NumberColumn("應上班天數",
                                                min_value=0, max_value=st.session_state.month_days, step=1),
                        },
                        hide_index=True, use_container_width=True, key="edit_targets_step2"
                    )
                    # 存回 session_state（下次重繪從這裡讀，不重算）
                    st.session_state.targets_df_ss = _edited_tgt
                    # 同步到 custom_targets
                    _custom_live = {}
                    for _trow in _edited_tgt.itertuples(index=False):
                        _match = ai_df[ai_df["姓名"] == _trow.姓名].index
                        if len(_match) > 0:
                            _custom_live[_match[0]] = max(0, int(_trow.應上班天數))
                    if _custom_live:
                        st.session_state.custom_targets = _custom_live
                else:
                    st.dataframe(st.session_state.targets_df_ss, hide_index=True, use_container_width=True)

            # ── 偏好加班人數設定 ──────────────────────────────────────────────
            st.write("### 📈 人力充足時的偏好加班設定（選填）")
            st.caption("設定每週特定星期與班別，在人力充足時系統會嘗試多補幾人上班。不影響基本配額，僅在已達最低配額後才生效。")

            _wday_labels = ["週一", "週二", "週三", "週四", "週五", "週六", "週日"]
            _shift_labels = ["D班", "E班", "N班", "12-8"]

            # 初始化預設值（全部為0）
            if st.session_state.get("extra_staffing_df") is None:
                _extra_default = []
                for _wl in _wday_labels:
                    _row = {"星期": _wl}
                    for _sl in _shift_labels:
                        _row[_sl] = 0
                    _extra_default.append(_row)
                st.session_state.extra_staffing_df = pd.DataFrame(_extra_default)

            _extra_edited = st.data_editor(
                st.session_state.extra_staffing_df,
                hide_index=True,
                use_container_width=False,
                column_config={
                    "星期": st.column_config.TextColumn("星期", disabled=True, width="small"),
                    "D班":  st.column_config.NumberColumn("D班 希望多加人數", min_value=0, max_value=4, step=1, width="small"),
                    "E班":  st.column_config.NumberColumn("E班 希望多加人數", min_value=0, max_value=4, step=1, width="small"),
                    "N班":  st.column_config.NumberColumn("N班 希望多加人數", min_value=0, max_value=4, step=1, width="small"),
                    "12-8": st.column_config.NumberColumn("12-8 希望多加人數", min_value=0, max_value=4, step=1, width="small"),
                },
                key="extra_staffing_editor"
            )
            st.session_state.extra_staffing_df = _extra_edited

            col_btn1, col_btn2 = st.columns([1, 4])
            with col_btn1:
                if st.button("⬅️ 回到第一步", type="secondary"):
                    st.session_state.step = 1
                    st.session_state.base_sched = None
                    st.session_state.ai_df = None
                    st.rerun()
            with col_btn2:
                if st.button("✅ 確認配額與底稿無誤，進入包班設定", type="primary"):
                    st.session_state.skill_cols = temp_skills
                    st.session_state.edited_weekly_df = edited_weekly_df
                    st.session_state.edited_quota_df = edited_quota_df
                    st.session_state.extra_staffing_df = _extra_edited
                    st.session_state.step = 3
                    st.rerun()
    else:
        with st.expander("已鎖定之設定表", expanded=False):
            st.dataframe(st.session_state.edited_quota_df, hide_index=True, use_container_width=True)

# ==========================================
# 📍 第三步：確認並自動安插包班人員
# ==========================================
if st.session_state.step >= 3:
    st.divider()
    st.header("3️⃣ 第三步：確認並預排包班人員")

    ai_df = st.session_state.ai_df
    month_days = st.session_state.month_days
    edited_quota_df = st.session_state.edited_quota_df

    with st.expander("⚙️ 調整每日人力配額（修改後需重新排班才生效）", expanded=False):
        _q3 = st.data_editor(st.session_state.edited_quota_df, hide_index=True, use_container_width=True, key="quota_editor_step3")
        if st.button("✅ 套用配額變更並重排（從第三步重算）", key="apply_quota_step3"):
            st.session_state.edited_quota_df = _q3
            st.session_state.pack_sched = None
            st.session_state.night_sched = None
            st.session_state.d_sched = None
            st.session_state.final_sched = None
            st.rerun()
    edited_quota_df = st.session_state.edited_quota_df
    
    pack_nurses = []
    for idx, row in ai_df.iterrows():
        pref = str(row.get("包班意願", "")).strip()
        if pref != "":
            pack_nurses.append({"姓名": row["姓名"], "職稱": row["職稱"], "包班意願": pref})
            
    if st.session_state.pack_sched is None:
        if pack_nurses:
            st.write("#### 📋 偵測到以下包班人員，請確認：")
            st.dataframe(pd.DataFrame(pack_nurses), use_container_width=False)
            
            col_btn_back, col_btn_go = st.columns([1, 4])
            with col_btn_back:
                if st.button("⬅️ 回到第二步", type="secondary"):
                    st.session_state.step = 2
                    st.session_state.pack_sched = None
                    st.rerun()
            with col_btn_go:
                if st.button("✅ 確認名單，開始預排包班", type="primary"):
                    with st.spinner("正在執行包班平滑區塊分配..."):
                        sched_df = st.session_state.base_sched.copy()
                        sched = {i: [""] + list(sched_df.iloc[i, 1:].values) for i in range(len(ai_df))}
                        
                        class_days_map = {i: [s.strip() for s in str(row.get("上課日期", "")).split(",") if s.strip().isdigit()] for i, row in ai_df.iterrows()}
                        cache_pref = {i: str(row.get("包班意願", "")).strip() for i, row in ai_df.iterrows()}
                        # 假日出勤能力快取
                        cache_can_sat3 = {i: str(row.get("能上週六", "")).strip() == "是" for i, row in ai_df.iterrows()}
                        cache_can_sun3 = {i: str(row.get("能上週日", "")).strip() == "是" for i, row in ai_df.iterrows()}
                        cache_can_nat3 = {i: str(row.get("能上國定假日", "")).strip() == "是" for i, row in ai_df.iterrows()}
                        cache_group3   = {i: str(row.get("組別", "")).strip().upper() for i, row in ai_df.iterrows()}
                        sat_list3 = st.session_state.saturdays_list
                        sun_list3 = st.session_state.sundays_list
                        nat_list3 = st.session_state.nat_holidays_list
                        illegal_next = {"D": ["N"], "E": ["D", "N", "12-8"], "12-8": ["N"], "N": []}

                        # ── 第三步鎖定格集合（預白班 / 公差 / 國定必上班別 / 上課日期）──────────────
                        # 均等化互換時，這些格的班別不可被移走
                        _prewhite_set3: set = set()
                        _mand_hol_set3: set = set()
                        _gongcha_set3:  set = set()
                        _class_set3:    set = set()
                        for _pi3, _pr3 in ai_df.iterrows():
                            for _dv3 in str(_pr3.get("預白日期", "")).split(","):
                                if _dv3.strip().isdigit() and 1 <= int(_dv3.strip()) <= month_days:
                                    _prewhite_set3.add((_pi3, int(_dv3.strip())))
                            for _item3 in str(_pr3.get("國定假日必上班別", "")).split(","):
                                for _delim3 in [":", "-"]:
                                    if _delim3 in _item3:
                                        _dvx3, _svx3 = _item3.split(_delim3, 1)
                                        if _dvx3.strip().isdigit() and 1 <= int(_dvx3.strip()) <= month_days:
                                            _mand_hol_set3.add((_pi3, int(_dvx3.strip())))
                                        break
                            for _dv3 in str(_pr3.get("公差日期", "")).split(","):
                                if _dv3.strip().isdigit() and 1 <= int(_dv3.strip()) <= month_days:
                                    _gongcha_set3.add((_pi3, int(_dv3.strip())))
                            for _dv3 in str(_pr3.get("上課日期", "")).split(","):
                                if _dv3.strip().isdigit() and 1 <= int(_dv3.strip()) <= month_days:
                                    _class_set3.add((_pi3, int(_dv3.strip())))
                        _locked_set3 = _prewhite_set3 | _mand_hol_set3 | _gongcha_set3 | _class_set3

                        def can_work_base(n_idx, s, d_int):
                            if sched[n_idx][d_int] not in ["", "上課"]: return False
                            # 假日出勤能力限制（包班人員有假日出勤義務，不受此限）
                            if cache_pref[n_idx] == "" and not can_work_holiday_check(n_idx, d_int, cache_can_sat3, cache_can_sun3, cache_can_nat3, sat_list3, sun_list3, nat_list3): return False
                            # 上課日不得排 E 或 12-8（適用所有護理師，包班亦同）
                            if str(d_int) in class_days_map.get(n_idx, []) and s not in ("D", "N"): return False

                            y_s = (sched[n_idx][d_int - 1] or "") if d_int > 1 else ""
                            t_s = (sched[n_idx][d_int + 1] or "") if d_int < month_days else ""

                            y_s_base = "D" if (y_s.startswith("D") or y_s == "公差") else y_s
                            if y_s == "上課": y_s_base = s

                            t_s_base = "D" if (t_s.startswith("D") or t_s == "公差") else t_s
                            if t_s == "上課": t_s_base = s
                            
                            if is_work(y_s) and s in illegal_next.get(y_s_base, []): return False
                            if is_work(t_s) and t_s_base in illegal_next.get(s, []): return False

                            s_consec = 1
                            for bd in range(d_int - 1, 0, -1):
                                if is_work(sched[n_idx][bd]): s_consec += 1
                                else: break
                            for fd in range(d_int + 1, month_days + 1):
                                if is_work(sched[n_idx][fd]): s_consec += 1
                                else: break
                            if s_consec > 5: return False

                            _has_pb3 = bool((st.session_state.get("prev_month_buffer") or {}).get(str(ai_df.at[n_idx,"姓名"]).strip() if n_idx in ai_df.index else "", {}))
                            w_min = (d_int - 13) if _has_pb3 else max(1, d_int - 13)
                            w_max = min(d_int, month_days - 13) if month_days >= 14 else 1
                            for start_d in range(w_min, w_max + 1):
                                end_d = min(month_days, start_d + 13)
                                worked_in_window = 0
                                for curr_d in range(start_d, end_d + 1):
                                    if curr_d == d_int: continue
                                    if is_work(_xmonth_shift(n_idx, curr_d, sched, ai_df, month_days)): worked_in_window += 1
                                if worked_in_window + 1 > 12: return False
                            if not week_variety_ok(sched, n_idx, s, d_int, st.session_state.first_wday, month_days): return False
                            return True

                        # ── 包班人員排班輔助函數 ─────────────────────────
                        def get_pref_s(pref):
                            return "N" if "大夜" in pref else ("E" if "小夜" in pref else ("12-8" if "中" in pref else "D"))

                        def en_quota_full3(s_type, d_int):
                            """回傳 True 表示當日該班別人力配額已滿（不可再加）
                            支援 E / N / 12-8 / D 四種班別"""
                            _q_map = {"E": "E班", "N": "N班", "12-8": "12-8", "D": "D班"}
                            if s_type not in _q_map:
                                return False
                            q_col3 = _q_map[s_type]
                            row_d3 = edited_quota_df[edited_quota_df["日期"] == str(d_int)]
                            if row_d3.empty:
                                return False
                            try:
                                req3 = int(row_d3.iloc[0][q_col3])
                            except (KeyError, ValueError):
                                return False
                            if s_type == "D":
                                curr3 = sum(1 for i in ai_df.index
                                            if isinstance(sched[i][d_int], str) and sched[i][d_int].startswith("D"))
                            else:
                                curr3 = sum(1 for i in ai_df.index if sched[i][d_int] == s_type)
                            return curr3 >= req3

                        pack_indices3 = [i for i in ai_df.index if cache_pref[i] != ""]
                        holiday_days_set3 = set(st.session_state.holiday_list)
                        pref_s_set = set(get_pref_s(cache_pref[i]) for i in pack_indices3)

                        PACK_MIN_SHIFTS = 15

                        # 預先建立每位包班人員的 max_target 快取
                        _toff3 = st.session_state.target_off
                        def get_max_target3(idx, row):
                            el = calc_extra_leaves(row, month_days, sat_list3, sun_list3, nat_list3,
                                                   target_off=_toff3)
                            mt = month_days - _toff3 - el
                            if st.session_state.custom_targets and idx in st.session_state.custom_targets:
                                mt = st.session_state.custom_targets[idx]
                            return mt
                        max_target3 = {idx: get_max_target3(idx, row) for idx, row in ai_df.iterrows() if cache_pref[idx] != ""}

                        # ── 公平分配目標預計算：各包班班別全月總配額 ÷ 同組人數 ──────────
                        # 第一、二階段以此為上限，避免少數護士搶佔過多配額
                        # 第三階段兜底再統一補足至 PACK_MIN_SHIFTS（15班下限）
                        def _get_group_quota_col(ps):
                            if ps == "12-8": return "12-8"
                            if ps in ("D", "E", "N"): return f"{ps}班"
                            return None

                        pack_fair_target = {}
                        for _ps in pref_s_set:
                            _group = [i for i in pack_indices3 if get_pref_s(cache_pref[i]) == _ps]
                            _gsize = len(_group)
                            _qcol  = _get_group_quota_col(_ps)
                            if _gsize == 0 or _qcol is None:
                                pack_fair_target[_ps] = PACK_MIN_SHIFTS
                                continue
                            try:
                                _total_q = sum(
                                    int(edited_quota_df[edited_quota_df["日期"] == str(d)].iloc[0][_qcol])
                                    for d in range(1, month_days + 1)
                                    if not edited_quota_df[edited_quota_df["日期"] == str(d)].empty
                                )
                                pack_fair_target[_ps] = _total_q // _gsize
                            except (KeyError, ValueError, IndexError):
                                pack_fair_target[_ps] = PACK_MIN_SHIFTS

                        # ── 第一階段：上課日先個別處理 ──
                        for idx, row in ai_df.iterrows():
                            pref = cache_pref[idx]
                            if pref == "": continue
                            pref_s = get_pref_s(pref)
                            for d_int in range(1, month_days + 1):
                                if sched[idx][d_int] == "上課":
                                    if not en_quota_full3(pref_s, d_int) and can_work_base(idx, pref_s, d_int):
                                        sched[idx][d_int] = pref_s
                                    # E/N包班：配額滿或無法排包班班次時，保持上課狀態（不改成D班）
                                    elif pref_s not in ("E", "N"):
                                        sched[idx][d_int] = "D"

                        # ── 第一階段（續）：非假日天 — 以日為外迴圈均分包班班別 ──
                        # 每日對同班別組依「本班別累計班數最少」排序，確保各人均等分配
                        # 同班數時，以「今日排班可形成連班」為次排序（避免上一休一）
                        def pack_streak_key3(i, d):
                            y = sched[i][d - 1] if d > 1 else ""
                            t = sched[i][d + 1] if d < month_days else ""
                            if is_work(y) and is_work(t): return -2   # 填補空隙最優先
                            if is_work(y) or is_work(t): return -1    # 延伸連班次之
                            return 0                                   # 孤立班最後

                        for d_int in range(1, month_days + 1):
                            if d_int in holiday_days_set3: continue
                            for pref_s in pref_s_set:
                                group = [i for i in pack_indices3 if get_pref_s(cache_pref[i]) == pref_s]
                                if not group: continue
                                # 主排：本班別累計班數升序；次排：索引（穩定、公平）
                                # 移除連班型態加分，防止早取班者形成正反饋連排優勢
                                group_sorted = sorted(group, key=lambda i: (
                                    sum(1 for v in sched[i] if v == pref_s),
                                    i
                                ))
                                for idx in group_sorted:
                                    if sum(1 for v in sched[idx] if is_work(v)) >= max_target3[idx]: continue
                                    # 公平分配上限：Stage 1 最多排到 PACK_MIN_SHIFTS（15班下限）
                                    # 超過後暫停，由 Stage 3 日外迴圈均等補足至 max_target
                                    if sum(1 for v in sched[idx] if v == pref_s) >= PACK_MIN_SHIFTS: continue
                                    # ── 假日E班硬性上限：E包班護師每月假日出勤 ≤ 上限 ──────────────
                                    # 上限 = ceil(假日E配額總量 / E包班人數)，避免集中在少數人
                                    if d_int in holiday_days_set3:
                                        _hol_set1 = holiday_days_set3
                                        _e_pack_cnt1 = sum(1 for j in pack_indices3 if get_pref_s(cache_pref[j]) == pref_s)
                                        _hol_e_demand1 = sum(
                                            int(edited_quota_df[edited_quota_df["日期"]==str(hd)].iloc[0].get(f"{pref_s}班", 0))
                                            for hd in _hol_set1
                                            if not edited_quota_df[edited_quota_df["日期"]==str(hd)].empty
                                        )
                                        _hol_e_cap1 = max(2, -(-_hol_e_demand1 // max(_e_pack_cnt1, 1)))  # ceil除法
                                        _hol_e_worked1 = sum(1 for hd in _hol_set1 if sched[idx][hd] == pref_s)
                                        if _hol_e_worked1 >= _hol_e_cap1: continue
                                    if en_quota_full3(pref_s, d_int): break  # 當日額滿，跳下一班別
                                    if can_work_base(idx, pref_s, d_int) and group_cap_ok(idx, pref_s, d_int, sched, cache_group3):
                                        sched[idx][d_int] = pref_s

                        # ── 第二階段：假日班次均衡分配（全月總班數少者優先） ──
                        # 依包班班別分組，在每個假日中輪流給總班數最少的護理師
                        for d_int in sorted(holiday_days_set3):
                            for pref_s in pref_s_set:
                                group = [i for i in pack_indices3 if get_pref_s(cache_pref[i]) == pref_s]
                                if not group: continue
                                # 主排：假日出勤最少者優先（均衡假日負擔）
                                # 次排：包班班次最少者；末排：索引
                                _hol_set_s2 = set(sat_list3) | set(sun_list3) | set(nat_list3)
                                group_sorted = sorted(
                                    group,
                                    key=lambda i: (
                                        sum(1 for hd in _hol_set_s2 if is_work(sched[i][hd])),
                                        sum(1 for v in sched[i] if v == pref_s),
                                        i
                                    )
                                )
                                for idx in group_sorted:
                                    if sum(1 for v in sched[idx] if is_work(v)) >= max_target3[idx]: continue
                                    # 公平分配上限：Stage 2 同樣最多到 PACK_MIN_SHIFTS（15班）
                                    if sum(1 for v in sched[idx] if v == pref_s) >= PACK_MIN_SHIFTS: continue
                                    if en_quota_full3(pref_s, d_int): break  # 此假日 E/N 額已滿
                                    if can_work_base(idx, pref_s, d_int) and group_cap_ok(idx, pref_s, d_int, sched, cache_group3):
                                        sched[idx][d_int] = pref_s
                                        break  # 每班別在此假日只優先排一人（下次輪到下一位）

                        # ── 第三階段：均等補足至應上班天數 ──
                        # 改用「日為外迴圈、人為內迴圈」，與第一、二階段相同結構
                        # 每日讓「總出勤最少 → 包班班次最少」的人優先，確保公平填充至 max_target
                        for d_int in range(1, month_days + 1):
                            for pref_s in pref_s_set:
                                if en_quota_full3(pref_s, d_int): continue
                                group3 = [i for i in pack_indices3 if get_pref_s(cache_pref[i]) == pref_s]
                                if not group3: continue
                                _hol_set_s3 = set(sat_list3) | set(sun_list3) | set(nat_list3)
                                _is_hol_d3 = d_int in _hol_set_s3
                                group3_sorted = sorted(group3, key=lambda i: (
                                    # 假日天：假日出勤最少者最優先；平日天：總出勤最少者最優先
                                    sum(1 for hd in _hol_set_s3 if is_work(sched[i][hd])) if _is_hol_d3
                                    else sum(1 for v in sched[i] if is_work(v)),
                                    sum(1 for v in sched[i] if is_work(v)),                               # 次排：總出勤
                                    sum(1 for v in sched[i] if v == pref_s) / max(max_target3[i], 1),    # 再排：達標比例
                                ))
                                # 假日和平日都用 len//2 限制，讓每天輪流讓不同人取E
                                # 排序已依假日出勤最少者優先，不需要用_day_limit=1來強制輪替
                                _day_limit3 = max(1, len(group3) // 2)
                                _day_placed3 = 0   # 本日已排入人數
                                for idx in group3_sorted:
                                    if _day_placed3 >= _day_limit3: break  # 本日已達上限，讓位給其他人
                                    if sum(1 for v in sched[idx] if is_work(v)) >= max_target3[idx]: continue
                                    # ── 假日E班硬性上限（Stage 3 同樣套用）──────────────────────
                                    if d_int in _hol_set_s3:
                                        _e_pack_cnt3 = sum(1 for j in pack_indices3 if get_pref_s(cache_pref[j]) == pref_s)
                                        _hol_e_demand3 = sum(
                                            int(edited_quota_df[edited_quota_df["日期"]==str(hd)].iloc[0].get(f"{pref_s}班", 0))
                                            for hd in _hol_set_s3
                                            if not edited_quota_df[edited_quota_df["日期"]==str(hd)].empty
                                        )
                                        _hol_e_cap3 = max(2, -(-_hol_e_demand3 // max(_e_pack_cnt3, 1)))
                                        _hol_e_worked3 = sum(1 for hd in _hol_set_s3 if sched[idx][hd] == pref_s)
                                        if _hol_e_worked3 >= _hol_e_cap3: continue
                                    if en_quota_full3(pref_s, d_int): break
                                    if can_work_base(idx, pref_s, d_int) and group_cap_ok(idx, pref_s, d_int, sched, cache_group3):
                                        sched[idx][d_int] = pref_s
                                        _day_placed3 += 1

                        # ── 第四階段：包班天數讓渡均衡 ─────────────────────────
                        # 若仍有人未達 15 班下限，嘗試從同組班次較多的人員讓渡可交換的日期
                        def _pack_min_info(idx, row):
                            """回傳 (pref_s, min_pack, cur_count)
                            12-8 包班不設下限（mp=0），其餘包班維持 PACK_MIN_SHIFTS 下限。"""
                            pref  = cache_pref[idx]
                            if pref == "": return None
                            ps    = get_pref_s(pref)
                            el    = calc_extra_leaves(row, month_days, sat_list3, sun_list3, nat_list3,
                                                      target_off=_toff3)
                            mt    = month_days - _toff3 - el
                            if st.session_state.custom_targets and idx in st.session_state.custom_targets:
                                mt = st.session_state.custom_targets[idx]
                            mp    = 0 if ps == "12-8" else min(PACK_MIN_SHIFTS, mt)
                            cur   = sum(1 for v in sched[idx] if v == ps)
                            return ps, mp, cur

                        # 收集仍不足者（可能多輪才能補足，最多迭代 5 輪）
                        for _rebal_round in range(5):
                            _under = [
                                (idx, *_pack_min_info(idx, ai_df.loc[idx])[:2])
                                for idx in ai_df.index
                                if cache_pref[idx] != ""
                                and _pack_min_info(idx, ai_df.loc[idx])[2] < _pack_min_info(idx, ai_df.loc[idx])[1]
                            ]
                            if not _under: break  # 全部達標，退出

                            _made_progress = False
                            for b_idx, b_pref_s, b_min in _under:
                                for d_int in range(1, month_days + 1):
                                    if sum(1 for v in sched[b_idx] if v == b_pref_s) >= b_min: break
                                    if sched[b_idx][d_int] not in ["", "上課"]: continue
                                    if not can_work_base(b_idx, b_pref_s, d_int): continue

                                    # 若配額未滿，直接補（漏網之魚）
                                    if not en_quota_full3(b_pref_s, d_int):
                                        sched[b_idx][d_int] = b_pref_s
                                        _made_progress = True
                                        continue

                                    # 配額已滿 → 尋找同組可讓渡者
                                    _peers = sorted(
                                        [i for i in pack_indices3
                                         if get_pref_s(cache_pref[i]) == b_pref_s
                                         and i != b_idx
                                         and sched[i][d_int] == b_pref_s],
                                        key=lambda i: -sum(1 for d in range(1, month_days+1) if sched[i][d] == b_pref_s)
                                    )
                                    for a_idx in _peers:
                                        # a 讓出此班後的班數
                                        a_after = sum(
                                            1 for d in range(1, month_days+1)
                                            if d != d_int and sched[a_idx][d] == b_pref_s
                                        )
                                        _, a_min, _ = _pack_min_info(a_idx, ai_df.loc[a_idx])

                                        # ── 情況 1：a 讓出後仍 ≥ 下限（直接讓渡）──
                                        if a_after >= a_min:
                                            # ★ 保護：a 的班若為鎖定格（國定必上班別/公差/預白），不可讓渡
                                            if (a_idx, d_int) in _locked_set3:
                                                continue
                                            sched[a_idx][d_int] = ""
                                            if can_work_base(b_idx, b_pref_s, d_int) and group_cap_ok(b_idx, b_pref_s, d_int, sched, cache_group3):
                                                sched[b_idx][d_int] = b_pref_s
                                                _made_progress = True
                                                break
                                            else:
                                                sched[a_idx][d_int] = b_pref_s  # 復原

                                        # ── 情況 2：a 讓出後會跌破下限 → 嘗試「日期對調」──
                                        # a 讓出 d_int 給 b，同時 a 去搶 b 沒排到的某天 d2
                                        else:
                                            # ★ 保護：a 的班若為鎖定格，不可讓渡
                                            if (a_idx, d_int) in _locked_set3:
                                                continue
                                            sched[a_idx][d_int] = ""
                                            if not can_work_base(b_idx, b_pref_s, d_int):
                                                sched[a_idx][d_int] = b_pref_s  # b 無法接，復原
                                                continue
                                            # b 可接此日；找 a 能補的替代日 d2
                                            _swapped = False
                                            for d2 in range(1, month_days + 1):
                                                if d2 == d_int: continue
                                                if sched[a_idx][d2] not in ["", "上課"]: continue
                                                if not en_quota_full3(b_pref_s, d2): continue  # d2 配額未滿，a 直接排即可（不需對調）
                                                if (can_work_base(a_idx, b_pref_s, d2)
                                                        and group_cap_ok(b_idx, b_pref_s, d_int, sched, cache_group3)
                                                        and group_cap_ok(a_idx, b_pref_s, d2, sched, cache_group3)):
                                                    sched[b_idx][d_int] = b_pref_s
                                                    sched[a_idx][d2]    = b_pref_s
                                                    _made_progress = True
                                                    _swapped = True
                                                    break
                                            if not _swapped:
                                                sched[a_idx][d_int] = b_pref_s  # 復原
                                            if _swapped:
                                                break

                            if not _made_progress: break  # 本輪無任何進展，停止

                        # ── 第五階段：E/N包班護士補充班次（讓渡均衡後仍不足時的最後手段）─────
                        # E班包班 → 補充 12-8 班；N班包班 → 補充 D班（白班）
                        # 此步驟在包班班次排入、讓渡均衡完成後執行，再接常規夜班排班
                        def _supp_count(row_sched, ss):
                            """計算補充班次數（D班含所有D開頭變體）"""
                            if ss == "D":
                                return sum(1 for v in row_sched if isinstance(v, str) and v.startswith("D"))
                            return sum(1 for v in row_sched if v == ss)

                        # ── 第五階段：E/N包班護士補充班次（讓渡均衡後仍不足時的最後手段）─────
                        # E班包班 → 補充 12-8 班；N班包班 → 補充 D班（白班）
                        # 此步驟在包班班次排入、讓渡均衡完成後執行，再接常規夜班排班
                        for idx, row in ai_df.iterrows():
                            pref = cache_pref[idx]
                            if pref == "": continue
                            pref_s = get_pref_s(pref)
                            if pref_s not in ("E", "N"): continue  # 僅適用 E/N 包班護士
                            extra_leaves = calc_extra_leaves(row, month_days, sat_list3, sun_list3, nat_list3,
                                                              target_off=_toff3)
                            max_target = month_days - _toff3 - extra_leaves
                            if st.session_state.custom_targets and idx in st.session_state.custom_targets:
                                max_target = st.session_state.custom_targets[idx]
                            min_pack = 0 if pref_s == "12-8" else min(PACK_MIN_SHIFTS, max_target)
                            supp_s = "12-8" if pref_s == "E" else "D"
                            for d_int in range(1, month_days + 1):
                                pack_now = sum(1 for v in sched[idx] if v == pref_s)
                                if pack_now < min_pack: break
                                if sum(1 for v in sched[idx] if is_work(v)) >= max_target: break
                                if sched[idx][d_int] not in ["", "上課"]: continue
                                if en_quota_full3(supp_s, d_int): continue
                                if can_work_base(idx, supp_s, d_int) and group_cap_ok(idx, supp_s, d_int, sched, cache_group3):
                                    sched[idx][d_int] = supp_s

                        # ── 包班下限檢查：警示以包班班別（pref_s）班數為準 ──
                        # 補充班次（12-8 / D班）不計入警示判斷，但顯示於表格供參考
                        _pack_warnings3 = []
                        for idx, row in ai_df.iterrows():
                            pref = cache_pref[idx]
                            if pref == "": continue
                            pref_s = get_pref_s(pref)
                            extra_leaves = calc_extra_leaves(row, month_days, sat_list3, sun_list3, nat_list3,
                                                              target_off=_toff3)
                            max_target = month_days - _toff3 - extra_leaves
                            if st.session_state.custom_targets and idx in st.session_state.custom_targets:
                                max_target = st.session_state.custom_targets[idx]
                            min_pack = 0 if pref_s == "12-8" else min(PACK_MIN_SHIFTS, max_target)
                            actual_pack = sum(1 for v in sched[idx] if v == pref_s)
                            if actual_pack < min_pack:
                                supp_s = "12-8" if pref_s == "E" else ("D" if pref_s in ("N", "12-8") else "")
                                supp_count = _supp_count(sched[idx], supp_s) if supp_s else 0
                                _pack_warnings3.append({
                                    "姓名": row["姓名"],
                                    "包班班別": pref_s,
                                    "應達下限": min_pack,
                                    f"{pref_s}班實排": actual_pack,
                                    "補充班次": f"{supp_s}×{supp_count}" if supp_count > 0 else "—",
                                    "合計": actual_pack + supp_count,
                                    "差距": min_pack - actual_pack,
                                })
                        st.session_state.pack_warnings = _pack_warnings3

                        pack_df = pd.DataFrame({"姓名": ai_df["姓名"]})
                        for d in range(1, month_days + 1):
                            pack_df[str(d)] = [sched[i][d] for i in ai_df.index]

                        st.session_state.pack_sched = pack_df
                        st.rerun()
        else:
            st.info("本月無包班人員，您可以直接進入下一步。")
            col_btn_back, col_btn_go = st.columns([1, 4])
            with col_btn_back:
                if st.button("⬅️ 回到第二步", type="secondary"):
                    st.session_state.step = 2
                    st.session_state.pack_sched = None
                    st.rerun()
            with col_btn_go:
                if st.button("⏩ 直接進入下一步", type="primary"):
                    st.session_state.pack_sched = st.session_state.base_sched.copy()
                    st.session_state.step = 4
                    st.rerun()

    if st.session_state.pack_sched is not None:
        # ── 包班下限警示 ──────────────────────────────────────────────────
        _pw = st.session_state.get("pack_warnings", [])
        if _pw:
            st.error(f"⚠️ 以下 {len(_pw)} 位包班人員在遵守人力配額及勞基法規定下，**無法達到包班 15 班下限**（系統已執行讓渡均衡及白班補班，仍屬結構性限制）。請調整 E/N 班每日配額上限，或至第 7 步手動補排：")
            _pw_df = pd.DataFrame(_pw)
            st.dataframe(_pw_df, use_container_width=False, hide_index=True)
        else:
            st.success("✅ 包班人員安插完畢！")
        with st.expander("📄 點擊展開包班安插結果", expanded=True):
            _edit_pack = st.checkbox("🖊️ 開啟手動編輯模式", value=False, key="chk_edit_pack_sched")
            if _edit_pack:
                st.caption("💡 直接點擊儲存格輸入班別代碼，修改即時儲存")
                _edited_pack = st.data_editor(
                    st.session_state.pack_sched,
                    column_config=make_sched_col_config(month_days),
                    use_container_width=True, hide_index=True, key="edit_pack_sched"
                )
                st.session_state.pack_sched = _edited_pack
            else:
                _day_cols_p = [str(i) for i in range(1, month_days + 1)]
                _disp_pack = st.session_state.pack_sched.copy()
                for _c in _day_cols_p:
                    _disp_pack[_c] = _disp_pack[_c].apply(abbrev_display)
                # 預白班：D 格標為 Dx
                _disp_pack = apply_prewhite_dx(_disp_pack, st.session_state.ai_df, month_days)
                st.dataframe(
                    build_schedule_with_counts(_disp_pack, st.session_state.pack_sched, _day_cols_p, st.session_state.ai_df),
                    use_container_width=True
                )

        # ── 包班人員上班天數摘要 ──────────────────────────────────────────────
        _pack_summary_rows = []
        for _ps_idx, _ps_row in ai_df.iterrows():
            _ps_pref = str(_ps_row.get("包班意願", "")).strip()
            if _ps_pref == "": continue
            _ps_pref_s = ("N" if "大夜" in _ps_pref else ("E" if "小夜" in _ps_pref else ("12-8" if "中" in _ps_pref else "D")))
            _ps_el = calc_extra_leaves(_ps_row, month_days,
                                       st.session_state.saturdays_list,
                                       st.session_state.sundays_list,
                                       st.session_state.nat_holidays_list,
                                       target_off=st.session_state.target_off)
            _ps_max = month_days - st.session_state.target_off - _ps_el
            if st.session_state.get("custom_targets") and _ps_idx in st.session_state.custom_targets:
                _ps_max = st.session_state.custom_targets[_ps_idx]
            _ps_sched_vals = list(st.session_state.pack_sched.iloc[_ps_idx, 1:].values) if _ps_idx < len(st.session_state.pack_sched) else []
            _ps_pack_cnt  = sum(1 for v in _ps_sched_vals if v == _ps_pref_s)
            _ps_total_cnt = sum(1 for v in _ps_sched_vals if is_work(str(v).strip()))
            _ps_supp_s    = "12-8" if _ps_pref_s == "E" else "D"
            _ps_supp_cnt  = sum(1 for v in _ps_sched_vals if (str(v).startswith("D") if _ps_supp_s == "D" else str(v) == _ps_supp_s))
            _ps_gap       = _ps_max - _ps_total_cnt
            if _ps_gap <= 0:
                _ps_status = "✅ 達標"
            elif _ps_gap == 1:
                _ps_status = "⚠️ 少 1 天"
            else:
                _ps_status = f"🚨 欠 {_ps_gap} 天"
            _pack_summary_rows.append({
                "姓名":        _ps_row["姓名"],
                "包班班別":    _ps_pref_s,
                f"{_ps_pref_s}班": _ps_pack_cnt,
                "補充班":      _ps_supp_cnt,
                "總出勤":      _ps_total_cnt,
                "應上班":      _ps_max,
                "狀態":        _ps_status,
            })
        if _pack_summary_rows:
            with st.expander("📊 包班人員上班天數摘要", expanded=True):
                _ps_df = pd.DataFrame(_pack_summary_rows)
                st.dataframe(_ps_df, use_container_width=False, hide_index=True)

        display_safety_radar(st.session_state.pack_sched, edited_quota_df, st.session_state.ai_df)

        # ── 空格充裕度預警（進入第四步前）──────────────────────────────────────
        _sat_pw = st.session_state.saturdays_list
        _sun_pw = st.session_state.sundays_list
        _nat_pw = st.session_state.nat_holidays_list
        _hol_set_pw = set(_sat_pw) | set(_sun_pw) | set(_nat_pw)
        _weekday_count_pw = sum(1 for _d in range(1, month_days + 1) if _d not in _hol_set_pw)
        _toff_pw = st.session_state.target_off

        # 計算 personal_targets（與第四步相同邏輯）
        _pt_pw = {}
        for _idx_pw, _row_pw in ai_df.iterrows():
            _title_pw = str(_row_pw.get("職稱", "")).strip()
            if _title_pw in NO_HOL_SET:
                _el_pw = calc_extra_leaves(_row_pw, month_days, _sat_pw, _sun_pw, _nat_pw, target_off=0)
                _pt_pw[_idx_pw] = max(0, _weekday_count_pw - _el_pw)
            else:
                _el_pw = calc_extra_leaves(_row_pw, month_days, _sat_pw, _sun_pw, _nat_pw, target_off=_toff_pw)
                _pt_pw[_idx_pw] = max(0, month_days - _toff_pw - _el_pw)
        if st.session_state.get("custom_targets"):
            for _idx_pw in ai_df.index:
                if _idx_pw in st.session_state.custom_targets:
                    _pt_pw[_idx_pw] = st.session_state.custom_targets[_idx_pw]

        # 一般護理師：職稱不在 ADMIN_TITLES 且包班意願為空白
        _gen_pw = [
            _i for _i, _r in ai_df.iterrows()
            if str(_r.get("職稱", "")).strip() not in ADMIN_TITLES
            and str(_r.get("包班意願", "")).strip() == ""
        ]

        _total_slots_pw = month_days * len(_gen_pw)

        # 掃描 pack_sched
        _pack_df_pw = st.session_state.pack_sched
        _day_cols_pw = [str(_d) for _d in range(1, month_days + 1)]
        _filled_pw   = 0   # 非空格（任何已有值的格子）
        _worked_pw   = {}  # 各護理師已排工作班數（供欠班計算）
        for _i_pw in _gen_pw:
            _w_pw = 0
            for _dc_pw in _day_cols_pw:
                _v_pw = str(_pack_df_pw.loc[_i_pw, _dc_pw] if _dc_pw in _pack_df_pw.columns else "").strip()
                if _v_pw not in ("", "nan"):
                    _filled_pw += 1
                if is_work(_v_pw):
                    _w_pw += 1
            _worked_pw[_i_pw] = _w_pw

        _remaining_pw = _total_slots_pw - _filled_pw
        _deficit_pw   = sum(max(0, _pt_pw.get(_i_pw, 0) - _worked_pw.get(_i_pw, 0)) for _i_pw in _gen_pw)

        if _remaining_pw < _deficit_pw:
            st.error(
                f"⚠️ 預警：剩餘可排空格（{_remaining_pw} 格）少於欠班缺口（{_deficit_pw} 天），"
                f"建議回頭提高每日配額或減少包班天數。"
            )
        else:
            st.success(
                f"✅ 空格充裕，剩餘 {_remaining_pw} 格，欠班缺口 {_deficit_pw} 天，繼續排班無虞。"
            )

        if st.session_state.step == 3:
            col_btn_back, col_btn_go = st.columns([1, 4])
            with col_btn_back:
                if st.button("⬅️ 回到第二步 (重調配額)", type="secondary"):
                    st.session_state.step = 2
                    st.session_state.pack_sched = None
                    st.rerun()
            with col_btn_go:
                if st.button("✅ 確認包班無誤，排入常規夜班", type="primary"):
                    st.session_state.step = 4
                    st.rerun()

# ==========================================
# 📍 第四步：自動安插常規夜班 (包含流動與控台防護)
# ==========================================
if st.session_state.step >= 4:
    st.divider()
    st.header("4️⃣ 第四步：排入常規夜班")

    ai_df = st.session_state.ai_df
    month_days = st.session_state.month_days
    edited_quota_df = st.session_state.edited_quota_df

    with st.expander("⚙️ 調整每日人力配額（修改後需重新排班才生效）", expanded=False):
        _q4 = st.data_editor(st.session_state.edited_quota_df, hide_index=True, use_container_width=True, key="quota_editor_step4")
        if st.button("✅ 套用配額變更並重排（從第四步重算）", key="apply_quota_step4"):
            st.session_state.edited_quota_df = _q4
            st.session_state.night_sched = None
            st.session_state.d_sched = None
            st.session_state.final_sched = None
            st.rerun()
    edited_quota_df = st.session_state.edited_quota_df
    
    if st.session_state.night_sched is None:
        col_btn_back, col_btn_go = st.columns([1, 4])
        with col_btn_back:
            if st.button("⬅️ 回到第三步", type="secondary"):
                st.session_state.step = 3
                st.session_state.night_sched = None
                st.rerun()
        with col_btn_go:
            if st.button("✅ 啟動防護網並均分夜班 (E / N / 12-8)", type="primary"):
                with st.spinner("正在執行附帶流動/階梯控台安全鎖的夜班均分矩陣..."):
                    sched_df = st.session_state.pack_sched.copy()
                    sched = {i: [""] + list(sched_df.iloc[i, 1:].values) for i in range(len(ai_df))}
                    
                    cache_preg = {i: str(row.get("孕/育嬰免夜班", "")).strip() == "是" for i, row in ai_df.iterrows()}
                    cache_night = {i: str(row.get("夜班資格", "")).strip() for i, row in ai_df.iterrows()}  # 空白 = 無夜班資格
                    cache_pref = {i: str(row.get("包班意願", "")).strip() for i, row in ai_df.iterrows()}
                    cache_title = {i: str(row.get("職稱", "")).strip() for i, row in ai_df.iterrows()}
                    cache_circ = {i: str(row.get("流動資格", "")).strip() == "是" for i, row in ai_df.iterrows()}
                    cache_leader_str = {i: str(row.get("控台資格", "")).strip() for i, row in ai_df.iterrows()}
                    class_days_map = {i: [s.strip() for s in str(row.get("上課日期", "")).split(",") if s.strip().isdigit()] for i, row in ai_df.iterrows()}
                    # 假日出勤能力快取（第四步）
                    cache_can_sat4 = {i: str(row.get("能上週六", "")).strip() == "是" for i, row in ai_df.iterrows()}
                    cache_can_sun4 = {i: str(row.get("能上週日", "")).strip() == "是" for i, row in ai_df.iterrows()}
                    cache_can_nat4 = {i: str(row.get("能上國定假日", "")).strip() == "是" for i, row in ai_df.iterrows()}
                    cache_group4   = {i: str(row.get("組別", "")).strip().upper() for i, row in ai_df.iterrows()}
                    # ── 預休接軌快取：記錄每位護理師的預休（O）日期集合 ──
                    cache_pre_off4 = {
                        i: {int(d.strip()) for d in str(row.get("預休日期", "")).split(",") if d.strip().isdigit()}
                        for i, row in ai_df.iterrows()
                    }
                    sat_list4 = st.session_state.saturdays_list
                    sun_list4 = st.session_state.sundays_list
                    nat_list4 = st.session_state.nat_holidays_list

                    personal_targets = {}
                    _hol_set4 = set(sat_list4) | set(sun_list4) | set(nat_list4)
                    _weekday_count4 = sum(1 for d in range(1, month_days + 1) if d not in _hol_set4)
                    _toff4 = st.session_state.target_off
                    for idx, row in ai_df.iterrows():
                        _title4 = str(row.get("職稱", "")).strip()
                        if _title4 in NO_HOL_SET:
                            # 護理長/副護理長/助理/傷兵/組長：不排假日班，預休/長假/特假均直接扣平日
                            # target_off=0 → 所有平日假別（O+長假+特假）都算扣減
                            extra_leaves = calc_extra_leaves(row, month_days, sat_list4, sun_list4, nat_list4, target_off=0)
                            personal_targets[idx] = max(0, _weekday_count4 - extra_leaves)
                        else:
                            # 一般護理師：O+長假先佔用 target_off，超出才扣工作天；特假直接扣
                            extra_leaves = calc_extra_leaves(row, month_days, sat_list4, sun_list4, nat_list4, target_off=_toff4)
                            personal_targets[idx] = max(0, month_days - _toff4 - extra_leaves)
                    # 套用手動調整的應上班天數
                    if st.session_state.custom_targets:
                        for idx in ai_df.index:
                            if idx in st.session_state.custom_targets:
                                personal_targets[idx] = st.session_state.custom_targets[idx]

                    illegal_next = {"D": ["N"], "E": ["D", "N", "12-8"], "12-8": ["N"], "N": []}

                    # ── 第四步鎖定格集合（預白班 / 公差 / 國定必上班別 / 上課日期）──────────────
                    # E/N 均等化互換時，這些格的班別不可被移走
                    _prewhite_set4: set = set()
                    _mand_hol_set4: set = set()
                    _gongcha_set4:  set = set()
                    _class_set4:    set = set()
                    for _pi4, _pr4 in ai_df.iterrows():
                        for _dv4 in str(_pr4.get("預白日期", "")).split(","):
                            if _dv4.strip().isdigit() and 1 <= int(_dv4.strip()) <= month_days:
                                _prewhite_set4.add((_pi4, int(_dv4.strip())))
                        for _item4 in str(_pr4.get("國定假日必上班別", "")).split(","):
                            for _delim4 in [":", "-"]:
                                if _delim4 in _item4:
                                    _dvx4, _svx4 = _item4.split(_delim4, 1)
                                    if _dvx4.strip().isdigit() and 1 <= int(_dvx4.strip()) <= month_days:
                                        _mand_hol_set4.add((_pi4, int(_dvx4.strip())))
                                    break
                        for _dv4 in str(_pr4.get("公差日期", "")).split(","):
                            if _dv4.strip().isdigit() and 1 <= int(_dv4.strip()) <= month_days:
                                _gongcha_set4.add((_pi4, int(_dv4.strip())))
                        for _dv4 in str(_pr4.get("上課日期", "")).split(","):
                            if _dv4.strip().isdigit() and 1 <= int(_dv4.strip()) <= month_days:
                                _class_set4.add((_pi4, int(_dv4.strip())))
                    _locked_set4 = _prewhite_set4 | _mand_hol_set4 | _gongcha_set4 | _class_set4

                    def can_work_base(n_idx, s, d_int, strict_wow=True):
                        if sched[n_idx][d_int] not in ["", "上課"]: return False
                        # 假日出勤能力限制（包班人員有假日出勤義務，不受此限）
                        if cache_pref[n_idx] == "" and not can_work_holiday_check(n_idx, d_int, cache_can_sat4, cache_can_sun4, cache_can_nat4, sat_list4, sun_list4, nat_list4): return False
                        if cache_preg[n_idx] and s in ("E", "N"): return False  # 母性保護僅禁 E/N；12-8 仍可排
                        if cache_title[n_idx] in ADMIN_TITLES: return False

                        # 上課日不得排 E 或 12-8（適用所有護理師，包班亦同）
                        if str(d_int) in class_days_map.get(n_idx, []) and s not in ("D", "N"): return False

                        qual = cache_night[n_idx]
                        if s == "N" and qual != "大夜": return False
                        if s == "E" and qual not in ["大夜", "小夜"]: return False
                        
                        worked = sum(1 for x in sched[n_idx] if is_work(x))
                        if worked >= personal_targets[n_idx]: return False

                        y_s = (sched[n_idx][d_int - 1] or "") if d_int > 1 else ""
                        t_s = (sched[n_idx][d_int + 1] or "") if d_int < month_days else ""
                        y_s_base = "D" if y_s.startswith("D") or y_s in ("上課", "公差") else y_s
                        t_s_base = "D" if t_s.startswith("D") or t_s in ("上課", "公差") else t_s

                        if is_work(y_s) and s in illegal_next.get(y_s_base, []): return False
                        if is_work(t_s) and t_s_base in illegal_next.get(s, []): return False

                        if strict_wow:
                            yy_s = sched[n_idx][d_int - 2] if d_int > 2 else "O"
                            tt_s = sched[n_idx][d_int + 2] if d_int < month_days - 1 else "O"
                            if not is_work(y_s) and is_work(yy_s): return False
                            if not is_work(t_s) and is_work(tt_s): return False

                        s_consec = 1
                        for bd in range(d_int - 1, 0, -1):
                            if is_work(sched[n_idx][bd]): s_consec += 1
                            else: break
                        for fd in range(d_int + 1, month_days + 1):
                            if is_work(sched[n_idx][fd]): s_consec += 1
                            else: break
                        if s_consec > 5: return False

                        _has_pb4 = bool((st.session_state.get("prev_month_buffer") or {}).get(str(ai_df.at[n_idx,"姓名"]).strip() if n_idx in ai_df.index else "", {}))
                        w_min = (d_int - 13) if _has_pb4 else max(1, d_int - 13)
                        w_max = min(d_int, month_days - 13) if month_days >= 14 else 1
                        for start_d in range(w_min, w_max + 1):
                            end_d = min(month_days, start_d + 13)
                            worked_in_window = 0
                            for curr_d in range(start_d, end_d + 1):
                                if curr_d == d_int: continue
                                if is_work(_xmonth_shift(n_idx, curr_d, sched, ai_df, month_days)): worked_in_window += 1
                            if worked_in_window + 1 > 12: return False
                        if not week_variety_ok(sched, n_idx, s, d_int, st.session_state.first_wday, month_days): return False
                        return True

                    elig_night_nurses = [i for i in ai_df.index if cache_pref[i] == "" and cache_night[i] in ("大夜", "小夜") and not cache_preg[i] and cache_title[i] not in ADMIN_TITLES]
                    total_night_demand = sum(int(edited_quota_df.iloc[d-1]["N班"]) + int(edited_quota_df.iloc[d-1]["E班"]) for d in range(1, month_days + 1))
                    pack_night_supply = sum(sum(1 for v in sched[i] if v in ["N","E"]) for i in ai_df.index if cache_pref[i] != "")
                    remaining_night_demand = max(0, total_night_demand - pack_night_supply)
                    target_night = remaining_night_demand // len(elig_night_nurses) if elig_night_nurses else 0

                    _MIN_BLOCK = 3
                    _block_remaining = {s: {} for s in ["N", "E"]}
                    _elig_night = [i for i in ai_df.index
                                   if cache_pref[i] == "" and cache_night[i] in ("大夜", "小夜")
                                   and not cache_preg[i] and cache_title[i] not in ADMIN_TITLES]
                    def _get_block_cap(s_type):
                        total_req = sum(
                            int(edited_quota_df[edited_quota_df["日期"] == str(d)].iloc[0][f"{s_type}班"])
                            for d in range(1, month_days + 1)
                            if not edited_quota_df[edited_quota_df["日期"] == str(d)].empty
                        )
                        return max(int(total_req / max(len(_elig_night), 1) * 2.0), _MIN_BLOCK + 1)
                    _block_cap = {s: _get_block_cap(s) for s in ["N", "E"]}

                    def assign_night_shifts(s_type):
                        q_col = f"{s_type}班"
                        for pass_num in [True, False]:
                            for _iter in range(month_days * 10):  # 最多迭代 month_days×10 次，防止無限迴圈
                                progress = False
                                deficits = []
                                for d in range(1, month_days + 1):
                                    req = int(edited_quota_df[edited_quota_df["日期"] == str(d)].iloc[0][q_col])
                                    curr = sum(1 for i in ai_df.index if sched[i][d] == s_type)
                                    if req > curr: deficits.append((d, req - curr, req))
                                
                                if not deficits: break
                                deficits.sort(key=lambda x: (x[1], random.random()), reverse=True)
                                
                                for d_int, defc, req in deficits:
                                    curr_nurses = [i for i in ai_df.index if sched[i][d_int] == s_type]
                                    curr_circ = sum(1 for i in curr_nurses if cache_circ[i])
                                    target_circ = (req + 1) // 2 if req > 0 else 0
                                    
                                    # 💡 階梯式控台檢測
                                    has_leader = False
                                    for i in curr_nurses:
                                        l_str = cache_leader_str[i]
                                        if not l_str: continue
                                        if "白" in l_str: has_leader = True
                                        elif "小" in l_str and s_type in ["E", "N"]: has_leader = True
                                        elif "大" in l_str and s_type == "N": has_leader = True
                                    
                                    available = [i for i in ai_df.index if can_work_base(i, s_type, d_int, strict_wow=pass_num)]
                                    available = [i for i in available if cache_pref[i] == ""]
                                    available = [i for i in available if group_cap_ok(i, s_type, d_int, sched, cache_group4)]
                                    if not available: continue

                                    # ── 最小塊優先：塊期內的護師直接插入，不參與競爭 ──
                                    _block_nurse = None
                                    for _bi in available:
                                        if _block_remaining[s_type].get(_bi, 0) > 0:
                                            _cur_n = sum(1 for v in sched[_bi] if v == s_type)
                                            if _cur_n >= _block_cap[s_type]:
                                                _block_remaining[s_type][_bi] = 0
                                                continue
                                            if not week_variety_ok(sched, _bi, s_type, d_int,
                                                                   st.session_state.first_wday, month_days):
                                                _block_remaining[s_type][_bi] = 0
                                                continue
                                            _block_nurse = _bi
                                            break

                                    def evaluate_nurse(idx):
                                        night_worked = sum(1 for v in sched[idx] if v in ["E", "N", "12-8"])
                                        score = 0
                                        # ── 個人化缺口比例加分（取代固定 target_night 比較）──────────────
                                        # 計算此人目前剩餘的可用空格數（空格 = 可能排夜班的位置）
                                        # 比例 = 已拿夜班 / 可用空格，比例越低代表「相對夜班不足」→ 越優先
                                        _avail4 = sum(
                                            1 for _d4 in range(1, month_days + 1)
                                            if sched[idx][_d4] == ""
                                        )
                                        _ratio4 = night_worked / max(_avail4, 1)
                                        # 比例從 0（夜班最少/空格最多）到 1（全排滿），加分從 500萬 遞減到 0
                                        score += int((1.0 - _ratio4) * 5_000_000)

                                        # ── 連班型態感知：避免上一休一 ──
                                        _y4 = sched[idx][d_int - 1] if d_int > 1 else ""
                                        _t4 = sched[idx][d_int + 1] if d_int < month_days else ""
                                        if is_work(_y4) and is_work(_t4):
                                            score += 2_000_000   # 填補孤立休假空隙 → W-W-W（最優）
                                        elif is_work(_y4) or is_work(_t4):
                                            score += 500_000     # 延伸既有連班
                                        else:
                                            score -= 2_000_000   # 兩側皆休，形成孤立班（最差）

                                        # 💡 階梯式控台權重
                                        l_str_idx = cache_leader_str[idx]
                                        is_leader_for_shift = False
                                        if "白" in l_str_idx: is_leader_for_shift = True
                                        elif "小" in l_str_idx and s_type in ["E", "N"]: is_leader_for_shift = True
                                        elif "大" in l_str_idx and s_type == "N": is_leader_for_shift = True

                                        if not has_leader and is_leader_for_shift: score += 50000000
                                        if cache_circ[idx] and curr_circ < target_circ: score += 10000000

                                        # ── 精確預休接軌加分：緩衝日是否浪費可排班位 ──
                                        # 緩衝日（d+1）若本來就是休/週末/預休 → 不浪費空格 → 加分（完美接軌）
                                        # 緩衝日（d+1）若是一般工作日        → 損失1個可排班位 → 懲罰（浪費日）
                                        _next_d4 = d_int + 1
                                        _buf_is_natural_rest4 = (
                                            _next_d4 in cache_pre_off4.get(idx, set())  # 個人預休 O
                                            or _next_d4 in set(sat_list4)               # 週六
                                            or _next_d4 in set(sun_list4)               # 週日
                                            or _next_d4 in set(nat_list4)               # 國定假日
                                        )
                                        if _buf_is_natural_rest4:
                                            score += 8_000_000   # 完美接軌：緩衝日完全不浪費
                                        else:
                                            score -= 3_000_000   # 緩衝日浪費1個工作日空格，懲罰

                                        # ── Block Preference：同班別聚類加分（跳過緩衝日）+ 異種夜班懲罰 ──
                                        # N 班後有強制緩衝日，直接看前後各 2 格才能偵測到 N-N 連排
                                        _night_types4 = ("E", "N", "12-8")
                                        _prev1_b4 = sched[idx][d_int - 1] if d_int > 1 else ""
                                        _prev2_b4 = sched[idx][d_int - 2] if d_int > 2 else ""
                                        _next1_b4 = sched[idx][d_int + 1] if d_int < month_days else ""
                                        _next2_b4 = sched[idx][d_int + 2] if d_int + 2 <= month_days else ""
                                        _block_direct = (_prev1_b4 == s_type or _next1_b4 == s_type)
                                        _block_skip   = (_prev2_b4 == s_type or _next2_b4 == s_type)
                                        if _block_direct:
                                            score += 4_000_000   # 直接相鄰同班別（E/12-8 連排）：最優
                                        elif _block_skip:
                                            score += 3_000_000   # 跳過緩衝日的同班別（N-N 連排模式）
                                        # 異種夜班相鄰懲罰（前後各 2 格內有不同夜班類型）
                                        _near4 = [_prev1_b4, _prev2_b4, _next1_b4, _next2_b4]
                                        if any(v in _night_types4 and v != s_type for v in _near4):
                                            score -= 2_500_000   # 異種夜班出現在前後 2 格，懲罰

                                        # ── 剩餘容量預判：排入此夜班後，後續有效空格能否補滿應上班天數 ──
                                        # 保守估計：夜班後至少損失 1 個強制緩衝日，有效空格再減 1
                                        _worked_now4 = sum(1 for x in sched[idx] if is_work(x))
                                        _still_need4 = personal_targets.get(idx, 0) - _worked_now4 - 1  # 排入本班後還需幾天
                                        _empty_after4 = sum(
                                            1 for _d4 in range(d_int + 2, month_days + 1)  # d_int+1 為緩衝日，從 d+2 起算
                                            if sched[idx][_d4] == ""
                                        )
                                        if _still_need4 > 0 and _empty_after4 < _still_need4:
                                            score -= 8_000_000  # 空格不足，大幅懲罰（但不硬阻，保留均分彈性）

                                        # ── 連五影子懲罰：預防夜班落在D班長連段尾端 ──────────
                                        # 邏輯：夜班左側若有>=4個連續可工作日（空格視為潛在D班），
                                        #       Step 5 補D班後最外圍那天會被連五擋死，損失1個補班空格
                                        # 以左側連段長度（含夜班本身）作為評估依據
                                        _hol_set4_cp = set(sat_list4) | set(sun_list4) | set(nat_list4)
                                        _left4_cp = 0
                                        for _bd4_cp in range(d_int - 1, 0, -1):
                                            _v4_cp = sched[idx][_bd4_cp]
                                            if _bd4_cp in _hol_set4_cp: break          # 週末：天然斷點
                                            if _v4_cp not in ("", "上課") and not is_work(_v4_cp): break  # 特殊假別
                                            _left4_cp += 1
                                        _chain4_cp = _left4_cp + 1  # 含夜班本身
                                        if _chain4_cp >= 5:
                                            score -= (_chain4_cp - 4) * 4_000_000  # 每多1天：-400萬懲罰
                                        # 右側連段預判（緩衝日d+1後的潛在D班長段）
                                        _right4_cp = 0
                                        for _fd4_cp in range(d_int + 2, month_days + 1):  # 跳過強制緩衝日
                                            _v4_cp = sched[idx][_fd4_cp]
                                            if _fd4_cp in _hol_set4_cp: break
                                            if _v4_cp not in ("", "上課") and not is_work(_v4_cp): break
                                            _right4_cp += 1
                                        if _right4_cp >= 5:
                                            score -= (_right4_cp - 4) * 2_000_000  # 右側風險較低，懲罰減半

                                        # 連五上限懲罰（超過 4 連班才懲罰）
                                        _sc = 1
                                        for _bd in range(d_int - 1, 0, -1):
                                            if is_work(sched[idx][_bd]): _sc += 1
                                            else: break
                                        for _fd in range(d_int + 1, month_days + 1):
                                            if is_work(sched[idx][_fd]): _sc += 1
                                            else: break
                                        if _sc >= 4: score -= (_sc - 3) * 2_000_000

                                        return score + random.random()
                                        
                                    if _block_nurse is not None:
                                        best_nurse = _block_nurse
                                    else:
                                        best_nurse = max(available, key=evaluate_nurse)
                                    # pass_num=False 時 can_work_base 不查 week_variety，此處補驗
                                    if not week_variety_ok(sched, best_nurse, s_type, d_int,
                                                           st.session_state.first_wday, month_days):
                                        progress = True  # 避免死迴圈
                                        break
                                    sched[best_nurse][d_int] = s_type
                                    # 更新塊追蹤器
                                    if _block_remaining[s_type].get(best_nurse, 0) > 0:
                                        _block_remaining[s_type][best_nurse] -= 1
                                    else:
                                        _block_remaining[s_type][best_nurse] = _MIN_BLOCK - 1
                                    progress = True
                                    break
                                if not progress: break  # 本輪無任何進展，提前結束

                    for s_t in ["N", "E"]:
                        assign_night_shifts(s_t)

                    # ── E/N 事後均等化（保證均等池內夜班差距 ≤ 1）─────────────
                    # 策略一（優先）：四格互換 — over[X]=E/N→"", under[X]=""→E/N,
                    #                            under[Y]=D→"",  over[Y]=""→D
                    # 策略二（備用）：單格轉讓 — 僅在 under 仍低於 personal_targets 時使用
                    _il4_eq = {"D": ["N"], "E": ["D","N","12-8"], "12-8": ["N"], "N": []}

                    def _can_en_nocheck4(n_idx, s, d_int):
                        """d_int 可合法排 E 或 N（不含 personal_targets 上限，適用互換）"""
                        if sched[n_idx][d_int] not in ["", "上課"]: return False
                        if cache_preg[n_idx] and s in ("E", "N"): return False
                        qual4 = cache_night[n_idx]
                        if s == "N" and qual4 != "大夜": return False
                        if s == "E" and qual4 not in ("大夜", "小夜"): return False
                        if not can_work_holiday_check(n_idx, d_int, cache_can_sat4, cache_can_sun4, cache_can_nat4, sat_list4, sun_list4, nat_list4): return False
                        _y4e = sched[n_idx][d_int - 1] if d_int > 1 else ""
                        _t4e = sched[n_idx][d_int + 1] if d_int < month_days else ""
                        _yb4e = "D" if (_y4e.startswith("D") or _y4e in ("上課","公差")) else _y4e
                        _tb4e = "D" if (_t4e.startswith("D") or _t4e in ("上課","公差")) else _t4e
                        if is_work(_y4e) and s in _il4_eq.get(_yb4e, []): return False
                        if is_work(_t4e) and _tb4e in _il4_eq.get(s, []): return False
                        _sc4e = 1
                        for _bd4 in range(d_int - 1, 0, -1):
                            if is_work(sched[n_idx][_bd4]): _sc4e += 1
                            else: break
                        for _fd4 in range(d_int + 1, month_days + 1):
                            if is_work(sched[n_idx][_fd4]): _sc4e += 1
                            else: break
                        if _sc4e > 5: return False
                        return True

                    def _can_D_nocheck4(n_idx, d_int):
                        """d_int 可合法排 D（不含 personal_targets 上限，適用互換）"""
                        if sched[n_idx][d_int] not in ["", "上課"]: return False
                        _y4d = sched[n_idx][d_int - 1] if d_int > 1 else ""
                        _t4d = sched[n_idx][d_int + 1] if d_int < month_days else ""
                        _yb4d = "D" if (_y4d.startswith("D") or _y4d in ("上課","公差")) else _y4d
                        _tb4d = "D" if (_t4d.startswith("D") or _t4d in ("上課","公差")) else _t4d
                        if is_work(_y4d) and "D" in _il4_eq.get(_yb4d, []): return False
                        if is_work(_t4d) and _tb4d in _il4_eq.get("D", []): return False
                        _sc4d = 1
                        for _bd4 in range(d_int - 1, 0, -1):
                            if is_work(sched[n_idx][_bd4]): _sc4d += 1
                            else: break
                        for _fd4 in range(d_int + 1, month_days + 1):
                            if is_work(sched[n_idx][_fd4]): _sc4d += 1
                            else: break
                        if _sc4d > 5: return False
                        return True

                    _en_elig_set4 = set(elig_night_nurses)  # 非包班、非孕育嬰、具大/小夜資格

                    for _nit4 in range(500):
                        _nc4 = {i: sum(1 for v in sched[i] if v in ("E", "N"))
                                for i in _en_elig_set4}
                        if not _nc4: break
                        _nmax4 = max(_nc4.values())
                        _nmin4 = min(_nc4.values())
                        if _nmax4 - _nmin4 <= 1: break

                        _over_l4  = [i for i, c in _nc4.items() if c == _nmax4]
                        _under_l4 = [i for i, c in _nc4.items() if c == _nmin4]

                        _swapped4 = False
                        for _ov4 in _over_l4:
                            if _swapped4: break
                            for _un4 in _under_l4:
                                if _swapped4: break
                                for _d4 in range(1, month_days + 1):
                                    if _swapped4: break
                                    _ov_shift4 = sched[_ov4][_d4]
                                    if _ov_shift4 not in ("E", "N"): continue
                                    # ★ 保護：over 的班若為鎖定格（國定必上班別/公差/預白），不可移走
                                    if (_ov4, _d4) in _locked_set4: continue
                                    if sched[_un4][_d4] not in ("", "上課"): continue
                                    # ★ 保護：under 的目標格若為鎖定格（上課日等），不可填入 E/N
                                    if (_un4, _d4) in _locked_set4: continue
                                    if not _can_en_nocheck4(_un4, _ov_shift4, _d4): continue

                                    # ── 策略一：四格互換（兩人總班數不變）──────
                                    _four4 = False
                                    for _wd4 in range(1, month_days + 1):
                                        if _wd4 == _d4: continue
                                        if sched[_un4][_wd4] != "D": continue
                                        # ★ 保護：under 的平日班若為鎖定格，不可清除
                                        if (_un4, _wd4) in _locked_set4: continue
                                        if not _can_D_nocheck4(_ov4, _wd4): continue
                                        sched[_ov4][_d4] = ""
                                        sched[_un4][_d4] = _ov_shift4
                                        sched[_un4][_wd4] = ""
                                        sched[_ov4][_wd4] = "D"
                                        _swapped4 = True
                                        _four4 = True
                                        break

                                    # ── 策略二：單格轉讓（under 仍低於目標時備用）──
                                    if not _four4:
                                        if sum(1 for x in sched[_un4] if is_work(x)) < personal_targets.get(_un4, 0):
                                            sched[_ov4][_d4] = ""
                                            sched[_un4][_d4] = _ov_shift4
                                            _swapped4 = True

                        if not _swapped4:
                            break  # 找不到可交換組合，停止

                    # ── 12-8 中班排入（E/N 均等化後接著排）──────────────────────────────────
                    # 資格：大夜/小夜/中班 夜班資格，或母性保護；排除行政職稱
                    def can_work_12_8_s4(n_idx, d_int):
                        if sched[n_idx][d_int] not in ["", "上課"]: return False
                        if str(d_int) in class_days_map.get(n_idx, []): return False  # 上課日不可排 12-8
                        if cache_pref[n_idx] == "" and cache_night[n_idx] not in ("大夜", "小夜", "中班") and not cache_preg[n_idx]: return False
                        if cache_pref[n_idx] == "" and not can_work_holiday_check(n_idx, d_int, cache_can_sat4, cache_can_sun4, cache_can_nat4, sat_list4, sun_list4, nat_list4): return False
                        worked = sum(1 for x in sched[n_idx] if is_work(x))
                        if worked >= personal_targets[n_idx]: return False
                        y_s = (sched[n_idx][d_int - 1] or "") if d_int > 1 else ""
                        t_s = (sched[n_idx][d_int + 1] or "") if d_int < month_days else ""
                        y_s_base = "D" if (y_s.startswith("D") or y_s in ("上課", "公差")) else y_s
                        t_s_base = "D" if (t_s.startswith("D") or t_s in ("上課", "公差")) else t_s
                        if is_work(y_s) and "12-8" in illegal_next.get(y_s_base, []): return False
                        if is_work(t_s) and t_s_base in illegal_next.get("12-8", []): return False
                        s_c = 1
                        for bd in range(d_int - 1, 0, -1):
                            if is_work(sched[n_idx][bd]): s_c += 1
                            else: break
                        for fd in range(d_int + 1, month_days + 1):
                            if is_work(sched[n_idx][fd]): s_c += 1
                            else: break
                        if s_c > 5: return False
                        _has_pb4b = bool((st.session_state.get("prev_month_buffer") or {}).get(str(ai_df.at[n_idx,"姓名"]).strip() if n_idx in ai_df.index else "", {}))
                        w_min = (d_int - 13) if _has_pb4b else max(1, d_int - 13)
                        w_max = min(d_int, month_days - 13) if month_days >= 14 else 1
                        for start_d in range(w_min, w_max + 1):
                            end_d = min(month_days, start_d + 13)
                            worked_win = 1
                            for cd in range(start_d, end_d + 1):
                                if cd == d_int: continue
                                if is_work(_xmonth_shift(n_idx, cd, sched, ai_df, month_days)): worked_win += 1
                            if worked_win > 12: return False
                        if not week_variety_ok(sched, n_idx, "12-8", d_int, st.session_state.first_wday, month_days): return False
                        return True

                    # 可排 12-8 人員：常規人員 + 包小夜人員（優先）+ 欠班的 E 包班護師（補充班）
                    # E 包班欠班者加入，讓她們在 Step 4 就能優先搶到 12-8，比 Step 5 更早一步
                    _e_pack_deficit_s4 = {
                        i for i in ai_df.index
                        if "小夜" in cache_pref.get(i, "")
                        and sum(1 for v in sched[i] if is_work(v)) < personal_targets.get(i, 0)
                    }
                    elig_12_8_s4 = [
                        i for i in ai_df.index
                        if cache_title[i] not in ADMIN_TITLES
                        and (cache_pref[i] == "" or "小夜" in cache_pref[i] or i in _e_pack_deficit_s4)
                    ]
                    total_12_8_demand_s4 = sum(
                        int(edited_quota_df[edited_quota_df["日期"] == str(d)].iloc[0]["12-8"])
                        for d in range(1, month_days + 1)
                    )
                    pack_12_8_supply_s4 = sum(
                        sum(1 for v in sched[i] if v == "12-8")
                        for i in ai_df.index if cache_pref[i] != "" and "小夜" not in cache_pref[i]
                    )
                    remaining_12_8_s4 = max(0, total_12_8_demand_s4 - pack_12_8_supply_s4)
                    target_12_8_s4 = remaining_12_8_s4 // len(elig_12_8_s4) if elig_12_8_s4 else 0

                    # E+N+12-8 夜班均等池（差距 ≤ 1）
                    # 排除：包班意願不為空、職稱為「組長」；必須具夜班資格
                    elig_night_s4 = [
                        i for i in ai_df.index
                        if cache_pref[i] == ""
                        and cache_title[i] != "組長"
                        and cache_night[i] != ""
                    ]
                    _already_night_s4 = sum(
                        sum(1 for v in sched[i] if v in ("E", "N", "12-8"))
                        for i in elig_night_s4
                    )
                    target_night_s4 = (
                        (_already_night_s4 + remaining_12_8_s4) // len(elig_night_s4)
                        if elig_night_s4 else 0
                    )

                    def assign_12_8_shifts_s4():
                        for pass_num in [True, False]:
                            for _iter in range(month_days * 10):
                                progress = False
                                deficits = []
                                for d in range(1, month_days + 1):
                                    req = int(edited_quota_df[edited_quota_df["日期"] == str(d)].iloc[0]["12-8"])
                                    curr = sum(1 for i in ai_df.index if sched[i][d] == "12-8")
                                    if req > curr:
                                        leave_count = sum(1 for i in ai_df.index if not is_work(sched[i][d]) and sched[i][d] != "")
                                        deficits.append((d, req - curr, leave_count))
                                if not deficits: break
                                deficits.sort(key=lambda x: (x[1], x[2], random.random()), reverse=True)
                                for d_int, defc, _ in deficits:
                                    curr_nurses_12 = [i for i in ai_df.index if sched[i][d_int] == "12-8"]
                                    curr_circ_12   = sum(1 for i in curr_nurses_12 if cache_circ[i])
                                    day_q_12       = int(edited_quota_df[edited_quota_df["日期"] == str(d_int)].iloc[0]["12-8"])
                                    target_circ_12 = (day_q_12 + 1) // 2 if day_q_12 > 0 else 0
                                    available = [i for i in elig_12_8_s4 if can_work_12_8_s4(i, d_int) and group_cap_ok(i, "12-8", d_int, sched, cache_group4)]
                                    if not available: continue
                                    def score_12_8_s4(idx):
                                        night_count = sum(1 for v in sched[idx] if v in ("E", "N", "12-8"))
                                        score = 0
                                        if "小夜" in cache_pref[idx]: score += 100_000_000  # 包小夜優先
                                        # 欠班的 E 包班護師：極高優先（僅次於包小夜本人），讓她們先搶到 12-8 補充班
                                        if idx in _e_pack_deficit_s4:
                                            _deficit_days = personal_targets.get(idx, 0) - sum(1 for v in sched[idx] if is_work(v))
                                            score += 80_000_000 + _deficit_days * 1_000_000  # 欠越多越優先
                                        if idx in elig_night_s4:
                                            score += (target_night_s4 - night_count) * 5_000_000
                                        else:
                                            count_12_8 = sum(1 for v in sched[idx] if v == "12-8")
                                            score += (target_12_8_s4 - count_12_8) * 5_000_000
                                        if cache_circ[idx] and curr_circ_12 < target_circ_12: score += 10_000_000
                                        _y4a = sched[idx][d_int - 1] if d_int > 1 else ""
                                        _t4a = sched[idx][d_int + 1] if d_int < month_days else ""
                                        if is_work(_y4a) and is_work(_t4a): score += 2_000_000
                                        elif is_work(_y4a) or is_work(_t4a): score += 500_000
                                        else: score -= 2_000_000
                                        _sc = 1
                                        for _bd in range(d_int - 1, 0, -1):
                                            if is_work(sched[idx][_bd]): _sc += 1
                                            else: break
                                        for _fd in range(d_int + 1, month_days + 1):
                                            if is_work(sched[idx][_fd]): _sc += 1
                                            else: break
                                        if _sc >= 4: score -= (_sc - 3) * 2_000_000
                                        return score + random.random()
                                    best = max(available, key=score_12_8_s4)
                                    sched[best][d_int] = "12-8"
                                    progress = True
                                    break
                                if not progress: break

                    assign_12_8_shifts_s4()

                    # ── 12-8 事後均等化（E+N+12-8 合計差距 ≤ 1）────────────────────────────
                    # 策略一（優先）：四格互換 — over[X]: 12-8→"", over[Y]: ""→D
                    #                            under[X]: ""→12-8, under[Y]: D→""
                    # 策略二（備用）：單格轉讓 — 僅在 under 仍低於 personal_targets 時使用
                    def _can_12_8_nocheck4b(n_idx, d_int):
                        """d_int 可合法排 12-8（不含 personal_targets 上限，適用均等化互換）"""
                        if sched[n_idx][d_int] not in ["", "上課"]: return False
                        if cache_pref[n_idx] == "" and cache_night[n_idx] not in ("大夜", "小夜", "中班") and not cache_preg[n_idx]: return False
                        if cache_pref[n_idx] == "" and not can_work_holiday_check(n_idx, d_int, cache_can_sat4, cache_can_sun4, cache_can_nat4, sat_list4, sun_list4, nat_list4): return False
                        y_s = (sched[n_idx][d_int - 1] or "") if d_int > 1 else ""
                        t_s = (sched[n_idx][d_int + 1] or "") if d_int < month_days else ""
                        y_b = "D" if (y_s.startswith("D") or y_s in ("上課", "公差")) else y_s
                        t_b = "D" if (t_s.startswith("D") or t_s in ("上課", "公差")) else t_s
                        if is_work(y_s) and "12-8" in _il4_eq.get(y_b, []): return False
                        if is_work(t_s) and t_b in _il4_eq.get("12-8", []): return False
                        s_c = 1
                        for bd in range(d_int - 1, 0, -1):
                            if is_work(sched[n_idx][bd]): s_c += 1
                            else: break
                        for fd in range(d_int + 1, month_days + 1):
                            if is_work(sched[n_idx][fd]): s_c += 1
                            else: break
                        if s_c > 5: return False
                        return True

                    _night_elig_set4b = set(elig_night_s4)
                    for _nit4b in range(500):
                        _nc4b = {i: sum(1 for v in sched[i] if v in ("E", "N", "12-8"))
                                 for i in _night_elig_set4b}
                        if not _nc4b: break
                        _nmax4b = max(_nc4b.values())
                        _nmin4b = min(_nc4b.values())
                        if _nmax4b - _nmin4b <= 1: break

                        _over_l4b  = [i for i, c in _nc4b.items() if c == _nmax4b]
                        _under_l4b = [i for i, c in _nc4b.items() if c == _nmin4b]

                        _swapped4b = False
                        for _ov4b in _over_l4b:
                            if _swapped4b: break
                            for _un4b in _under_l4b:
                                if _swapped4b: break
                                for _d4b in range(1, month_days + 1):
                                    if _swapped4b: break
                                    if sched[_ov4b][_d4b] != "12-8": continue
                                    if (_ov4b, _d4b) in _locked_set4: continue
                                    if sched[_un4b][_d4b] not in ["", "上課"]: continue
                                    if (_un4b, _d4b) in _locked_set4: continue
                                    if not _can_12_8_nocheck4b(_un4b, _d4b): continue

                                    _four4b = False
                                    for _wd4b in range(1, month_days + 1):
                                        if _wd4b == _d4b: continue
                                        if sched[_un4b][_wd4b] != "D": continue
                                        if (_un4b, _wd4b) in _locked_set4: continue
                                        if not _can_D_nocheck4(_ov4b, _wd4b): continue
                                        if (_ov4b, _d4b) in _locked_set4: continue
                                        sched[_ov4b][_d4b] = ""
                                        sched[_un4b][_d4b] = "12-8"
                                        sched[_un4b][_wd4b] = ""
                                        sched[_ov4b][_wd4b] = "D"
                                        _swapped4b = True
                                        _four4b = True
                                        break

                                    if not _four4b:
                                        _un4b_worked = sum(1 for x in sched[_un4b] if is_work(x))
                                        if _un4b_worked < personal_targets.get(_un4b, 0):
                                            sched[_ov4b][_d4b] = ""
                                            sched[_un4b][_d4b] = "12-8"
                                            _swapped4b = True

                        if not _swapped4b:
                            break  # 找不到可交換組合，停止

                    # ── 統一後處理均等化（E+N+12-8 全班種可互換，差距 ≤ 1）──────────────
                    # 對象：elig_night_s4（大夜＋小夜＋中班，非包班，非組長）
                    # 邏輯：找出 E/N/12-8 合計最多(over)與最少(under)的護理師，
                    #       嘗試把 over 在某天的夜班直接交給 under（含鄰班/連五/資格驗證）
                    # 策略一（優先）：四格互換 over[X]→"", over[Y]→D; under[X]→夜班, under[Y]→""
                    # 策略二（備用）：單格轉讓（under 仍未達 personal_targets 時）

                    def _can_any_night_unified4(n_idx, s, d_int):
                        """d_int 能合法排入 s（E/N/12-8），供統一均等化互換使用"""
                        if sched[n_idx][d_int] not in ["", "上課"]: return False
                        if cache_preg[n_idx] and s in ("E", "N"): return False
                        qual4u = cache_night[n_idx]
                        if s == "N" and qual4u != "大夜": return False
                        if s == "E" and qual4u not in ("大夜", "小夜"): return False
                        if s == "12-8" and qual4u not in ("大夜", "小夜", "中班") and not cache_preg[n_idx]: return False
                        if cache_pref[n_idx] == "" and not can_work_holiday_check(
                                n_idx, d_int, cache_can_sat4, cache_can_sun4, cache_can_nat4,
                                sat_list4, sun_list4, nat_list4): return False
                        _yu = (sched[n_idx][d_int - 1] or "") if d_int > 1 else ""
                        _tu = (sched[n_idx][d_int + 1] or "") if d_int < month_days else ""
                        _ybu = "D" if (_yu.startswith("D") or _yu in ("上課", "公差")) else _yu
                        _tbu = "D" if (_tu.startswith("D") or _tu in ("上課", "公差")) else _tu
                        if is_work(_yu) and s in _il4_eq.get(_ybu, []): return False
                        if is_work(_tu) and _tbu in _il4_eq.get(s, []): return False
                        _scu = 1
                        for _bdu in range(d_int - 1, 0, -1):
                            if is_work(sched[n_idx][_bdu]): _scu += 1
                            else: break
                        for _fdu in range(d_int + 1, month_days + 1):
                            if is_work(sched[n_idx][_fdu]): _scu += 1
                            else: break
                        if _scu > 5: return False
                        return True

                    _unified_elig_set4 = set(elig_night_s4)
                    for _nitu in range(800):
                        _ncu = {i: sum(1 for v in sched[i] if v in ("E", "N", "12-8"))
                                for i in _unified_elig_set4}
                        if not _ncu: break
                        _nmaxu = max(_ncu.values())
                        _nminu = min(_ncu.values())
                        if _nmaxu - _nminu <= 1: break

                        _over_lu  = [i for i, c in _ncu.items() if c == _nmaxu]
                        _under_lu = [i for i, c in _ncu.items() if c == _nminu]

                        _swappedu = False
                        for _ovu in _over_lu:
                            if _swappedu: break
                            for _unu in _under_lu:
                                if _swappedu: break
                                for _du in range(1, month_days + 1):
                                    if _swappedu: break
                                    _svu = sched[_ovu][_du]
                                    if _svu not in ("E", "N", "12-8"): continue
                                    if (_ovu, _du) in _locked_set4: continue
                                    if sched[_unu][_du] not in ["", "上課"]: continue
                                    if (_unu, _du) in _locked_set4: continue
                                    if not _can_any_night_unified4(_unu, _svu, _du): continue

                                    _fouru = False
                                    for _wdu in range(1, month_days + 1):
                                        if _wdu == _du: continue
                                        if sched[_unu][_wdu] != "D": continue
                                        if (_unu, _wdu) in _locked_set4: continue
                                        if not _can_D_nocheck4(_ovu, _wdu): continue
                                        sched[_ovu][_du]  = ""
                                        sched[_unu][_du]  = _svu
                                        sched[_unu][_wdu] = ""
                                        sched[_ovu][_wdu] = "D"
                                        _swappedu = True
                                        _fouru = True
                                        break

                                    if not _fouru:
                                        _unu_worked = sum(1 for x in sched[_unu] if is_work(x))
                                        if _unu_worked < personal_targets.get(_unu, 0):
                                            sched[_ovu][_du] = ""
                                            sched[_unu][_du] = _svu
                                            _swappedu = True
                                            break

                        if not _swappedu:
                            break  # 無法再縮小差距，停止

                    night_df = pd.DataFrame({"姓名": ai_df["姓名"]})
                    for d in range(1, month_days + 1):
                        night_df[str(d)] = [sched[i][d] for i in ai_df.index]
                        
                    st.session_state.night_sched = night_df
                    st.rerun()

    if st.session_state.night_sched is not None:
        st.success("✅ 常規夜班 (E / N / 12-8) 已全數排入完畢！(具備流動與階梯控台防護)")
        
        elig_night_nurses = [
            i for i in ai_df.index
            if str(ai_df.loc[i, "包班意願"]).strip() == ""
            and (str(ai_df.loc[i, "夜班資格"]).strip() in ("大夜", "小夜", "中班")
                 or str(ai_df.loc[i, "孕/育嬰免夜班"]).strip() == "是")
            and str(ai_df.loc[i, "職稱"]).strip() not in ADMIN_TITLES
        ]

        night_stats = []
        for i in elig_night_nurses:
            nurse_name = ai_df.loc[i, "姓名"]
            s_vals = list(st.session_state.night_sched.iloc[i, 1:].values)
            n_count = s_vals.count("N")
            e_count = s_vals.count("E")
            m_count = s_vals.count("12-8")
            night_stats.append({
                "姓名": nurse_name,
                "E班(小夜)": e_count,
                "N班(大夜)": n_count,
                "12-8(中班)": m_count,
                "夜班總計(E+N+12-8)": n_count + e_count + m_count
            })

        st.info("📊 **夜班(E/N/12-8)分配明細**：以下為本月參與輪替人員的各別夜班獲派數量 (不含包班)。")
        st.dataframe(pd.DataFrame(night_stats).sort_values(by="夜班總計(E+N+12-8)", ascending=False), use_container_width=False)
        
        with st.expander("📄 點擊展開含夜班之排班結果", expanded=True):
            _edit_night = st.checkbox("🖊️ 開啟手動編輯模式", value=False, key="chk_edit_night_sched")
            if _edit_night:
                st.caption("💡 直接點擊儲存格輸入班別代碼，修改即時儲存")
                _edited_night = st.data_editor(
                    st.session_state.night_sched,
                    column_config=make_sched_col_config(month_days),
                    use_container_width=True, hide_index=True, key="edit_night_sched"
                )
                st.session_state.night_sched = _edited_night
            else:
                _day_cols_n = [str(i) for i in range(1, month_days + 1)]
                _disp_night = st.session_state.night_sched.copy()
                for _c in _day_cols_n:
                    _disp_night[_c] = _disp_night[_c].apply(abbrev_display)
                # 預白班：D 格標為 Dx
                _disp_night = apply_prewhite_dx(_disp_night, st.session_state.ai_df, month_days)
                st.dataframe(
                    build_schedule_with_counts(_disp_night, st.session_state.night_sched, _day_cols_n, st.session_state.ai_df),
                    use_container_width=True
                )

        display_safety_radar(st.session_state.night_sched, edited_quota_df, st.session_state.ai_df)

        # ── 假日人力覆蓋率摘要 ───────────────────────────────────────────────────
        st.write("#### 📅 假日人力覆蓋率摘要")
        _hol_sat4  = st.session_state.saturdays_list
        _hol_sun4  = st.session_state.sundays_list
        _hol_nat4  = st.session_state.nat_holidays_list
        _hol_all4  = sorted(set(_hol_sat4) | set(_hol_sun4) | set(_hol_nat4))
        _fw4       = st.session_state.first_wday          # 0=Mon … 6=Sun
        _wday_lbl  = ["一", "二", "三", "四", "五", "六", "日"]
        _night_df4 = st.session_state.night_sched

        # 建立職稱快取（用於排除 NO_HOL_ADMIN）
        _title_cache4 = {i: str(ai_df.loc[i, "職稱"]).strip() for i in ai_df.index}

        _hol_rows4 = []
        _has_critical4 = False
        for _hd4 in _hol_all4:
            _hd_str4 = str(_hd4)
            # 當日 D班配額
            _quota_d4 = int(edited_quota_df.iloc[_hd4 - 1]["D班"]) if _hd4 <= len(edited_quota_df) else 0
            # 計算已有班別人數（排除 NO_HOL_ADMIN）
            _worked_hd4 = 0
            if _hd_str4 in _night_df4.columns:
                for _i4 in ai_df.index:
                    if _title_cache4[_i4] in NO_HOL_ADMIN:
                        continue
                    _v4 = str(_night_df4.loc[_i4, _hd_str4] if _i4 in _night_df4.index else "").strip()
                    if is_work(_v4):
                        _worked_hd4 += 1
            # 星期別
            _wd4 = (_fw4 + _hd4 - 1) % 7
            if _hd4 in set(_hol_nat4) and _hd4 not in set(_hol_sat4) | set(_hol_sun4):
                _wlabel4 = f"國定({_wday_lbl[_wd4]})"
            elif _hd4 in set(_hol_sat4):
                _wlabel4 = "週六"
            else:
                _wlabel4 = "週日"
            # 覆蓋率
            _cov4 = (_worked_hd4 / _quota_d4 * 100) if _quota_d4 > 0 else 0.0
            if _cov4 < 60:
                _status4 = "🚨 嚴重不足"
                _has_critical4 = True
            elif _cov4 < 100:
                _status4 = "⚠️ 部分不足"
            else:
                _status4 = "✅ 已達標"
            _hol_rows4.append({
                "日期": _hd4,
                "星期別": _wlabel4,
                "D班配額": _quota_d4,
                "已排人數": _worked_hd4,
                "覆蓋率": f"{_cov4:.0f}%",
                "狀態": _status4,
            })

        st.dataframe(pd.DataFrame(_hol_rows4), use_container_width=False, hide_index=True)
        if _has_critical4:
            st.warning("⚠️ 以上假日人力不足，建議至護理師名單調整『能上週六/週日/國定假日』欄位後重排。")

        if st.session_state.step == 4:
            col_btn_back, col_btn_go = st.columns([1, 4])
            with col_btn_back:
                if st.button("⬅️ 重新安插夜班", type="secondary"):
                    st.session_state.night_sched = None
                    st.rerun()
            with col_btn_go:
                if st.button("✅ 確認夜班無誤，前往第五步（排滿白班）", type="primary"):
                    st.session_state.step = 5
                    st.rerun()

# ==========================================
# 📍 第五步：排滿白班
# ==========================================
if st.session_state.step >= 5:
    st.divider()
    st.header("5️⃣ 第五步：自動排滿白班")

    ai_df = st.session_state.ai_df
    month_days = st.session_state.month_days
    edited_quota_df = st.session_state.edited_quota_df
    edited_weekly_df = st.session_state.edited_weekly_df
    holiday_days_list = st.session_state.holiday_list

    with st.expander("⚙️ 調整每日人力配額（修改後需重新排班才生效）", expanded=False):
        _q5 = st.data_editor(st.session_state.edited_quota_df, hide_index=True, use_container_width=True, key="quota_editor_step5")
        if st.button("✅ 套用配額變更並重排（從第五步重算）", key="apply_quota_step5"):
            st.session_state.edited_quota_df = _q5
            st.session_state.d_sched = None
            st.session_state.final_sched = None
            st.rerun()
    edited_quota_df = st.session_state.edited_quota_df
    
    if st.session_state.d_sched is None:
        with st.spinner("正在排入白班，執行缺額救援運算..."):
            sched_df = st.session_state.night_sched.copy()
            sched = {i: [""] + list(sched_df.iloc[i, 1:].values) for i in range(len(ai_df))}
            
            cache_skills = {i: [x.strip() for x in str(row.get("次專科能力", "")).split(",")] for i, row in ai_df.iterrows()}
            cache_circ = {i: str(row.get("流動資格", "")).strip() == "是" for i, row in ai_df.iterrows()}
            cache_title = {i: str(row.get("職稱", "")).strip() for i, row in ai_df.iterrows()}
            cache_pref = {i: str(row.get("包班意願", "")).strip() for i, row in ai_df.iterrows()}
            cache_preg5 = {i: str(row.get("孕/育嬰免夜班", "")).strip() == "是" for i, row in ai_df.iterrows()}
            cache_night5 = {i: str(row.get("夜班資格", "")).strip() for i, row in ai_df.iterrows()}  # 空白 = 無夜班資格
            cache_leader_str = {i: str(row.get("控台資格", "")).strip() for i, row in ai_df.iterrows()}
            class_days_map = {i: [s.strip() for s in str(row.get("上課日期", "")).split(",") if s.strip().isdigit()] for i, row in ai_df.iterrows()}
            # 假日出勤能力快取（第五步）
            cache_can_sat5 = {i: str(row.get("能上週六", "")).strip() == "是" for i, row in ai_df.iterrows()}
            cache_can_sun5 = {i: str(row.get("能上週日", "")).strip() == "是" for i, row in ai_df.iterrows()}
            cache_can_nat5 = {i: str(row.get("能上國定假日", "")).strip() == "是" for i, row in ai_df.iterrows()}
            cache_group5   = {i: str(row.get("組別", "")).strip().upper() for i, row in ai_df.iterrows()}
            sat_list5 = st.session_state.saturdays_list
            sun_list5 = st.session_state.sundays_list
            nat_list5 = st.session_state.nat_holidays_list

            personal_targets = {}
            _hol_set5 = set(sat_list5) | set(sun_list5) | set(nat_list5)
            _weekday_count5 = sum(1 for d in range(1, month_days + 1) if d not in _hol_set5)
            _toff5 = st.session_state.target_off
            for idx, row in ai_df.iterrows():
                title_idx = str(row.get("職稱", "")).strip()
                if title_idx in NO_HOL_SET:
                    # 護理長/副護理長/助理/傷兵/組長：不排假日班，預休/長假/特假均直接扣平日
                    # target_off=0 → 所有平日假別（O+長假+特假）都算扣減
                    extra_leaves = calc_extra_leaves(row, month_days, sat_list5, sun_list5, nat_list5, target_off=0)
                    personal_targets[idx] = max(0, _weekday_count5 - extra_leaves)
                else:
                    # 一般護理師：O+長假先佔用 target_off，超出才扣工作天；特假直接扣
                    extra_leaves = calc_extra_leaves(row, month_days, sat_list5, sun_list5, nat_list5, target_off=_toff5)
                    personal_targets[idx] = max(0, month_days - _toff5 - extra_leaves)
            # 套用手動調整的應上班天數
            if st.session_state.custom_targets:
                for idx in ai_df.index:
                    if idx in st.session_state.custom_targets:
                        personal_targets[idx] = st.session_state.custom_targets[idx]

            # ── 預先設定保護集合（預白班 / 公差 / 國定必上班別 / 上課日期）────────────────
            # 這些班次由護理長在 Excel 中預先指定，均等化互換時絕對不可移動
            _prewhite_set5: set = set()   # (idx, day) → 預白班
            _mand_hol_set5: set = set()   # (idx, day) → 國定假日必上班別
            _gongcha_set5:  set = set()   # (idx, day) → 公差（已存為 "公差"，但保留集合供一致性檢查）
            _class_set5:    set = set()   # (idx, day) → 上課日期
            for _pi5, _pr5 in ai_df.iterrows():
                for _dv in str(_pr5.get("預白日期", "")).split(","):
                    if _dv.strip().isdigit() and 1 <= int(_dv.strip()) <= month_days:
                        _prewhite_set5.add((_pi5, int(_dv.strip())))
                for _item in str(_pr5.get("國定假日必上班別", "")).split(","):
                    for _delim in [":", "-"]:
                        if _delim in _item:
                            _dv = _item.split(_delim)[0].strip()
                            if _dv.isdigit() and 1 <= int(_dv) <= month_days:
                                _mand_hol_set5.add((_pi5, int(_dv)))
                            break
                for _dv in str(_pr5.get("公差日期", "")).split(","):
                    if _dv.strip().isdigit() and 1 <= int(_dv.strip()) <= month_days:
                        _gongcha_set5.add((_pi5, int(_dv.strip())))
                for _dv in str(_pr5.get("上課日期", "")).split(","):
                    if _dv.strip().isdigit() and 1 <= int(_dv.strip()) <= month_days:
                        _class_set5.add((_pi5, int(_dv.strip())))
            # 統合保護集合：均等化互換時，這些 (人員, 日期) 組合的班次不可被移動
            _locked_set5 = _prewhite_set5 | _mand_hol_set5 | _gongcha_set5 | _class_set5

            illegal_next = {"D": ["N"], "E": ["D", "N", "12-8"], "12-8": ["N"], "N": []}

            def can_work_base(n_idx, s, d_int, strict_wow=True, week_variety_override=False):
                if sched[n_idx][d_int] not in ["", "上課"]: return False
                # 行政職稱（組長/護理長/副護理長）只能上白班
                if cache_title[n_idx] in ADMIN_TITLES and s != "D": return False
                # 護理長/副護理長/組長：行政班，完全不排假日班
                if cache_title[n_idx] in NO_HOL_SET and d_int in set(sat_list5) | set(sun_list5) | set(nat_list5): return False
                # 假日出勤能力限制（包班人員有假日出勤義務，不受此限）
                if cache_pref[n_idx] == "" and not can_work_holiday_check(n_idx, d_int, cache_can_sat5, cache_can_sun5, cache_can_nat5, sat_list5, sun_list5, nat_list5): return False
                if str(d_int) in class_days_map.get(n_idx, []) and s not in ("D", "N"): return False  # 上課日禁E/12-8

                worked = sum(1 for x in sched[n_idx] if is_work(x))
                if worked >= personal_targets[n_idx]: return False

                y_s = (sched[n_idx][d_int - 1] or "") if d_int > 1 else ""
                t_s = (sched[n_idx][d_int + 1] or "") if d_int < month_days else ""
                y_s_base = "D" if y_s.startswith("D") or y_s in ("上課", "公差") else y_s
                t_s_base = "D" if t_s.startswith("D") or t_s in ("上課", "公差") else t_s

                if is_work(y_s) and s in illegal_next.get(y_s_base, []): return False
                if is_work(t_s) and t_s_base in illegal_next.get(s, []): return False

                if strict_wow:
                    yy_s = sched[n_idx][d_int - 2] if d_int > 2 else "O"
                    tt_s = sched[n_idx][d_int + 2] if d_int < month_days - 1 else "O"
                    if not is_work(y_s) and is_work(yy_s): return False
                    if not is_work(t_s) and is_work(tt_s): return False

                s_consec = 1
                for bd in range(d_int - 1, 0, -1):
                    if is_work(sched[n_idx][bd]): s_consec += 1
                    else: break
                for fd in range(d_int + 1, month_days + 1):
                    if is_work(sched[n_idx][fd]): s_consec += 1
                    else: break
                if s_consec > 5: return False

                _has_pb5 = bool((st.session_state.get("prev_month_buffer") or {}).get(str(ai_df.at[n_idx,"姓名"]).strip() if n_idx in ai_df.index else "", {}))
                _w_min = (d_int - 13) if _has_pb5 else max(1, d_int - 13)
                _w_max = min(d_int, month_days - 13) if month_days >= 14 else 1
                for _sd in range(_w_min, _w_max + 1):
                    _ed = min(month_days, _sd + 13)
                    _ww = 1
                    for _cd in range(_sd, _ed + 1):
                        if _cd == d_int: continue
                        if is_work(_xmonth_shift(n_idx, _cd, sched, ai_df, month_days)): _ww += 1
                    if _ww > 12: return False
                if not week_variety_override and not week_variety_ok(sched, n_idx, s, d_int, st.session_state.first_wday, month_days): return False
                return True

            # ── 排入白班（D）至 personal_targets ──────────────────────────────
            elig_hol_nurses = [i for i in ai_df.index if cache_pref[i] == "" and cache_title[i] not in ADMIN_TITLES]
            total_hol_demand = sum(int(edited_quota_df.iloc[d-1]["D班"]) + int(edited_quota_df.iloc[d-1]["E班"]) + int(edited_quota_df.iloc[d-1]["N班"]) + int(edited_quota_df.iloc[d-1]["12-8"]) for d in holiday_days_list)
            # ── 修正：扣除包班人員已排入假日班次，避免 target_hol 偏高造成一般護理師假日過度排班 ──
            _pack_hol_supply = sum(
                1 for i in ai_df.index if cache_pref[i] != ""
                for hd in holiday_days_list if is_work(sched[i][hd])
            )
            target_hol = max(0, (total_hol_demand - _pack_hol_supply) // len(elig_hol_nurses)) if elig_hol_nurses else 0

            def assign_d_shifts(target_skill=None):
                for pass_num in [True, False]:
                    for _iter in range(month_days * 10):  # 最多迭代 month_days×10 次，防止無限迴圈
                        progress = False
                        deficits = []
                        for d in range(1, month_days + 1):
                            req = int(edited_quota_df[edited_quota_df["日期"] == str(d)].iloc[0]["D班"])
                            if target_skill:
                                curr_w_day = ["週一", "週二", "週三", "週四", "週五", "週六", "週日"][(st.session_state.first_wday + d - 1) % 7]
                                # 週六、週日不設次專科需求，直接視為 0
                                if curr_w_day in ("週六", "週日"):
                                    req = 0
                                else:
                                    _wdf_match = edited_weekly_df[edited_weekly_df["星期"] == curr_w_day] if not edited_weekly_df.empty else pd.DataFrame()
                                    req = int(_wdf_match.iloc[0][f"{target_skill}需求"]) if not _wdf_match.empty else 0
                            
                            curr = sum(1 for i in ai_df.index if sched[i][d] == "D"
                                       and cache_title[i] not in NO_HOL_ADMIN
                                       and (not target_skill or target_skill in cache_skills[i]))
                            if req > curr: deficits.append((d, req - curr))
                        
                        if not deficits: break
                        deficits.sort(key=lambda x: (x[1], random.random()), reverse=True)
                        
                        for d_int, defc in deficits:
                            is_today_holiday = (d_int in holiday_days_list)
                            
                            curr_nurses = [i for i in ai_df.index if sched[i][d_int] == "D" and cache_title[i] not in NO_HOL_ADMIN]
                            curr_circ = sum(1 for i in curr_nurses if cache_circ[i])
                            day_d_quota = int(edited_quota_df[edited_quota_df["日期"] == str(d_int)].iloc[0]["D班"])
                            target_circ = (day_d_quota + 1) // 2 if day_d_quota > 0 else 0
                            
                            has_leader = False
                            for i in curr_nurses:
                                if "白" in cache_leader_str[i]: has_leader = True
                            
                            available = [i for i in ai_df.index if can_work_base(i, "D", d_int, strict_wow=pass_num)]
                            if target_skill: available = [i for i in available if target_skill in cache_skills[i]]
                            available = [i for i in available if cache_pref[i] == ""]
                            if not available: continue
                            
                            def evaluate_nurse(idx):
                                hol_worked = sum(1 for hd in holiday_days_list if is_work(sched[idx][hd]))
                                score = 0

                                if is_today_holiday:
                                    if hol_worked < target_hol: score += 3000000
                                    elif hol_worked == target_hol: score += 1000000
                                    else: score -= (hol_worked * 1000000)

                                # ── 應上班天數達標優先：落後越多越優先（每欠1天加80萬分）──
                                # 目的：防止月初休假較多的護理師（因「孤立班」評分劣勢）持續被跳過，
                                #       導致月末欠班無法補回。
                                _worked_now = sum(1 for v in sched[idx] if is_work(v))
                                _needed_now = personal_targets[idx] - _worked_now
                                if _needed_now > 0:
                                    score += _needed_now * 800_000

                                # ── 連班型態感知：避免上一休一 ──
                                _y5b = sched[idx][d_int - 1] if d_int > 1 else ""
                                _t5b = sched[idx][d_int + 1] if d_int < month_days else ""
                                if is_work(_y5b) and is_work(_t5b):
                                    score += 2_000_000   # 填補孤立休假空隙（最優）
                                elif is_work(_y5b) or is_work(_t5b):
                                    score += 500_000     # 延伸既有連班
                                elif _needed_now >= 2:
                                    pass                 # 欠班≥2天：不懲罰孤立班，讓欠班優先填滿月末空格
                                else:
                                    score -= 2_000_000   # 兩側皆休，孤立班（最差）

                                if not has_leader and "白" in cache_leader_str[idx]:
                                    score += 50000000
                                if cache_circ[idx] and curr_circ < target_circ:
                                    score += 10000000

                                # 連五上限懲罰
                                _sc = 1
                                for _bd in range(d_int - 1, 0, -1):
                                    if is_work(sched[idx][_bd]): _sc += 1
                                    else: break
                                for _fd in range(d_int + 1, month_days + 1):
                                    if is_work(sched[idx][_fd]): _sc += 1
                                    else: break
                                if _sc >= 4: score -= (_sc - 3) * 2_000_000
                                score += group_d_score(idx, d_int, set(sat_list5), set(sun_list5), sched, cache_group5)

                                # ── 前瞻懲罰：若排今天會讓連班數延伸至週末導致週末無法出勤，
                                #    且該週末仍短缺此護理師所屬組別，則給予重懲 ──
                                _grp5_la = cache_group5.get(idx, "")
                                if _grp5_la in ("A", "B"):
                                    _sat5s = set(sat_list5)
                                    _sun5s = set(sun_list5)
                                    # 往後看最多 5 天（連五上限）
                                    for _ld5 in range(d_int + 1, min(d_int + 6, month_days + 1)):
                                        if _ld5 not in (_sat5s | _sun5s):
                                            continue
                                        # 確認該週末仍短缺此組
                                        if _ld5 in _sat5s:
                                            _mn5 = _GROUP_D_SAT_MIN.get(_grp5_la, 0)
                                        else:
                                            _mn5 = _GROUP_D_SUN_MIN.get(_grp5_la, 0)
                                        if _mn5 == 0:
                                            continue
                                        _curr_wknd5 = sum(
                                            1 for _ii in ai_df.index
                                            if cache_group5.get(_ii) == _grp5_la
                                            and isinstance(sched[_ii][_ld5], str)
                                            and sched[_ii][_ld5].startswith("D")
                                        )
                                        if _curr_wknd5 >= _mn5:
                                            continue
                                        # 模擬：排今天後往前連班數
                                        _back5 = 0
                                        for _bd5 in range(d_int - 1, 0, -1):
                                            if is_work(sched[idx][_bd5]): _back5 += 1
                                            else: break
                                        # 若連班 + 今天 + 到週末的距離 > 5，此護理師無法排週末
                                        if _back5 + 1 + (_ld5 - d_int) > 5:
                                            score -= 35_000_000
                                        break  # 只懲罰最近一個有需求的週末

                                return score + random.random()

                            best_nurse = max(available, key=evaluate_nurse)
                            sched[best_nurse][d_int] = "D"
                            progress = True
                            break
                        if not progress: break  # 本輪無任何進展，提前結束

            # ── 週末 A/B 組別預先卡位：在填入平日 D 班前，先確保週末達到組別最低配額 ──
            # 原理：先鎖定週末名額 → 主排班迴圈的連五上限檢查會自動阻止相同護理師
            #       在週末前塞滿連續平日，從根本消除「B組全部休在同一週末」的問題。
            def pre_assign_group_weekend_d_s5():
                _sat5p = set(sat_list5)
                _sun5p = set(sun_list5)
                _hol5p = _sat5p | _sun5p | set(nat_list5)
                # (day, group, min_count) — 週六只要求 A；週日要求 A 和 B
                _reqs5p = []
                for _d in sorted(_sat5p):
                    for _g, _mn in _GROUP_D_SAT_MIN.items():
                        if _mn > 0: _reqs5p.append((_d, _g, _mn))
                for _d in sorted(_sun5p):
                    for _g, _mn in _GROUP_D_SUN_MIN.items():
                        if _mn > 0: _reqs5p.append((_d, _g, _mn))

                for _d5p, _grp5p, _min5p in _reqs5p:
                    _curr5p = sum(
                        1 for i in ai_df.index
                        if cache_group5.get(i) == _grp5p
                        and isinstance(sched[i][_d5p], str)
                        and sched[i][_d5p].startswith("D")
                    )
                    _need5p = _min5p - _curr5p
                    if _need5p <= 0:
                        continue
                    # 確認 D 班當日配額尚有餘裕
                    _qr5p = edited_quota_df[edited_quota_df["日期"] == str(_d5p)]
                    if _qr5p.empty:
                        continue
                    try:
                        _quota5p = int(_qr5p.iloc[0]["D班"])
                    except (KeyError, ValueError):
                        continue
                    # 候選人：同組、非包班、非行政、可合法出勤（放寬 strict_wow 允許孤立班）
                    _cands5p = [
                        i for i in ai_df.index
                        if cache_group5.get(i) == _grp5p
                        and cache_pref[i] == ""
                        and cache_title[i] not in ADMIN_TITLES
                        and can_work_base(i, "D", _d5p, strict_wow=False)
                    ]
                    # 公平排序：假日出勤天數少的優先 → 總出勤天數少的優先
                    _cands5p.sort(key=lambda i: (
                        sum(1 for _hd in _hol5p if is_work(sched[i][_hd])),
                        sum(1 for v in sched[i] if is_work(v))
                    ))
                    for _ci5p in _cands5p:
                        if _need5p <= 0:
                            break
                        _tot5p = sum(
                            1 for _i in ai_df.index
                            if isinstance(sched[_i][_d5p], str) and sched[_i][_d5p].startswith("D")
                        )
                        if _tot5p >= _quota5p:
                            break
                        sched[_ci5p][_d5p] = "D"
                        _need5p -= 1

            pre_assign_group_weekend_d_s5()

            if st.session_state.skill_cols:
                for sk in st.session_state.skill_cols: assign_d_shifts(target_skill=sk)
            assign_d_shifts()

            for idx in ai_df.index:
                for d in range(1, month_days + 1):
                    if sched[idx][d] == "上課":
                        # 假日：先檢查 D班 配額，已達上限則保留「上課」不轉換
                        if d in _hol_set5:
                            _row_sc = edited_quota_df[edited_quota_df["日期"] == str(d)]
                            if not _row_sc.empty:
                                _req_sc = int(_row_sc.iloc[0]["D班"])
                                _curr_sc = sum(1 for _i in ai_df.index if sched[_i][d] == "D")
                                if _curr_sc >= _req_sc:
                                    continue  # 已達假日 D班 配額，保留「上課」
                        sched[idx][d] = "D"

            deficit_nurses = sorted(
                [i for i in ai_df.index if sum(1 for v in sched[i] if is_work(v)) < personal_targets[i]],
                key=lambda i: sum(1 for v in sched[i] if is_work(v)) - personal_targets[i]  # 欠班最多者優先（最負值排最前）
            )
            for n_idx in deficit_nurses:
                worked = sum(1 for v in sched[n_idx] if is_work(v))
                target = personal_targets[n_idx]
                # 包班人員補其包班班別；一般人員補 D
                _pref5 = cache_pref[n_idx]
                if _pref5:
                    _ps5 = "N" if "大夜" in _pref5 else ("E" if "小夜" in _pref5 else ("12-8" if "中" in _pref5 else "D"))
                    # N 包班 / 12-8 包班：仍欠班時允許補 D 班（N→D、12-8→D 均合法）
                    # E 包班：E→D 違規，維持原邏輯只補 E 或 12-8
                    f_s = "D" if _ps5 in ("N", "12-8") else _ps5
                else:
                    f_s = "D"
                for strict_wow, _wv_override in [(True, False), (False, False), (False, False)]:
                    if worked >= target: break
                    # 優先補填「兩側皆已是班」的空隙，其次補延伸連班，最後才補孤立空格
                    def _day_pat5(d):
                        y = sched[n_idx][d - 1] if d > 1 else ""
                        t = sched[n_idx][d + 1] if d < month_days else ""
                        if is_work(y) and is_work(t): return 0
                        if is_work(y) or is_work(t): return 1
                        return 2
                    days_list = sorted(range(1, month_days + 1), key=_day_pat5)
                    for d_int in days_list:
                        if worked >= target: break
                        eff_s5 = f_s  # 有效班別，可能在配額滿時降級為 12-8
                        if _pref5 and f_s in ("E", "N"):
                            q_col5 = f"{f_s}班"
                            row_d5 = edited_quota_df[edited_quota_df["日期"] == str(d_int)]
                            if not row_d5.empty:
                                req5 = int(row_d5.iloc[0][q_col5])
                                curr5 = sum(1 for i in ai_df.index if sched[i][d_int] == f_s)
                                if curr5 >= req5:
                                    # E包班：配額滿時嘗試補 12-8（E→12-8 合法）
                                    # N包班：已在上方設為 f_s="D"，不應進到這裡
                                    if f_s == "E":
                                        eff_s5 = "12-8"
                                    else:
                                        continue  # N班配額滿，跳過此天
                        # ── 補足 Pass 1：D班 / 12-8 每日配額上限檢查 ──
                        # 假日：嚴格遵守配額上限（假日人力有限）
                        # 平日：允許超出配額最多 3 人的緩衝（幫助欠班護理師補足），超過緩衝才截止
                        _row_q_p1 = edited_quota_df[edited_quota_df["日期"] == str(d_int)]
                        _is_hol_p1 = d_int in _hol_set5
                        _D_WEEKDAY_BUFFER = 3  # 平日D班每日最多允許超出配額的人數緩衝
                        if not _row_q_p1.empty:
                            try:
                                if eff_s5 == "D":
                                    if cache_title[n_idx] not in NO_HOL_ADMIN:
                                        _req_p1 = int(_row_q_p1.iloc[0]["D班"])
                                        _curr_p1 = sum(
                                            1 for i in ai_df.index
                                            if isinstance(sched[i][d_int], str) and sched[i][d_int].startswith("D")
                                            and cache_title[i] not in NO_HOL_ADMIN
                                        )
                                        if _is_hol_p1:
                                            if _curr_p1 >= _req_p1: continue  # 假日嚴守上限
                                        else:
                                            if _curr_p1 >= _req_p1 + _D_WEEKDAY_BUFFER: continue  # 平日允許超出最多3人
                                elif eff_s5 == "12-8":
                                    _req_p1 = int(_row_q_p1.iloc[0]["12-8"])
                                    _curr_p1 = sum(1 for i in ai_df.index if sched[i][d_int] == "12-8")
                                    if _curr_p1 >= _req_p1:
                                        continue
                            except (KeyError, ValueError):
                                pass
                        if can_work_base(n_idx, eff_s5, d_int, strict_wow=strict_wow, week_variety_override=_wv_override):
                            if eff_s5 in ("12-8", "E") and not group_cap_ok(n_idx, eff_s5, d_int, sched, cache_group5):
                                continue
                            sched[n_idx][d_int] = eff_s5
                            worked += 1

            # ── 補足 Pass 2：嚴守勞基法 & 配額，盡力再補到 personal_targets ──
            # 與 deficit_nurses 邏輯相同，但改為逐人全月掃描（不隨機），提高命中率
            _hol_set_f5 = set(sat_list5) | set(sun_list5) | set(nat_list5)
            def can_work_force5(n_idx, s, d_int, week_variety_override=False):
                """補足 Pass 2：完全遵守勞基法（連五/14日窗口）及 E/N 配額上限"""
                if sched[n_idx][d_int] not in ["", "上課"]: return False
                if cache_title[n_idx] in ADMIN_TITLES and s != "D": return False
                if cache_title[n_idx] in NO_HOL_SET and d_int in _hol_set_f5: return False
                # 包班人員有假日出勤義務，不受假日資格限制
                if cache_pref[n_idx] == "" and not can_work_holiday_check(n_idx, d_int, cache_can_sat5, cache_can_sun5,
                        cache_can_nat5, sat_list5, sun_list5, nat_list5): return False
                if str(d_int) in class_days_map.get(n_idx, []) and s not in ("D", "N"): return False  # 上課日禁E/12-8
                if sum(1 for x in sched[n_idx] if is_work(x)) >= personal_targets[n_idx]: return False
                y_s = (sched[n_idx][d_int - 1] or "") if d_int > 1 else ""
                t_s = (sched[n_idx][d_int + 1] or "") if d_int < month_days else ""
                y_s_base = "D" if y_s.startswith("D") or y_s in ("上課", "公差") else y_s
                t_s_base = "D" if t_s.startswith("D") or t_s in ("上課", "公差") else t_s
                if is_work(y_s) and s in illegal_next.get(y_s_base, []): return False
                if is_work(t_s) and t_s_base in illegal_next.get(s, []): return False
                # 連五上限（勞基法）
                s_c5 = 1
                for _bd in range(d_int - 1, 0, -1):
                    if is_work(sched[n_idx][_bd]): s_c5 += 1
                    else: break
                for _fd in range(d_int + 1, month_days + 1):
                    if is_work(sched[n_idx][_fd]): s_c5 += 1
                    else: break
                if s_c5 > 5: return False
                # 14 日 12 班窗口（含前月緩衝）
                _has_pbf = bool((st.session_state.get("prev_month_buffer") or {}).get(str(ai_df.at[n_idx,"姓名"]).strip() if n_idx in ai_df.index else "", {}))
                _w_min = (d_int - 13) if _has_pbf else max(1, d_int - 13)
                _w_max = min(d_int, month_days - 13) if month_days >= 14 else 1
                for _sd in range(_w_min, _w_max + 1):
                    _ed = min(month_days, _sd + 13)
                    _ww = 1
                    for _cd in range(_sd, _ed + 1):
                        if _cd == d_int: continue
                        if is_work(_xmonth_shift(n_idx, _cd, sched, ai_df, month_days)): _ww += 1
                    if _ww > 12: return False
                if not week_variety_override and not week_variety_ok(sched, n_idx, s, d_int, st.session_state.first_wday, month_days): return False
                return True

            for n_idx in sorted(ai_df.index, key=lambda i: sum(1 for v in sched[i] if is_work(v)) - personal_targets[i]):
                if sum(1 for v in sched[n_idx] if is_work(v)) >= personal_targets[n_idx]: continue
                _pref_f = cache_pref[n_idx]
                if _pref_f:
                    f_s_f = "N" if "大夜" in _pref_f else ("E" if "小夜" in _pref_f else ("12-8" if "中" in _pref_f else "D"))
                else:
                    f_s_f = "D"
                def _day_pat_f5(d):
                    y = sched[n_idx][d - 1] if d > 1 else ""
                    t = sched[n_idx][d + 1] if d < month_days else ""
                    if is_work(y) and is_work(t): return 0
                    if is_work(y) or is_work(t): return 1
                    return 2
                for d_int in sorted(range(1, month_days + 1), key=_day_pat_f5):
                    if sum(1 for v in sched[n_idx] if is_work(v)) >= personal_targets[n_idx]: break
                    eff_sf = f_s_f
                    if _pref_f and f_s_f in ("E", "N"):
                        _row_df = edited_quota_df[edited_quota_df["日期"] == str(d_int)]
                        if not _row_df.empty:
                            _req_f  = int(_row_df.iloc[0][f"{f_s_f}班"])
                            _curr_f = sum(1 for i in ai_df.index if sched[i][d_int] == f_s_f)
                            if _curr_f >= _req_f:
                                # 包班人員不改排白班，直接跳過
                                continue
                    # ── 補足 Pass 2：D班 / 12-8 每日配額上限檢查 ──
                    # 假日：嚴格遵守配額上限；平日：允許超出配額最多 3 人的緩衝
                    _row_q_p2 = edited_quota_df[edited_quota_df["日期"] == str(d_int)]
                    _is_hol_p2 = d_int in _hol_set_f5
                    _D_WEEKDAY_BUFFER_P2 = 3  # 平日D班每日最多允許超出配額的人數緩衝
                    if not _row_q_p2.empty:
                        try:
                            if eff_sf == "D":
                                if cache_title[n_idx] not in NO_HOL_ADMIN:
                                    _req_p2 = int(_row_q_p2.iloc[0]["D班"])
                                    _curr_p2 = sum(
                                        1 for i in ai_df.index
                                        if isinstance(sched[i][d_int], str) and sched[i][d_int].startswith("D")
                                        and cache_title[i] not in NO_HOL_ADMIN
                                    )
                                    if _is_hol_p2:
                                        if _curr_p2 >= _req_p2: continue  # 假日嚴守上限
                                    else:
                                        if _curr_p2 >= _req_p2 + _D_WEEKDAY_BUFFER_P2: continue  # 平日允許超出最多3人
                            elif eff_sf == "12-8":
                                _req_p2 = int(_row_q_p2.iloc[0]["12-8"])
                                _curr_p2 = sum(1 for i in ai_df.index if sched[i][d_int] == "12-8")
                                if _curr_p2 >= _req_p2:
                                    continue
                        except (KeyError, ValueError):
                            pass
                    if can_work_force5(n_idx, eff_sf, d_int):
                        if eff_sf in ("12-8", "E") and not group_cap_ok(n_idx, eff_sf, d_int, sched, cache_group5):
                            continue
                        sched[n_idx][d_int] = eff_sf

            # ── 補足 Pass 2b：放寬週多樣性，再嘗試補足仍欠班人員 ──────────────────────
            for n_idx in sorted(ai_df.index, key=lambda i: sum(1 for v in sched[i] if is_work(v)) - personal_targets[i]):
                if sum(1 for v in sched[n_idx] if is_work(v)) >= personal_targets[n_idx]: continue
                _pref_2b = cache_pref[n_idx]
                f_s_2b = ("N" if "大夜" in _pref_2b else ("E" if "小夜" in _pref_2b else ("12-8" if "中" in _pref_2b else "D"))) if _pref_2b else "D"
                def _pat_2b(d):
                    y = sched[n_idx][d - 1] if d > 1 else ""
                    t = sched[n_idx][d + 1] if d < month_days else ""
                    if is_work(y) and is_work(t): return 0
                    if is_work(y) or is_work(t): return 1
                    return 2
                for d_int in sorted(range(1, month_days + 1), key=_pat_2b):
                    if sum(1 for v in sched[n_idx] if is_work(v)) >= personal_targets[n_idx]: break
                    eff_2b = f_s_2b
                    if _pref_2b and f_s_2b in ("E", "N"):
                        _row_2b = edited_quota_df[edited_quota_df["日期"] == str(d_int)]
                        if not _row_2b.empty:
                            if sum(1 for i in ai_df.index if sched[i][d_int] == f_s_2b) >= int(_row_2b.iloc[0][f"{f_s_2b}班"]):
                                continue
                    _rowq_2b = edited_quota_df[edited_quota_df["日期"] == str(d_int)]
                    if not _rowq_2b.empty:
                        try:
                            if eff_2b == "D" and cache_title[n_idx] not in NO_HOL_ADMIN:
                                _rq = int(_rowq_2b.iloc[0]["D班"])
                                _cq = sum(1 for i in ai_df.index if isinstance(sched[i][d_int], str) and sched[i][d_int].startswith("D") and cache_title[i] not in NO_HOL_ADMIN)
                                if d_int in _hol_set_f5:
                                    if _cq >= _rq: continue
                                else:
                                    if _cq >= _rq + 3: continue
                            elif eff_2b == "12-8":
                                if sum(1 for i in ai_df.index if sched[i][d_int] == "12-8") >= int(_rowq_2b.iloc[0]["12-8"]): continue
                        except (KeyError, ValueError): pass
                    if can_work_force5(n_idx, eff_2b, d_int, week_variety_override=False):
                        if eff_2b in ("12-8", "E") and not group_cap_ok(n_idx, eff_2b, d_int, sched, cache_group5): continue
                        sched[n_idx][d_int] = eff_2b

            # ── 補足 Pass 3：夜班班別補足（D 班已窮盡後，嘗試 E / N / 12-8）──────────
            # 對象：具夜班資格、非包班、非行政職、仍有欠班的一般護理師
            # 優先補工作日，假日若有缺也補；配額嚴守上限（不使用緩衝）
            _hol_set_p3 = set(sat_list5) | set(sun_list5) | set(nat_list5)
            _il_p3 = {"D": ["N"], "E": ["D","N","12-8"], "12-8": ["N"], "N": []}
            for n_idx in sorted(ai_df.index, key=lambda i: sum(1 for v in sched[i] if is_work(v)) - personal_targets[i]):
                if sum(1 for v in sched[n_idx] if is_work(v)) >= personal_targets[n_idx]: continue
                if cache_pref[n_idx]: continue          # 包班人員：只排其包班班別
                if cache_title[n_idx] in ADMIN_TITLES: continue  # 行政職：只能上 D
                if cache_night5[n_idx] == "": continue  # 無夜班資格：只能上 D
                # 依夜班資格決定可嘗試班別
                _alt_p3 = ["E", "N", "12-8"] if not cache_preg5[n_idx] else ["12-8"]
                for _s3 in _alt_p3:
                    if sum(1 for v in sched[n_idx] if is_work(v)) >= personal_targets[n_idx]: break
                    def _pat_p3(d):
                        y3 = sched[n_idx][d - 1] if d > 1 else ""
                        t3 = sched[n_idx][d + 1] if d < month_days else ""
                        if is_work(y3) and is_work(t3): return 0
                        if is_work(y3) or is_work(t3): return 1
                        return 2
                    for d_int in sorted(range(1, month_days + 1), key=_pat_p3):
                        if sum(1 for v in sched[n_idx] if is_work(v)) >= personal_targets[n_idx]: break
                        if sched[n_idx][d_int] != "": continue
                        # 配額上限（嚴守，不使用緩衝）
                        # 注意：12-8 的欄位名稱為 "12-8"，E/N 為 "E班"/"N班"
                        _rq_p3 = edited_quota_df[edited_quota_df["日期"] == str(d_int)]
                        if not _rq_p3.empty:
                            try:
                                _col3 = {"E": "E班", "N": "N班", "12-8": "12-8"}.get(_s3, f"{_s3}班")
                                _req3 = int(_rq_p3.iloc[0][_col3])
                                _cur3 = sum(1 for i in ai_df.index if sched[i][d_int] == _s3)
                                if _cur3 >= _req3: continue
                            except (KeyError, ValueError): pass
                        # 假日出勤能力
                        if not can_work_holiday_check(n_idx, d_int, cache_can_sat5, cache_can_sun5,
                                                       cache_can_nat5, sat_list5, sun_list5, nat_list5): continue
                        # 相鄰班別規定
                        _y3 = sched[n_idx][d_int - 1] if d_int > 1 else ""
                        _t3 = sched[n_idx][d_int + 1] if d_int < month_days else ""
                        _yb3 = "D" if (_y3.startswith("D") or _y3 in ("上課","公差")) else _y3
                        _tb3 = "D" if (_t3.startswith("D") or _t3 in ("上課","公差")) else _t3
                        if is_work(_y3) and _s3 in _il_p3.get(_yb3, []): continue
                        if is_work(_t3) and _tb3 in _il_p3.get(_s3, []): continue
                        # 連五上限
                        _sc3 = 1
                        for _bd3 in range(d_int - 1, 0, -1):
                            if is_work(sched[n_idx][_bd3]): _sc3 += 1
                            else: break
                        for _fd3 in range(d_int + 1, month_days + 1):
                            if is_work(sched[n_idx][_fd3]): _sc3 += 1
                            else: break
                        if _sc3 > 5: continue
                        # 14 日窗口（含前月緩衝）
                        _14ok3 = True
                        _has_pb3p = bool((st.session_state.get("prev_month_buffer") or {}).get(str(ai_df.at[n_idx,"姓名"]).strip() if n_idx in ai_df.index else "", {}))
                        _wm3 = (d_int - 13) if _has_pb3p else max(1, d_int - 13)
                        _wx3 = min(d_int, month_days - 13) if month_days >= 14 else 1
                        for _sd3 in range(_wm3, _wx3 + 1):
                            _ed3 = min(month_days, _sd3 + 13)
                            _ww3 = sum(1 for _cd3 in range(_sd3, _ed3 + 1)
                                       if _cd3 != d_int and is_work(_xmonth_shift(n_idx, _cd3, sched, ai_df, month_days)))
                            if _ww3 + 1 > 12: _14ok3 = False; break
                        if not _14ok3: continue
                        if not week_variety_ok(sched, n_idx, _s3, d_int, st.session_state.first_wday, month_days): continue
                        if _s3 in ("12-8", "E") and not group_cap_ok(n_idx, _s3, d_int, sched, cache_group5): continue
                        sched[n_idx][d_int] = _s3

            # ── 補足 Final Pass：放寬週多樣性，補齊仍欠班人員 ──────────────────────────────
            # 保留所有勞基法限制（連五 / 14日窗口 / 相鄰班別 / 配額），唯一放寬：週多樣性限制
            # 僅對仍欠班的護師執行，不影響已達標人員
            _hol_set_fp = set(sat_list5) | set(sun_list5) | set(nat_list5)
            _il_fp = {"D": ["N"], "E": ["D", "N", "12-8"], "12-8": ["N"], "N": ["D", "E", "12-8"]}
            for n_idx in sorted(ai_df.index, key=lambda i: sum(1 for v in sched[i] if is_work(v)) - personal_targets[i]):
                if sum(1 for v in sched[n_idx] if is_work(v)) >= personal_targets[n_idx]: continue
                _pref_fp = cache_pref[n_idx]
                if _pref_fp:
                    _fs_fp = "N" if "大夜" in _pref_fp else ("E" if "小夜" in _pref_fp else ("12-8" if "中" in _pref_fp else "D"))
                else:
                    _fs_fp = "D"
                def _pat_fp(d):
                    _y = sched[n_idx][d - 1] if d > 1 else ""
                    _t = sched[n_idx][d + 1] if d < month_days else ""
                    if is_work(_y) and is_work(_t): return 0
                    if is_work(_y) or is_work(_t): return 1
                    return 2
                for d_int in sorted(range(1, month_days + 1), key=_pat_fp):
                    if sum(1 for v in sched[n_idx] if is_work(v)) >= personal_targets[n_idx]: break
                    if sched[n_idx][d_int] not in ["", "上課"]: continue
                    _eff_fp = _fs_fp
                    # 包班人員：E/N 配額滿則跳過（不改排白班）
                    if _pref_fp and _eff_fp in ("E", "N"):
                        _rq_fp = edited_quota_df[edited_quota_df["日期"] == str(d_int)]
                        if not _rq_fp.empty:
                            if sum(1 for i in ai_df.index if sched[i][d_int] == _eff_fp) >= int(_rq_fp.iloc[0][f"{_eff_fp}班"]):
                                continue
                    # 配額上限（D班平日允許緩衝3人，假日嚴守）
                    _rq2_fp = edited_quota_df[edited_quota_df["日期"] == str(d_int)]
                    _is_hol_fp = d_int in _hol_set_fp
                    if not _rq2_fp.empty:
                        try:
                            if _eff_fp == "D" and cache_title[n_idx] not in NO_HOL_ADMIN:
                                _req_fp = int(_rq2_fp.iloc[0]["D班"])
                                _cur_fp = sum(1 for i in ai_df.index if isinstance(sched[i][d_int], str) and sched[i][d_int].startswith("D") and cache_title[i] not in NO_HOL_ADMIN)
                                if _is_hol_fp:
                                    if _cur_fp >= _req_fp: continue
                                else:
                                    if _cur_fp >= _req_fp + 3: continue
                            elif _eff_fp == "12-8":
                                if sum(1 for i in ai_df.index if sched[i][d_int] == "12-8") >= int(_rq2_fp.iloc[0]["12-8"]): continue
                        except (KeyError, ValueError): pass
                    # 假日出勤能力
                    if cache_pref[n_idx] == "" and not can_work_holiday_check(n_idx, d_int, cache_can_sat5, cache_can_sun5, cache_can_nat5, sat_list5, sun_list5, nat_list5): continue
                    if cache_title[n_idx] in NO_HOL_SET and _is_hol_fp: continue
                    if cache_title[n_idx] in ADMIN_TITLES and _eff_fp != "D": continue
                    # 上課日限制
                    if str(d_int) in class_days_map.get(n_idx, []) and _eff_fp not in ("D", "N"): continue
                    # 相鄰班別（勞基法）
                    _y_fp = (sched[n_idx][d_int - 1] or "") if d_int > 1 else ""
                    _t_fp = (sched[n_idx][d_int + 1] or "") if d_int < month_days else ""
                    _yb_fp = "D" if (_y_fp.startswith("D") or _y_fp in ("上課", "公差")) else _y_fp
                    _tb_fp = "D" if (_t_fp.startswith("D") or _t_fp in ("上課", "公差")) else _t_fp
                    if is_work(_y_fp) and _eff_fp in _il_fp.get(_yb_fp, []): continue
                    if is_work(_t_fp) and _tb_fp in _il_fp.get(_eff_fp, []): continue
                    # 連五上限（勞基法）
                    _sc_fp = 1
                    for _bd_fp in range(d_int - 1, 0, -1):
                        if is_work(sched[n_idx][_bd_fp]): _sc_fp += 1
                        else: break
                    for _fd_fp in range(d_int + 1, month_days + 1):
                        if is_work(sched[n_idx][_fd_fp]): _sc_fp += 1
                        else: break
                    if _sc_fp > 5: continue
                    # 14日12班窗口（含前月緩衝）
                    _14ok_fp = True
                    _has_pbfp = bool((st.session_state.get("prev_month_buffer") or {}).get(str(ai_df.at[n_idx,"姓名"]).strip() if n_idx in ai_df.index else "", {}))
                    _wm_fp = (d_int - 13) if _has_pbfp else max(1, d_int - 13)
                    _wx_fp = min(d_int, month_days - 13) if month_days >= 14 else 1
                    for _sd_fp in range(_wm_fp, _wx_fp + 1):
                        _ed_fp = min(month_days, _sd_fp + 13)
                        _ww_fp = sum(1 for _cd in range(_sd_fp, _ed_fp + 1) if _cd != d_int and is_work(_xmonth_shift(n_idx, _cd, sched, ai_df, month_days)))
                        if _ww_fp + 1 > 12: _14ok_fp = False; break
                    if not _14ok_fp: continue
                    if not week_variety_ok(sched, n_idx, _eff_fp, d_int, st.session_state.first_wday, month_days): continue
                    if _eff_fp in ("12-8", "E") and not group_cap_ok(n_idx, _eff_fp, d_int, sched, cache_group5): continue
                    sched[n_idx][d_int] = _eff_fp

            # ── 偏好加班 Pass：人力充足時，依設定多補指定班別人數 ──────────────────────────
            # 僅在已達最低配額後才執行，對每日每班別嘗試多補「extra_staffing_df」設定的人數
            # 保留所有勞基法限制，不強制，找不到合法人員就跳過
            _extra_df = st.session_state.get("extra_staffing_df")
            _wday_names_ep = ["週一", "週二", "週三", "週四", "週五", "週六", "週日"]
            _shift_col_map_ep = {"D班": "D", "E班": "E", "N班": "N", "12-8": "12-8"}

            if _extra_df is not None:
                _hol_set_ep = set(sat_list5) | set(sun_list5) | set(nat_list5)
                for d_int in range(1, month_days + 1):
                    _fw_ep = (st.session_state.first_wday + d_int - 1) % 7
                    _wday_label_ep = _wday_names_ep[_fw_ep]
                    _extra_row = _extra_df[_extra_df["星期"] == _wday_label_ep]
                    if _extra_row.empty: continue

                    for _scol_ep, _stype_ep in _shift_col_map_ep.items():
                        try:
                            _extra_cnt_ep = int(_extra_row.iloc[0][_scol_ep])
                        except (KeyError, ValueError):
                            continue
                        if _extra_cnt_ep <= 0: continue

                        # 計算當日已有多少人排此班別
                        _curr_ep = sum(
                            1 for i in ai_df.index
                            if (sched[i][d_int] == _stype_ep if _stype_ep != "D"
                                else (isinstance(sched[i][d_int], str) and sched[i][d_int].startswith("D")
                                      and cache_title[i] not in NO_HOL_ADMIN))
                        )
                        # 取得當日基本配額
                        _rq_ep = edited_quota_df[edited_quota_df["日期"] == str(d_int)]
                        if _rq_ep.empty: continue
                        try:
                            _base_quota_ep = int(_rq_ep.iloc[0][_scol_ep])
                        except (KeyError, ValueError):
                            continue

                        # 已達基本配額才開始補偏好加班
                        if _curr_ep < _base_quota_ep: continue
                        _target_ep = _base_quota_ep + _extra_cnt_ep
                        if _curr_ep >= _target_ep: continue

                        # 找可合法排此班別的人員（已達個人目標天數的人不強制加班）
                        _need_ep = min(_target_ep - _curr_ep, 4)  # 單日單班最多補 4 人
                        _cands_ep = []
                        for i in ai_df.index:
                            if sched[i][d_int] not in ["", "上課"]: continue
                            if cache_title[i] in ADMIN_TITLES and _stype_ep != "D": continue
                            if cache_title[i] in NO_HOL_SET and d_int in _hol_set_ep: continue
                            if sum(1 for x in sched[i] if is_work(x)) >= personal_targets.get(i, 0): continue
                            if not can_work_base(i, _stype_ep, d_int, strict_wow=False, week_variety_override=False): continue
                            if _stype_ep in ("E", "12-8") and not group_cap_ok(i, _stype_ep, d_int, sched, cache_group5): continue
                            _cands_ep.append(i)

                        # 排序：偏好已有連班的人（避免製造孤立班）
                        _cands_ep.sort(key=lambda i: (
                            0 if (is_work(sched[i][d_int-1] if d_int>1 else "") or
                                  is_work(sched[i][d_int+1] if d_int<month_days else "")) else 1,
                            sum(1 for v in sched[i] if is_work(v))
                        ))

                        for i in _cands_ep[:_need_ep]:
                            sched[i][d_int] = _stype_ep

            # ── 傷兵/助理 最終兜底：強制填滿所有平日空格（不套用任何勞基法限制）──────────────
            # 傷兵/助理：沒有預假，平日全上白班，不計入單位人力配額
            # 放在所有排班邏輯最後，確保不受任何約束漏排
            _nohol_hol_set = set(sat_list5) | set(sun_list5) | set(nat_list5)
            for _nohol_idx in ai_df.index:
                if cache_title[_nohol_idx] not in NO_HOL_ADMIN:
                    continue
                for _nohol_d in range(1, month_days + 1):
                    if _nohol_d in _nohol_hol_set:
                        continue  # 假日不排
                    if sched[_nohol_idx][_nohol_d] not in ["", "上課"]:
                        continue  # 已有班別（含特殊假別）→ 保留
                    sched[_nohol_idx][_nohol_d] = "D"

            # ── 假日出勤事後均等化 ─────────────────────────────────────────
            # 修正 1：擴展互換支援 D/E/N/12-8 所有班別（同班別才能互換）
            # 修正 2：分兩個均等化池
            #   - 常規池：非包班、非行政、至少能上一種假日（全部同池，不分假日能力）
            #   - 包班池：依包班班別分組，同組包班人員互相均等
            _hol_day_set = set(holiday_days_list)
            _wday_set    = set(range(1, month_days + 1)) - _hol_day_set

            def _legal_place_shift(n_idx, d_int, s):
                """檢查在 d_int 放 s 班是否合法（相鄰規定 + 連五 + 假日出勤能力）"""
                if sched[n_idx][d_int] not in ["", "上課"]: return False
                y_s = (sched[n_idx][d_int - 1] or "") if d_int > 1 else ""
                t_s = (sched[n_idx][d_int + 1] or "") if d_int < month_days else ""
                y_b = "D" if (y_s.startswith("D") or y_s in ("上課", "公差")) else y_s
                t_b = "D" if (t_s.startswith("D") or t_s in ("上課", "公差")) else t_s
                if is_work(y_s) and s in illegal_next.get(y_b, []): return False
                if is_work(t_s) and t_b in illegal_next.get(s, []): return False
                sc = 1
                for bd in range(d_int - 1, 0, -1):
                    if is_work(sched[n_idx][bd]): sc += 1
                    else: break
                for fd in range(d_int + 1, month_days + 1):
                    if is_work(sched[n_idx][fd]): sc += 1
                    else: break
                if sc > 5: return False
                if not can_work_holiday_check(n_idx, d_int, cache_can_sat5, cache_can_sun5,
                                              cache_can_nat5, sat_list5, sun_list5, nat_list5):
                    return False
                return True

            def _equalize_holiday_pool(pool_set):
                """對指定人員集合執行假日出勤均等化（支援 D/E/N/12-8 所有班別互換）"""
                if not pool_set: return
                # ── 假日目標：以池內最高應上班天數為基準，所有人使用同一目標 ──
                # 請假多的人與全勤同仁使用相同假日目標，確保假日出勤公平分攤
                _hol_count      = len(_hol_day_set)
                _max_pt         = max((personal_targets.get(i, 0) for i in pool_set), default=0)
                _hol_target_val = round(_max_pt / max(month_days, 1) * _hol_count)
                _hol_target     = {i: _hol_target_val for i in pool_set}
                for _hit in range(500):
                    # 計算有效假日消耗數：實際出勤 + 特殊假別（公假/喪假/病假等）均視為「已消耗」
                    # 目的：避免因特殊假別而讓護理師在均等化中被誤認為「假日出勤不足」
                    _hc = {i: sum(1 for hd in _hol_day_set
                                  if sched[i][hd] not in ("", "O"))
                           for i in pool_set}
                    if not _hc: break
                    _hmax = max(_hc.values())
                    _hmin = min(_hc.values())

                    # 停止條件：全體差距 ≤ 1 且無人低於個人假日目標
                    _below_target = [i for i, c in _hc.items() if c < _hol_target[i]]
                    if _hmax - _hmin <= 1 and not _below_target:
                        break

                    # OVER：高於個人目標者優先讓出；若無，退回全體最大值
                    _personal_over = [i for i, c in _hc.items() if c > _hol_target[i]]
                    if _personal_over:
                        _over_lh = sorted(_personal_over, key=lambda i: -_hc[i])
                    else:
                        _over_lh = [i for i, c in _hc.items() if c == _hmax]

                    # UNDER：低於個人目標者優先補足；若無，退回全體最小值
                    if _below_target:
                        _under_lh = sorted(_below_target, key=lambda i: _hc[i])
                    else:
                        _under_lh = [i for i, c in _hc.items() if c == _hmin]

                    _swapped_h = False
                    for _ov_h in _over_lh:
                        if _swapped_h: break
                        for _un_h in _under_lh:
                            if _swapped_h: break
                            for _hd in sorted(_hol_day_set):
                                if _swapped_h: break
                                _ov_raw = sched[_ov_h][_hd]
                                # over 必須在假日有正規班別
                                if not is_work(_ov_raw): continue
                                if _ov_raw in ("上課", "公差"): continue
                                # ★ 保護：over 的假日班若為預白/公差/國定必上班，不可移動
                                if (_ov_h, _hd) in _locked_set5: continue
                                # 正規化班別（D開頭→D）
                                _swap_s = "D" if str(_ov_raw).startswith("D") else _ov_raw
                                if _swap_s not in ("D", "E", "N", "12-8"): continue
                                # under 在同一假日必須有空
                                if sched[_un_h][_hd] not in ["", "上課"]: continue
                                # under 在這假日能合法上 _swap_s
                                if not _legal_place_shift(_un_h, _hd, _swap_s): continue
                                # 找平日：under 有 _swap_s，over 有空，可執行四格互換
                                for _wd in sorted(_wday_set):
                                    _un_wd_raw = sched[_un_h][_wd]
                                    _un_wd_s = "D" if str(_un_wd_raw).startswith("D") else _un_wd_raw
                                    if _un_wd_s != _swap_s: continue
                                    # ★ 保護：under 的平日班若為預白/公差/國定必上班，不可移動
                                    if (_un_h, _wd) in _locked_set5: continue
                                    if sched[_ov_h][_wd] not in ["", "上課"]: continue
                                    if not _legal_place_shift(_ov_h, _wd, _swap_s): continue
                                    # 四格互換：總出勤天數兩人不變，配額不變
                                    sched[_ov_h][_hd] = ""
                                    sched[_un_h][_hd] = _swap_s
                                    sched[_un_h][_wd] = ""
                                    sched[_ov_h][_wd] = _swap_s
                                    _swapped_h = True
                                    break

                    # ── 策略二：單格轉讓（四格互換無解時備用）────────────────────
                    # over 讓出假日班（出勤-1），under 在同一假日接班（出勤+1）
                    # 條件：over讓出後仍達 personal_targets；under接班後不超標；under有假日出勤資格
                    if not _swapped_h:
                        for _ov_h in _over_lh:
                            if _swapped_h: break
                            for _un_h in _under_lh:
                                if _swapped_h: break
                                for _hd in sorted(_hol_day_set):
                                    if _swapped_h: break
                                    _ov_raw = sched[_ov_h][_hd]
                                    if not is_work(_ov_raw): continue
                                    if _ov_raw in ("上課", "公差"): continue
                                    if (_ov_h, _hd) in _locked_set5: continue
                                    _swap_s = "D" if str(_ov_raw).startswith("D") else _ov_raw
                                    if _swap_s not in ("D", "E", "N", "12-8"): continue
                                    if sched[_un_h][_hd] not in ["", "上課"]: continue
                                    if (_un_h, _hd) in _locked_set5: continue
                                    # under 在此假日能合法排 _swap_s
                                    if not _legal_place_shift(_un_h, _hd, _swap_s): continue
                                    # over 讓出後仍達個人應上班天數（不低於 personal_targets）
                                    _ov_worked = sum(1 for v in sched[_ov_h] if is_work(v))
                                    if _ov_worked - 1 < personal_targets.get(_ov_h, 0): continue
                                    # under 接班後不超出個人應上班天數
                                    _un_worked = sum(1 for v in sched[_un_h] if is_work(v))
                                    if _un_worked + 1 > personal_targets.get(_un_h, 0): continue
                                    # 執行單格轉讓
                                    sched[_ov_h][_hd] = ""
                                    sched[_un_h][_hd] = _swap_s
                                    _swapped_h = True
                                    break
                    # ── 策略三：換日不加班（under 已達標但假日出勤仍偏低）──────────
                    # under 把某個「平日班」挪到「假日」：平日變休、假日補班、總班數不變
                    # 適用：包班人員因日曆結構導致全部班次落在平日（假日出勤 = 0）的情況
                    if not _swapped_h:
                        for _un_h in _under_lh:
                            if _swapped_h: break
                            for _hd in sorted(_hol_day_set):
                                if _swapped_h: break
                                if sched[_un_h][_hd] not in ["", "上課"]: continue
                                if (_un_h, _hd) in _locked_set5: continue
                                for _wd in sorted(_wday_set):
                                    if _swapped_h: break
                                    _un_wd_raw = sched[_un_h][_wd]
                                    if not is_work(_un_wd_raw): continue
                                    if _un_wd_raw in ("上課", "公差"): continue
                                    if (_un_h, _wd) in _locked_set5: continue
                                    _mv_s = "D" if str(_un_wd_raw).startswith("D") else _un_wd_raw
                                    if _mv_s not in ("D", "E", "N", "12-8"): continue
                                    # 假日能合法排此班別（連班/休息間隔等）
                                    if not _legal_place_shift(_un_h, _hd, _mv_s): continue
                                    _s3_qcol = {"D": "D班", "E": "E班", "N": "N班", "12-8": "12-8"}.get(_mv_s, "D班")
                                    # ★ 確認假日目的地未超出配額上限（避免集中填入同一假日）
                                    _hd_qrow = edited_quota_df[edited_quota_df["日期"] == str(_hd)]
                                    if not _hd_qrow.empty:
                                        _hd_quota = int(_hd_qrow.iloc[0].get(_s3_qcol, 0))
                                        _hd_curr = sum(
                                            1 for _qi in ai_df.index
                                            if (str(sched[_qi][_hd]).strip().startswith("D") if _mv_s == "D"
                                                else str(sched[_qi][_hd]).strip() == _mv_s)
                                        )
                                        if _hd_curr >= _hd_quota: continue  # 假日已達配額，換下一個假日
                                    # ★ 確認讓出平日班後，該平日仍符合配額下限
                                    _wd_qrow = edited_quota_df[edited_quota_df["日期"] == str(_wd)]
                                    if not _wd_qrow.empty:
                                        _wd_quota_min = int(_wd_qrow.iloc[0].get(_s3_qcol, 0))
                                        _wd_curr = sum(
                                            1 for _qi in ai_df.index
                                            if (str(sched[_qi][_wd]).strip().startswith("D") if _mv_s == "D"
                                                else str(sched[_qi][_wd]).strip() == _mv_s)
                                        )
                                        if _wd_curr - 1 < _wd_quota_min: continue  # 讓出後低於配額，跳過
                                    # 執行換日：平日變休，假日補班，總班數不變
                                    sched[_un_h][_wd] = ""
                                    sched[_un_h][_hd] = _mv_s
                                    _swapped_h = True
                                    break
                    if not _swapped_h:
                        break

            # ── 假日出勤均等化：單一大池，目標差距 ≤ 1 ─────────────────────────
            # 納入：所有非行政職、非組長人員
            #        且至少能上一種假日（週六/週日/國定其中一種）或有包班意願
            # 排除：NO_HOL_ADMIN（護理長/副護理長/助理/傷兵）、組長
            _hol_elig_set = set(
                i for i in ai_df.index
                if cache_title[i] not in NO_HOL_ADMIN
                and cache_title[i] != "組長"
                and (cache_can_sat5[i] or cache_can_sun5[i] or cache_can_nat5[i]
                     or cache_pref[i] != "")
            )
            _equalize_holiday_pool(_hol_elig_set)

            # ── 班數達標檢查：收集未完美符合 personal_targets 的警示 ──
            _target_warnings5 = []
            for n_idx in ai_df.index:
                _actual5 = sum(1 for v in sched[n_idx] if is_work(v))
                _tgt5    = personal_targets[n_idx]
                if _actual5 != _tgt5:
                    _target_warnings5.append({
                        "姓名":     ai_df.at[n_idx, "姓名"],
                        "應上班天數": _tgt5,
                        "實際排入": _actual5,
                        "差距":     _tgt5 - _actual5,
                    })
            st.session_state.target_warnings = _target_warnings5

            # ✅ 第五步完成：儲存純白班結果，不含加班線
            d_df = pd.DataFrame({"姓名": ai_df["姓名"]})
            for d in range(1, month_days + 1):
                d_df[str(d)] = [sched[i][d] for i in ai_df.index]

            st.session_state.d_sched = d_df
            st.session_state.personal_targets = personal_targets
            st.rerun()

    if st.session_state.d_sched is not None:
        # ── 班數達標警示 ──────────────────────────────────────────────────
        _tw = st.session_state.get("target_warnings", [])
        if _tw:
            st.error(f"⚠️ 以下 {len(_tw)} 人在遵守人力配額及勞基法規定下，**上班天數無法完美符合應上班天數**，請手動調整或檢查配額設定：")
            _tw_df = pd.DataFrame(_tw)
            st.dataframe(_tw_df, use_container_width=False, hide_index=True)
        else:
            st.success("✅ 12-8 + 白班已全數排入，所有人上班天數完美符合應上班天數！")

        # ── 換班配對推薦（僅在最終班表產出後顯示）──────────────────────────────
        # 規則：同一天真正對調——欠班者得到超班者的班，超班者改為休；
        #       預假日（O / 特殊假別）與預班日（預白 / 上課 / 公差）均不可異動。
        if st.session_state.get("final_sched") is not None:
            st.write("#### 🔄 換班配對推薦")

            _pt_sw    = st.session_state.personal_targets
            _fs_sw    = st.session_state.final_sched
            _dcols_sw = [str(_d) for _d in range(1, month_days + 1)]
            _sat_sw   = set(st.session_state.saturdays_list)
            _sun_sw   = set(st.session_state.sundays_list)
            _nat_sw   = set(st.session_state.nat_holidays_list)
            _hol_sw   = _sat_sw | _sun_sw | _nat_sw
            _il_sw    = {"D": ["N"], "E": ["D", "N", "12-8"], "12-8": ["N"], "N": []}

            # ── 鎖定日判斷：預假 / 預班 / 上課 / 公差 均不可異動 ──────────────
            def _is_locked_sw(nidx, day):
                _cell = str(_fs_sw.at[nidx, str(day)] if str(day) in _fs_sw.columns and nidx in _fs_sw.index else "").strip()
                # 特殊值直接鎖定
                if _cell in ("O", "上課", "公差"): return True
                # 特殊假別（非空白、非標準班別代碼）
                _standard = {"", "D", "E", "N", "12-8"}
                if _cell not in _standard and not (_cell.startswith("D") and _cell[1:].isdigit()): return True
                # 預休日期（ai_df 欄位）
                _pre_off = {int(s.strip()) for s in str(ai_df.at[nidx, "預休日期"] if "預休日期" in ai_df.columns else "").split(",") if s.strip().isdigit()}
                if day in _pre_off: return True
                # 預白日期（ai_df 欄位）
                _pre_wh = {int(s.strip()) for s in str(ai_df.at[nidx, "預白日期"] if "預白日期" in ai_df.columns else "").split(",") if s.strip().isdigit()}
                if day in _pre_wh: return True
                # 上課日期（ai_df 欄位）
                _class_d = {int(s.strip()) for s in str(ai_df.at[nidx, "上課日期"] if "上課日期" in ai_df.columns else "").split(",") if s.strip().isdigit()}
                if day in _class_d: return True
                # 公差日期（ai_df 欄位）
                _gongcha = {int(s.strip()) for s in str(ai_df.at[nidx, "公差日期"] if "公差日期" in ai_df.columns else "").split(",") if s.strip().isdigit()}
                if day in _gongcha: return True
                return False

            # 計算每人實際上班天數
            _wk_sw = {
                _i: sum(1 for _dc in _dcols_sw
                        if is_work(str(_fs_sw.at[_i, _dc] if _dc in _fs_sw.columns else "").strip()))
                for _i in ai_df.index
            }

            # 僅一般護理師（職稱不在 ADMIN_TITLES、包班意願為空白）可納入配對
            _gen_sw = [_i for _i in ai_df.index
                       if str(ai_df.at[_i, "職稱"]).strip() not in ADMIN_TITLES
                       and str(ai_df.at[_i, "包班意願"]).strip() == ""]

            # 欠班人員（依欠差降冪，僅一般護理師）
            _under_sw = sorted(
                [_i for _i in _gen_sw if _wk_sw[_i] < _pt_sw.get(_i, 0)],
                key=lambda _i: _pt_sw.get(_i, 0) - _wk_sw[_i],
                reverse=True
            )
            # 超班人員（僅一般護理師）
            _over_sw = [_i for _i in _gen_sw if _wk_sw[_i] > _pt_sw.get(_i, 0)]

            def _consec_sw(nidx, day):
                """排入 day 後，連續工作天 ≤ 5？"""
                _c = 1
                for _b in range(day - 1, 0, -1):
                    if is_work(str(_fs_sw.at[nidx, str(_b)] if str(_b) in _fs_sw.columns else "").strip()): _c += 1
                    else: break
                for _f in range(day + 1, month_days + 1):
                    if is_work(str(_fs_sw.at[nidx, str(_f)] if str(_f) in _fs_sw.columns else "").strip()): _c += 1
                    else: break
                return _c <= 5

            def _adj_sw(nidx, shift, day):
                """鄰班合法性（illegal_next）"""
                _y = str(_fs_sw.at[nidx, str(day - 1)] if day > 1 and str(day - 1) in _fs_sw.columns else "").strip()
                _t = str(_fs_sw.at[nidx, str(day + 1)] if day < month_days and str(day + 1) in _fs_sw.columns else "").strip()
                _yb = "D" if (_y.startswith("D") or _y in ("上課", "公差")) else _y
                _tb = "D" if (_t.startswith("D") or _t in ("上課", "公差")) else _t
                if is_work(_y) and shift in _il_sw.get(_yb, []): return False
                if is_work(_t) and _tb in _il_sw.get(shift, []): return False
                return True

            def _hol_sw_ok(nidx, day):
                """假日出勤資格"""
                if day not in _hol_sw: return True
                if day in _sat_sw and str(ai_df.at[nidx, "能上週六"]).strip() != "是": return False
                if day in _sun_sw and str(ai_df.at[nidx, "能上週日"]).strip() != "是": return False
                if day in _nat_sw and str(ai_df.at[nidx, "能上國定假日"]).strip() != "是": return False
                return True

            _recs_sw = []
            _done_under = set()

            for _un in _under_sw:
                if _un in _done_under: continue
                for _d in range(1, month_days + 1):
                    _dkey = str(_d)
                    if _dkey not in _fs_sw.columns: continue
                    _un_cell = str(_fs_sw.at[_un, _dkey] if _un in _fs_sw.index else "").strip()
                    # 欠班者當日必須為空格（真正可排入）
                    if _un_cell not in ("", "nan"): continue
                    # 欠班者此日不可為鎖定日
                    if _is_locked_sw(_un, _d): continue

                    for _ov in _over_sw:
                        if _ov == _un: continue
                        _ov_val = str(_fs_sw.at[_ov, _dkey] if _ov in _fs_sw.index else "").strip()
                        # 超班者當日必須有可移出的工作班
                        if not is_work(_ov_val): continue
                        # 超班者此日不可為鎖定日（預假 / 預白 / 上課 / 公差）
                        if _is_locked_sw(_ov, _d): continue
                        # 對調後超班者工作天數仍 >= personal_targets
                        if _wk_sw[_ov] - 1 < _pt_sw.get(_ov, 0): continue
                        # 欠班者排入後連五合法
                        if not _consec_sw(_un, _d): continue
                        # 欠班者鄰班規則
                        _shift_base = "D" if _ov_val.startswith("D") else _ov_val
                        if not _adj_sw(_un, _shift_base, _d): continue
                        # 假日出勤資格
                        if not _hol_sw_ok(_un, _d): continue
                        # ✅ 合法對調配對
                        _recs_sw.append({
                            "欠班者":   ai_df.at[_un, "姓名"],
                            "超班者":   ai_df.at[_ov, "姓名"],
                            "對調日期": _d,
                            "班別":     _ov_val,
                            "_un":      _un,
                            "_ov":      _ov,
                        })
                        _done_under.add(_un)
                        break
                    if _un in _done_under: break

            if _recs_sw:
                st.caption("💡 對調方式：欠班者在該日得到超班者的班別；超班者當日改為休息。預假日與預班日均不納入配對。")
                st.dataframe(
                    pd.DataFrame(_recs_sw)[["欠班者", "超班者", "對調日期", "班別"]],
                    use_container_width=False, hide_index=True
                )
                if st.button("✅ 套用所有推薦換班", key="btn_apply_swap_recs"):
                    _fs_new = st.session_state.final_sched.copy()
                    for _rec in _recs_sw:
                        _dk = str(_rec["對調日期"])
                        _fs_new.at[_rec["_un"], _dk] = _rec["班別"]
                        _fs_new.at[_rec["_ov"], _dk] = ""
                    st.session_state.final_sched = _fs_new
                    st.rerun()
            else:
                st.info("✅ 目前無欠班情形，或欠班人員已無法找到合法對調配對，無推薦換班。")

        with st.expander("📄 點擊展開白班排班結果", expanded=True):
            _edit_d = st.checkbox("🖊️ 開啟手動編輯模式", value=False, key="chk_edit_d_sched")
            if _edit_d:
                st.caption("💡 直接點擊儲存格輸入班別代碼，修改即時儲存")
                _edited_d = st.data_editor(
                    st.session_state.d_sched,
                    column_config=make_sched_col_config(month_days),
                    use_container_width=True, hide_index=True, key="edit_d_sched"
                )
                st.session_state.d_sched = _edited_d
            else:
                _day_cols_d = [str(i) for i in range(1, month_days + 1)]
                _disp_d = st.session_state.d_sched.copy()
                for _c in _day_cols_d:
                    _disp_d[_c] = _disp_d[_c].apply(abbrev_display)
                # 預白班：D 格標為 Dx
                _disp_d = apply_prewhite_dx(_disp_d, ai_df, month_days)
                st.dataframe(
                    build_schedule_with_counts(_disp_d, st.session_state.d_sched, _day_cols_d, ai_df),
                    use_container_width=True
                )

        display_safety_radar(st.session_state.d_sched, edited_quota_df, ai_df)

        if st.session_state.step == 5:
            col_btn_back, col_btn_go = st.columns([1, 4])
            with col_btn_back:
                if st.button("⬅️ 退回重排白班", type="secondary"):
                    st.session_state.d_sched = None
                    st.rerun()
            with col_btn_go:
                if st.button("✅ 確認 12-8 + 白班無誤，進行加班線均分", type="primary"):
                    st.session_state.step = 6
                    st.rerun()

# ==========================================
# 📍 第六步：加班線均分 → 最終結算
# ==========================================
if st.session_state.step >= 6:
    st.divider()
    st.header("6️⃣ 第六步：均分加班線 → 最終結算")

    ai_df = st.session_state.ai_df
    month_days = st.session_state.month_days
    edited_quota_df = st.session_state.edited_quota_df
    holiday_days_list = st.session_state.holiday_list

    with st.expander("⚙️ 調整每日人力配額（修改後需重新排班才生效）", expanded=False):
        _q6 = st.data_editor(st.session_state.edited_quota_df, hide_index=True, use_container_width=True, key="quota_editor_step6")
        if st.button("✅ 套用配額變更並重排（從第六步重算）", key="apply_quota_step6"):
            st.session_state.edited_quota_df = _q6
            st.session_state.final_sched = None
            st.rerun()
    edited_quota_df = st.session_state.edited_quota_df

    if st.session_state.final_sched is None:
        with st.spinner("正在均分加班線並執行最終結算…"):
            sched_df = st.session_state.d_sched.copy()
            sched = {i: [""] + list(sched_df.iloc[i, 1:].values) for i in range(len(ai_df))}

            cache_skills   = {i: [x.strip() for x in str(row.get("次專科能力","")).split(",")] for i, row in ai_df.iterrows()}
            cache_circ     = {i: str(row.get("流動資格","")).strip() == "是" for i, row in ai_df.iterrows()}
            cache_title    = {i: str(row.get("職稱","")).strip() for i, row in ai_df.iterrows()}
            cache_pref     = {i: str(row.get("包班意願","")).strip() for i, row in ai_df.iterrows()}
            cache_preg6    = {i: str(row.get("孕/育嬰免夜班","")).strip() == "是" for i, row in ai_df.iterrows()}
            cache_night6   = {i: str(row.get("夜班資格","")).strip() for i, row in ai_df.iterrows()}
            cache_leader6  = {i: str(row.get("控台資格","")).strip() for i, row in ai_df.iterrows()}
            class_days_map = {i: [s.strip() for s in str(row.get("上課日期","")).split(",") if s.strip().isdigit()] for i, row in ai_df.iterrows()}
            cache_can_sat6 = {i: str(row.get("能上週六","")).strip() == "是" for i, row in ai_df.iterrows()}
            cache_can_sun6 = {i: str(row.get("能上週日","")).strip() == "是" for i, row in ai_df.iterrows()}
            cache_can_nat6 = {i: str(row.get("能上國定假日","")).strip() == "是" for i, row in ai_df.iterrows()}
            cache_group6   = {i: str(row.get("組別", "")).strip().upper() for i, row in ai_df.iterrows()}
            sat_list6 = st.session_state.saturdays_list
            sun_list6 = st.session_state.sundays_list
            nat_list6 = st.session_state.nat_holidays_list
            personal_targets = st.session_state.personal_targets

            illegal_next = {"D": ["N"], "E": ["D","N","12-8"], "12-8": ["N"], "N": []}

            # ── 第六步週末 A/B 組別補強：優先補足週末 D 班組別最低配額 ──
            _sat6p = set(sat_list6)
            _sun6p = set(sun_list6)
            _hol6p = _sat6p | _sun6p | set(nat_list6)
            _reqs6p = []
            for _d6p in sorted(_sat6p):
                for _g6p, _mn6p in _GROUP_D_SAT_MIN.items():
                    if _mn6p > 0: _reqs6p.append((_d6p, _g6p, _mn6p))
            for _d6p in sorted(_sun6p):
                for _g6p, _mn6p in _GROUP_D_SUN_MIN.items():
                    if _mn6p > 0: _reqs6p.append((_d6p, _g6p, _mn6p))

            for _day6p, _grp6p, _min6p in _reqs6p:
                _curr6p = sum(
                    1 for i in ai_df.index
                    if cache_group6.get(i) == _grp6p
                    and isinstance(sched[i][_day6p], str)
                    and sched[i][_day6p].startswith("D")
                )
                _need6p = _min6p - _curr6p
                if _need6p <= 0:
                    continue
                _qr6p = edited_quota_df[edited_quota_df["日期"] == str(_day6p)]
                if _qr6p.empty:
                    continue
                try:
                    _quota6p = int(_qr6p.iloc[0]["D班"])
                except (KeyError, ValueError):
                    continue
                _cands6p = []
                for i in ai_df.index:
                    if cache_group6.get(i) != _grp6p: continue
                    if cache_pref[i] != "": continue
                    if cache_title[i] in ADMIN_TITLES: continue
                    if sched[i][_day6p] != "": continue
                    if cache_pref[i] == "" and not can_work_holiday_check(
                            i, _day6p, cache_can_sat6, cache_can_sun6, cache_can_nat6,
                            sat_list6, sun_list6, nat_list6): continue
                    _worked6p = sum(1 for x in sched[i] if is_work(x))
                    if _worked6p >= personal_targets[i]: continue
                    # 連五檢查
                    _sc6p = 1
                    for _b6p in range(_day6p - 1, 0, -1):
                        if is_work(sched[i][_b6p]): _sc6p += 1
                        else: break
                    for _f6p in range(_day6p + 1, month_days + 1):
                        if is_work(sched[i][_f6p]): _sc6p += 1
                        else: break
                    if _sc6p > 5: continue
                    _cands6p.append(i)
                _cands6p.sort(key=lambda i: (
                    sum(1 for _hd in _hol6p if is_work(sched[i][_hd])),
                    sum(1 for v in sched[i] if is_work(v))
                ))
                for _ci6p in _cands6p:
                    if _need6p <= 0: break
                    _tot6p = sum(
                        1 for _i in ai_df.index
                        if isinstance(sched[_i][_day6p], str) and sched[_i][_day6p].startswith("D")
                    )
                    if _tot6p >= _quota6p: break
                    sched[_ci6p][_day6p] = "D"
                    _need6p -= 1

            # ── 清尾補班：12-8 排完後，補回仍未達 personal_targets 的空格 ──
            for n_idx in sorted(ai_df.index, key=lambda i: sum(1 for v in sched[i] if is_work(v)) - personal_targets[i]):
                _pref6 = cache_pref[n_idx]
                # 行政職稱（組長/護理長/副護理長）只能上白班；包班人員補其包班班別；其餘補 D
                if cache_title[n_idx] in ADMIN_TITLES:
                    _fill_s = "D"
                elif _pref6:
                    _fill_s = "N" if "大夜" in _pref6 else ("E" if "小夜" in _pref6 else ("12-8" if "中" in _pref6 else "D"))
                else:
                    _fill_s = "D"
                for d_int in range(1, month_days + 1):
                    if sched[n_idx][d_int] != "": continue
                    # 護理長/副護理長/組長：行政班，不填假日格
                    if cache_title[n_idx] in NO_HOL_SET and d_int in set(sat_list6) | set(sun_list6) | set(nat_list6): continue
                    worked_now = sum(1 for x in sched[n_idx] if is_work(x))
                    if worked_now >= personal_targets[n_idx]: break
                    # 每日配額上限檢查（E/N/D/12-8 全部班別）
                    # NO_HOL_ADMIN 不計入單位人力，不受 D班配額限制；配額計數也排除 NO_HOL_ADMIN
                    eff_fill6 = _fill_s
                    _row_q6 = edited_quota_df[edited_quota_df["日期"] == str(d_int)]
                    if not _row_q6.empty:
                        try:
                            if eff_fill6 in ("E", "N"):
                                _req6 = int(_row_q6.iloc[0][f"{eff_fill6}班"])
                                _curr6 = sum(1 for i in ai_df.index if sched[i][d_int] == eff_fill6)
                                if _curr6 >= _req6:
                                    continue  # 包班人員不改排白班，直接跳過
                            elif eff_fill6 == "D":
                                if cache_title[n_idx] not in NO_HOL_ADMIN:
                                    _req6 = int(_row_q6.iloc[0]["D班"])
                                    _curr6 = sum(1 for i in ai_df.index
                                                 if isinstance(sched[i][d_int], str) and sched[i][d_int].startswith("D")
                                                 and cache_title[i] not in NO_HOL_ADMIN)
                                    if _curr6 >= _req6: continue
                            elif eff_fill6 == "12-8":
                                _req6 = int(_row_q6.iloc[0]["12-8"])
                                _curr6 = sum(1 for i in ai_df.index if sched[i][d_int] == "12-8")
                                if _curr6 >= _req6:
                                    continue
                        except (KeyError, ValueError):
                            pass
                    # 勞基法：11 小時間距 + 連五
                    y_s6 = sched[n_idx][d_int - 1] if d_int > 1 else ""
                    t_s6 = sched[n_idx][d_int + 1] if d_int < month_days else ""
                    _il = {"D": ["N"], "E": ["D","N","12-8"], "12-8": ["N"], "N": []}
                    _ys_k = "D" if (y_s6.startswith("D") or y_s6 in ("上課", "公差")) else y_s6
                    _ts_k = "D" if (t_s6.startswith("D") or t_s6 in ("上課", "公差")) else t_s6
                    if is_work(y_s6) and eff_fill6 in _il.get(_ys_k, []): continue
                    if is_work(t_s6) and _ts_k in _il.get(eff_fill6, []): continue
                    _sc6 = 1
                    for _bd in range(d_int - 1, 0, -1):
                        if is_work(sched[n_idx][_bd]): _sc6 += 1
                        else: break
                    for _fd in range(d_int + 1, month_days + 1):
                        if is_work(sched[n_idx][_fd]): _sc6 += 1
                        else: break
                    if _sc6 > 5: continue
                    # 勞基法 §30-1：14 日窗口內上班天數 ≤ 12
                    _w6_min = max(1, d_int - 13)
                    _w6_max = min(d_int, month_days - 13) if month_days >= 14 else 1
                    _14d_ok = True
                    for _sd6 in range(_w6_min, _w6_max + 1):
                        _ed6 = min(month_days, _sd6 + 13)
                        _ww6 = sum(1 for _cd6 in range(_sd6, _ed6 + 1)
                                   if _cd6 != d_int and is_work(sched[n_idx][_cd6]))
                        if _ww6 + 1 > 12:
                            _14d_ok = False
                            break
                    if not _14d_ok: continue
                    # 假日出勤能力檢查（包班人員有假日出勤義務，不受此限）
                    if cache_pref[n_idx] == "" and not can_work_holiday_check(n_idx, d_int, cache_can_sat6, cache_can_sun6, cache_can_nat6, sat_list6, sun_list6, nat_list6): continue
                    if not week_variety_ok(sched, n_idx, eff_fill6, d_int, st.session_state.first_wday, month_days): continue
                    if eff_fill6 in ("12-8", "E") and not group_cap_ok(n_idx, eff_fill6, d_int, sched, cache_group6): continue
                    sched[n_idx][d_int] = eff_fill6

            # ── 清尾補班 Pass 2：夜班資格者補 E / N / 12-8（D 班已窮盡後）──────
            _hol_set6p2 = set(sat_list6) | set(sun_list6) | set(nat_list6)
            _il_p2 = {"D": ["N"], "E": ["D","N","12-8"], "12-8": ["N"], "N": []}
            for n_idx in sorted(ai_df.index, key=lambda i: sum(1 for v in sched[i] if is_work(v)) - personal_targets[i]):
                if sum(1 for v in sched[n_idx] if is_work(v)) >= personal_targets[n_idx]: continue
                if cache_pref[n_idx]: continue            # 包班人員已由上方 Pass 處理
                if cache_title[n_idx] in ADMIN_TITLES: continue  # 行政職稱只上 D
                if cache_night6[n_idx] == "": continue   # 無夜班資格：只能上 D，不在此 pass 處理
                _alt_p2 = ["E", "N", "12-8"] if not cache_preg6[n_idx] else ["12-8"]
                def _pat_p2(d, s=n_idx):
                    """優先排在現有班別多的週內，分散原則"""
                    if d in _hol_set6p2: return (1, d)  # 假日後排
                    return (0, d)
                for _s2 in _alt_p2:
                    if sum(1 for v in sched[n_idx] if is_work(v)) >= personal_targets[n_idx]: break
                    for d_int in sorted(range(1, month_days + 1), key=lambda d: _pat_p2(d)):
                        if sched[n_idx][d_int] != "": continue
                        if sum(1 for v in sched[n_idx] if is_work(v)) >= personal_targets[n_idx]: break
                        # 假日出勤資格
                        if d_int in _hol_set6p2:
                            if cache_title[n_idx] in NO_HOL_SET: continue
                            if not can_work_holiday_check(n_idx, d_int, cache_can_sat6, cache_can_sun6, cache_can_nat6, sat_list6, sun_list6, nat_list6): continue
                        # 夜班資格對應班別限制
                        _night6q = cache_night6[n_idx]
                        if _s2 == "N" and _night6q not in ("大夜",): continue
                        if _s2 == "E" and _night6q not in ("小夜", "大夜"): continue
                        if _s2 == "12-8" and _night6q not in ("大夜", "小夜", "中班"): continue
                        # 每日配額上限
                        _row_q6p2 = edited_quota_df[edited_quota_df["日期"] == str(d_int)]
                        if not _row_q6p2.empty:
                            try:
                                _col_p2 = {"E": "E班", "N": "N班", "12-8": "12-8"}.get(_s2)
                                if _col_p2:
                                    _req_p2 = int(_row_q6p2.iloc[0][_col_p2])
                                    _cur_p2 = sum(1 for i in ai_df.index if sched[i][d_int] == _s2)
                                    if _cur_p2 >= _req_p2: continue
                            except (KeyError, ValueError):
                                pass
                        # 勞基法：相鄰班別間距
                        _ys_p2 = sched[n_idx][d_int - 1] if d_int > 1 else ""
                        _ts_p2 = sched[n_idx][d_int + 1] if d_int < month_days else ""
                        _yk_p2 = "D" if (_ys_p2.startswith("D") or _ys_p2 in ("上課","公差")) else _ys_p2
                        _tk_p2 = "D" if (_ts_p2.startswith("D") or _ts_p2 in ("上課","公差")) else _ts_p2
                        if is_work(_ys_p2) and _s2 in _il_p2.get(_yk_p2, []): continue
                        if is_work(_ts_p2) and _tk_p2 in _il_p2.get(_s2, []): continue
                        # 連五
                        _sc_p2 = 1
                        for _b2 in range(d_int - 1, 0, -1):
                            if is_work(sched[n_idx][_b2]): _sc_p2 += 1
                            else: break
                        for _f2 in range(d_int + 1, month_days + 1):
                            if is_work(sched[n_idx][_f2]): _sc_p2 += 1
                            else: break
                        if _sc_p2 > 5: continue
                        # 14 日窗口
                        _w2_min = max(1, d_int - 13)
                        _w2_max = min(d_int, month_days - 13) if month_days >= 14 else 1
                        _14d_p2 = True
                        for _sd2 in range(_w2_min, _w2_max + 1):
                            _ed2 = min(month_days, _sd2 + 13)
                            _ww2 = sum(1 for _cd2 in range(_sd2, _ed2 + 1)
                                       if _cd2 != d_int and is_work(sched[n_idx][_cd2]))
                            if _ww2 + 1 > 12:
                                _14d_p2 = False
                                break
                        if not _14d_p2: continue
                        # 週多樣性 + 組別上限
                        if not week_variety_ok(sched, n_idx, _s2, d_int, st.session_state.first_wday, month_days): continue
                        if _s2 in ("12-8", "E") and not group_cap_ok(n_idx, _s2, d_int, sched, cache_group6): continue
                        sched[n_idx][d_int] = _s2

            # ── Step 6 夜班均等化：清尾補班 Pass 2 後重新拉平 E+N+12-8 差距 ≤ 1 ──
            # 建立 _locked_set6（預白班/公差/國定假日必上班別/上課日期，不可被互換移動）
            _prewhite_set6: set = set()
            _mand_hol_set6: set = set()
            _gongcha_set6:  set = set()
            _class_set6:    set = set()
            for _pi6, _pr6 in ai_df.iterrows():
                for _dv6 in str(_pr6.get("預白日期", "")).split(","):
                    if _dv6.strip().isdigit() and 1 <= int(_dv6.strip()) <= month_days:
                        _prewhite_set6.add((_pi6, int(_dv6.strip())))
                for _item6 in str(_pr6.get("國定假日必上班別", "")).split(","):
                    for _delim6 in [":", "-"]:
                        if _delim6 in _item6:
                            _dv6 = _item6.split(_delim6)[0].strip()
                            if _dv6.isdigit() and 1 <= int(_dv6) <= month_days:
                                _mand_hol_set6.add((_pi6, int(_dv6)))
                            break
                for _dv6 in str(_pr6.get("公差日期", "")).split(","):
                    if _dv6.strip().isdigit() and 1 <= int(_dv6.strip()) <= month_days:
                        _gongcha_set6.add((_pi6, int(_dv6.strip())))
                for _dv6 in str(_pr6.get("上課日期", "")).split(","):
                    if _dv6.strip().isdigit() and 1 <= int(_dv6.strip()) <= month_days:
                        _class_set6.add((_pi6, int(_dv6.strip())))
            _locked_set6 = _prewhite_set6 | _mand_hol_set6 | _gongcha_set6 | _class_set6

            # 均等池：非包班、非組長、具夜班資格
            _night_elig6 = [
                i for i in ai_df.index
                if cache_pref[i] == ""
                and cache_title[i] != "組長"
                and cache_night6[i] != ""
            ]
            _night_elig_set6 = set(_night_elig6)
            _il6_eq = {"D": ["N"], "E": ["D","N","12-8"], "12-8": ["N"], "N": []}

            def _can_12_8_nocheck6(n_idx, d_int):
                """Step 6：d_int 可合法排 12-8（不含 personal_targets 上限，適用互換）"""
                if sched[n_idx][d_int] not in ["", "上課"]: return False
                if cache_pref[n_idx] == "" and cache_night6[n_idx] not in ("大夜", "小夜", "中班") and not cache_preg6[n_idx]: return False
                if cache_pref[n_idx] == "" and not can_work_holiday_check(n_idx, d_int, cache_can_sat6, cache_can_sun6, cache_can_nat6, sat_list6, sun_list6, nat_list6): return False
                _y6e = sched[n_idx][d_int - 1] if d_int > 1 else ""
                _t6e = sched[n_idx][d_int + 1] if d_int < month_days else ""
                _yb6e = "D" if (_y6e.startswith("D") or _y6e in ("上課", "公差")) else _y6e
                _tb6e = "D" if (_t6e.startswith("D") or _t6e in ("上課", "公差")) else _t6e
                if is_work(_y6e) and "12-8" in _il6_eq.get(_yb6e, []): return False
                if is_work(_t6e) and _tb6e in _il6_eq.get("12-8", []): return False
                _sc6e = 1
                for _bd6 in range(d_int - 1, 0, -1):
                    if is_work(sched[n_idx][_bd6]): _sc6e += 1
                    else: break
                for _fd6 in range(d_int + 1, month_days + 1):
                    if is_work(sched[n_idx][_fd6]): _sc6e += 1
                    else: break
                if _sc6e > 5: return False
                return True

            def _can_D_nocheck6(n_idx, d_int):
                """Step 6：d_int 可合法排 D（不含 personal_targets 上限，適用互換）"""
                if sched[n_idx][d_int] not in ["", "上課"]: return False
                _y6e = sched[n_idx][d_int - 1] if d_int > 1 else ""
                _t6e = sched[n_idx][d_int + 1] if d_int < month_days else ""
                _yb6e = "D" if (_y6e.startswith("D") or _y6e in ("上課", "公差")) else _y6e
                _tb6e = "D" if (_t6e.startswith("D") or _t6e in ("上課", "公差")) else _t6e
                if is_work(_y6e) and "D" in _il6_eq.get(_yb6e, []): return False
                if is_work(_t6e) and _tb6e in _il6_eq.get("D", []): return False
                _sc6e = 1
                for _bd6 in range(d_int - 1, 0, -1):
                    if is_work(sched[n_idx][_bd6]): _sc6e += 1
                    else: break
                for _fd6 in range(d_int + 1, month_days + 1):
                    if is_work(sched[n_idx][_fd6]): _sc6e += 1
                    else: break
                if _sc6e > 5: return False
                return True

            # ── 夜班個人目標：全月E+N+12-8配額總數 ÷ 符合資格人數，所有人同一目標 ──
            # 不論個人請假多少，夜班目標一律相同，確保假日均等分攤
            _total_night_quota6 = 0
            for _d6q in range(1, month_days + 1):
                _row6q = edited_quota_df[edited_quota_df["日期"] == str(_d6q)]
                if not _row6q.empty:
                    _total_night_quota6 += (
                        int(_row6q.iloc[0].get("E班", 0)) +
                        int(_row6q.iloc[0].get("N班", 0)) +
                        int(_row6q.iloc[0].get("12-8", 0))
                    )
            _night_target_val6 = (
                round(_total_night_quota6 / len(_night_elig_set6))
                if _night_elig_set6 else 0
            )
            _night_target6 = {i: _night_target_val6 for i in _night_elig_set6}

            for _nit6 in range(500):
                _nc6 = {i: sum(1 for v in sched[i] if v in ("E", "N", "12-8"))
                        for i in _night_elig_set6}
                if not _nc6: break
                _nmax6 = max(_nc6.values())
                _nmin6 = min(_nc6.values())

                # 停止條件：全體差距 ≤ 1 且無人低於個人夜班目標
                _below_night6 = [i for i, c in _nc6.items() if c < _night_target6[i]]
                if _nmax6 - _nmin6 <= 1 and not _below_night6:
                    break

                # OVER：高於個人目標者優先讓出；若無，退回全體最大值
                _personal_over6 = [i for i, c in _nc6.items() if c > _night_target6[i]]
                if _personal_over6:
                    _over_l6 = sorted(_personal_over6, key=lambda i: -_nc6[i])
                else:
                    _over_l6 = [i for i, c in _nc6.items() if c == _nmax6]

                # UNDER：低於個人目標者優先補足；若無，退回全體最小值
                if _below_night6:
                    _under_l6 = sorted(_below_night6, key=lambda i: _nc6[i])
                else:
                    _under_l6 = [i for i, c in _nc6.items() if c == _nmin6]

                _swapped6 = False
                for _ov6 in _over_l6:
                    if _swapped6: break
                    for _un6 in _under_l6:
                        if _swapped6: break
                        for _d6 in range(1, month_days + 1):
                            if _swapped6: break
                            # over 必須有 12-8
                            if sched[_ov6][_d6] != "12-8": continue
                            # under 必須是空班或上課日
                            if sched[_un6][_d6] not in ["", "上課"]: continue
                            # under 在這天能排 12-8
                            if not _can_12_8_nocheck6(_un6, _d6): continue

                            # ── 策略一：四格互換（兩人總班數不變）──────────
                            _four_done6 = False
                            for _wd6 in range(1, month_days + 1):
                                if _wd6 == _d6: continue
                                if sched[_un6][_wd6] != "D": continue
                                # ★ 保護預白/公差/國定必上班
                                if (_un6, _wd6) in _locked_set6: continue
                                if not _can_D_nocheck6(_ov6, _wd6): continue
                                if (_ov6, _d6) in _locked_set6: continue
                                # 執行四格互換
                                sched[_ov6][_d6] = ""
                                sched[_un6][_d6] = "12-8"
                                sched[_un6][_wd6] = ""
                                sched[_ov6][_wd6] = "D"
                                _swapped6 = True
                                _four_done6 = True
                                break

                            # ── 策略二：單格轉讓（under 仍低於目標時備用）──
                            if not _four_done6:
                                _un_worked6 = sum(1 for x in sched[_un6] if is_work(x))
                                if _un_worked6 < personal_targets.get(_un6, 0):
                                    if (_ov6, _d6) not in _locked_set6:
                                        sched[_ov6][_d6] = ""
                                        sched[_un6][_d6] = "12-8"
                                        _swapped6 = True

                if not _swapped6:
                    break   # 找不到可交換組合，停止

            # ── 第六步欠班診斷：清尾補班後仍有欠班時，收集原因 ─────────────
            _hol_set6d = set(sat_list6) | set(sun_list6) | set(nat_list6)
            _il6d = {"D": ["N"], "E": ["D","N","12-8"], "12-8": ["N"], "N": []}
            _s6_deficits = []
            for n_idx in ai_df.index:
                _worked6f = sum(1 for v in sched[n_idx] if is_work(v))
                _tgt6f    = personal_targets[n_idx]
                if _worked6f >= _tgt6f: continue
                _gap6f = _tgt6f - _worked6f
                _fill6f = "D"
                _pref6d = cache_pref[n_idx]
                if _pref6d:
                    _fill6f = "N" if "大夜" in _pref6d else ("E" if "小夜" in _pref6d else ("12-8" if "中" in _pref6d else "D"))
                # 掃描各日空格，統計被哪類規定擋住
                _has_empty = False
                _law_block = 0    # 連五 / 14日窗口
                _quota_block = 0  # 每日配額已滿
                _hol_block = 0    # 假日出勤資格不足
                _adj_block = 0    # 相鄰班別規定
                for _d6f in range(1, month_days + 1):
                    if sched[n_idx][_d6f] != "": continue
                    _has_empty = True
                    if cache_title[n_idx] in NO_HOL_SET and _d6f in _hol_set6d: continue
                    if cache_pref[n_idx] == "" and not can_work_holiday_check(
                            n_idx, _d6f, cache_can_sat6, cache_can_sun6, cache_can_nat6,
                            sat_list6, sun_list6, nat_list6):
                        _hol_block += 1; continue
                    # 相鄰班別
                    _ys6d = sched[n_idx][_d6f - 1] if _d6f > 1 else ""
                    _ts6d = sched[n_idx][_d6f + 1] if _d6f < month_days else ""
                    _yb6d = "D" if (_ys6d.startswith("D") or _ys6d in ("上課","公差")) else _ys6d
                    _tb6d = "D" if (_ts6d.startswith("D") or _ts6d in ("上課","公差")) else _ts6d
                    if (is_work(_ys6d) and _fill6f in _il6d.get(_yb6d, [])) or \
                       (is_work(_ts6d) and _tb6d in _il6d.get(_fill6f, [])):
                        _adj_block += 1; continue
                    # 連五
                    _sc6f = 1
                    for _bd6f in range(_d6f - 1, 0, -1):
                        if is_work(sched[n_idx][_bd6f]): _sc6f += 1
                        else: break
                    for _fd6f in range(_d6f + 1, month_days + 1):
                        if is_work(sched[n_idx][_fd6f]): _sc6f += 1
                        else: break
                    if _sc6f > 5: _law_block += 1; continue
                    # 14日窗口（含前月緩衝）
                    _14ok6f = True
                    _has_pb6f = bool((st.session_state.get("prev_month_buffer") or {}).get(str(ai_df.at[n_idx,"姓名"]).strip() if n_idx in ai_df.index else "", {}))
                    _w6f_min = (_d6f - 13) if _has_pb6f else max(1, _d6f - 13)
                    _w6f_max = min(_d6f, month_days - 13) if month_days >= 14 else 1
                    for _sd6f in range(_w6f_min, _w6f_max + 1):
                        _ed6f = min(month_days, _sd6f + 13)
                        _ww6f = sum(1 for _cd6f in range(_sd6f, _ed6f + 1)
                                    if _cd6f != _d6f and is_work(_xmonth_shift(n_idx, _cd6f, sched, ai_df, month_days)))
                        if _ww6f + 1 > 12: _14ok6f = False; break
                    if not _14ok6f: _law_block += 1; continue
                    # 配額
                    _rq6f = edited_quota_df[edited_quota_df["日期"] == str(_d6f)]
                    if not _rq6f.empty:
                        try:
                            if _fill6f in ("E", "N"):
                                if sum(1 for i in ai_df.index if sched[i][_d6f] == _fill6f) >= int(_rq6f.iloc[0][f"{_fill6f}班"]):
                                    _quota_block += 1; continue
                            elif _fill6f == "D" and cache_title[n_idx] not in NO_HOL_ADMIN:
                                if sum(1 for i in ai_df.index
                                       if isinstance(sched[i][_d6f], str) and sched[i][_d6f].startswith("D")
                                       and cache_title[i] not in NO_HOL_ADMIN) >= int(_rq6f.iloc[0]["D班"]):
                                    _quota_block += 1; continue
                            elif _fill6f == "12-8":
                                if sum(1 for i in ai_df.index if sched[i][_d6f] == "12-8") >= int(_rq6f.iloc[0]["12-8"]):
                                    _quota_block += 1; continue
                        except (KeyError, ValueError):
                            pass
                _reasons6f = []
                if not _has_empty:
                    _reasons6f.append("全月空格已耗盡（無可插入天）")
                else:
                    if _law_block > 0:   _reasons6f.append(f"勞基法連五/14日窗口限制（{_law_block} 格受限）")
                    if _adj_block > 0:   _reasons6f.append(f"相鄰班別規定（{_adj_block} 格受限）")
                    if _quota_block > 0: _reasons6f.append(f"每日人力配額已達上限（{_quota_block} 格受限）")
                    if _hol_block > 0:   _reasons6f.append(f"假日出勤資格不足（{_hol_block} 格受限）")
                    if not _reasons6f:   _reasons6f.append("複合限制（連五+相鄰+假日同時作用）")
                _s6_deficits.append({
                    "姓名":      ai_df.at[n_idx, "姓名"],
                    "應上班天數": _tgt6f,
                    "實際排入":  _worked6f,
                    "差距天數":  _gap6f,
                    "無法補足原因": "；".join(_reasons6f),
                })
            st.session_state.s6_deficits = _s6_deficits

            # ── 加班線均分（最後階段）─────────────────────────────────
            ot_days_count = {n_idx: 0 for n_idx in ai_df.index}
            ot_history    = {n_idx: [] for n_idx in ai_df.index}

            for d_int in range(1, month_days + 1):
                if d_int in holiday_days_list: continue   # 假日不排加班線
                d_workers = [idx for idx in ai_df.index if sched[idx][d_int] == "D"]
                if not d_workers: continue

                eligible_for_ot = []
                for idx in d_workers:
                    if cache_pref[idx] != "": continue        # 包班人員不排加班線
                    if cache_title[idx] in NO_HOL_ADMIN: continue  # 護理長/副護理長不排加班線
                    y_s = sched[idx][d_int - 1] if d_int > 1 else ""
                    if y_s == "12-8": continue                # 昨天排12-8者跳過
                    eligible_for_ot.append(idx)

                # 以 OT 天數最少者優先（確保所有人 OT 天數一致）
                eligible_for_ot.sort(key=lambda x: (ot_days_count[x], random.random()))
                num_slots = min(14, len(eligible_for_ot))
                selected_for_ot = eligible_for_ot[:num_slots]

                heavy_ot_yesterday = {idx for idx in selected_for_ot
                                      if (sched[idx][d_int - 1] if d_int > 1 else "") in ["D1","D2","D3"]}

                def get_avg_ot(idx):
                    return sum(ot_history[idx]) / len(ot_history[idx]) if ot_history[idx] else 15.0

                # 組長（3 人）獨立一組：OT 天數與所有人一致，線位在組長間互相均分
                leaders      = [x for x in selected_for_ot if cache_title[x] == "組長"]
                # 上課日護理師（非組長）：排後段線位
                class_nurses = [x for x in selected_for_ot
                                if x not in leaders and str(d_int) in class_days_map.get(x, [])]
                # 一般護理師（非包班、非護理長/副護理長、非組長、非上課）：線位共同均分
                regs         = [x for x in selected_for_ot
                                if x not in leaders and x not in class_nurses]

                # 組長：在組長子群中依均值均分線位（高均者先拿低線）
                leaders.sort(key=lambda x: get_avg_ot(x), reverse=True)
                # 一般護理師：同池依均值均分線位
                regs.sort(key=lambda x: get_avg_ot(x), reverse=True)

                slots = list(range(1, num_slots + 1))
                assignments = {}

                # 組長先拿 6 號以後的線位
                for p in leaders:
                    v_s = [s for s in slots if s >= 6]
                    chosen = min(v_s) if v_s else (max(slots) if slots else None)
                    if chosen is not None:
                        slots.remove(chosen); assignments[p] = chosen

                # 上課日護理師拿 8 號以後的線位
                for p in class_nurses:
                    v_s = [s for s in slots if s >= 8]
                    if v_s:
                        chosen = min(v_s)
                        slots.remove(chosen); assignments[p] = chosen

                # 一般護理師：從剩餘線位依序分配（均分平均線）
                for p in regs:
                    if not slots: break
                    valid = [s for s in slots if s >= 4] if p in heavy_ot_yesterday else slots
                    chosen = min(valid) if valid else min(slots)
                    slots.remove(chosen); assignments[p] = chosen

                for idx in d_workers:
                    if cache_pref[idx] != "": continue
                    if cache_title[idx] in NO_HOL_ADMIN: continue  # 護理長/副護理長不記錄 OT
                    if idx in assignments:
                        line = assignments[idx]
                        sched[idx][d_int] = f"D{line}"
                        ot_days_count[idx] += 1
                        ot_history[idx].append(line)
                    else:
                        ot_history[idx].append(15)

            # ── OT 天數絕對均等後處理（與個人應上班天數無關）──────────────
            # 目標：所有可排加班線的人（非包班、非護理長/副護理長）最終 OT 天數相差 ≤ 1
            eligible_ot_set = set()
            for idx in ai_df.index:
                if cache_pref[idx] != "": continue
                if cache_title[idx] in NO_HOL_ADMIN: continue
                eligible_ot_set.add(idx)

            _max_iter = 500  # 防止無窮迴圈
            for _it in range(_max_iter):
                _counts = {i: ot_days_count[i] for i in eligible_ot_set}
                if not _counts: break
                _max_v = max(_counts.values())
                _min_v = min(_counts.values())
                if _max_v - _min_v <= 1: break

                # 找天數最多的人（over）和天數最少的人（under）
                over_list  = [i for i, c in _counts.items() if c == _max_v]
                under_list = [i for i, c in _counts.items() if c == _min_v]

                _swapped = False
                for _over in over_list:
                    if _swapped: break
                    for _under in under_list:
                        if _swapped: break
                        # 找一天：_over 有 Dx（加班線），_under 有 D（一般白班），且 _under 昨天不是 12-8
                        for _d in range(1, month_days + 1):
                            if _d in holiday_days_list: continue
                            _v_over  = sched[_over][_d]
                            _v_under = sched[_under][_d]
                            # over 必須有加班線
                            if not (isinstance(_v_over, str) and _v_over.startswith("D") and len(_v_over) > 1):
                                continue
                            # under 必須有一般白班 D
                            if _v_under != "D": continue
                            # under 昨天不可以是 12-8
                            _y_under = sched[_under][_d - 1] if _d > 1 else ""
                            if _y_under == "12-8": continue
                            # 執行交換：under 取得加班線，over 退回一般 D
                            _line = int(_v_over[1:])
                            sched[_under][_d] = _v_over
                            sched[_over][_d]  = "D"
                            ot_days_count[_under] += 1
                            ot_days_count[_over]  -= 1
                            # 更新 ot_history（移除 over 的那條線記錄，under 新增）
                            if _line in ot_history[_over]:
                                ot_history[_over].remove(_line)
                            ot_history[_under].append(_line)
                            _swapped = True
                            break

                if not _swapped:
                    break  # 找不到可交換的組合，停止

            # ── 產出最終班表 ───────────────────────────────────────────
            final_df = pd.DataFrame({"姓名": ai_df["姓名"]})
            for d in range(1, month_days + 1):
                final_df[str(d)] = [sched[i][d] for i in ai_df.index]

            st.session_state.final_sched  = final_df
            st.session_state.ot_history   = ot_history
            st.session_state.ot_days_count = ot_days_count

            # ── 公平性診斷報告 ─────────────────────────────────────────
            explanation_data = []
            # max_target 只取常規人員（排除 NO_HOL_ADMIN 的平日全上目標，避免誤標一般護理師「請假較多」）
            _regular_targets = [personal_targets[i] for i in ai_df.index
                                 if str(ai_df.at[i, "職稱"]).strip() not in NO_HOL_ADMIN]
            max_target = max(_regular_targets) if _regular_targets else 20
            for idx, row in ai_df.iterrows():
                name        = row["姓名"]
                title       = str(row["職稱"]).strip()
                is_protected= str(row["孕/育嬰免夜班"]).strip() == "是"
                pref_raw    = str(row["包班意願"]).strip()
                target      = personal_targets[idx]
                reasons = []
                if title in ADMIN_TITLES:
                    if title in NO_HOL_ADMIN:
                        reasons.append(f"行政職({title})：僅排平日白班，不參與假日班輪替，亦不排加班線。")
                    else:
                        reasons.append(f"管理職({title})：不參與常規夜班與假日班輪替；加班天數與其他同仁一致，線位在組長群中互相均分。")
                if is_protected:     reasons.append("母性保護：依法規禁止安排小夜班(E)及大夜班(N)；12-8 中班可正常排入。")
                if pref_raw != "":   reasons.append(f"包班人員({pref_raw})：優先保障其專屬班別達標，極少數落差係受「連五/14休2防護」或「預休衝堂」擠壓。")
                if title not in NO_HOL_ADMIN and target < max_target:
                    reasons.append(f"請假天數較多：本月依法應上班僅 {target} 天 (全滿為 {max_target} 天)。在「最高連五天」及「11小時休息」限制下，產生了數學極限落差。")
                if not reasons:
                    reasons.append("常規人力：數據已受 AI 強制追平，若有1天內之微小差距，屬勞基法防護所產生之數學必然結果。")
                explanation_data.append({
                    "姓名": name,
                    "職稱": "🛡️ 母性保護" if is_protected else (title if title in ADMIN_TITLES else "護理師"),
                    "本月應上班": target,
                    "差異原因診斷報告": " | ".join(reasons)
                })
            faq_data = [
                {"姓名": "📊 【系統總結】", "職稱": "", "本月應上班": "", "差異原因診斷報告": "為什麼大家天數無法完全一樣？"},
                {"姓名": "💡 解答一",      "職稱": "", "本月應上班": "", "差異原因診斷報告": "【數學極限(鴿籠原理)】：若A休10天，B休18天，A依法可分攤的夜班或加班數，物理上絕對不可能與B相同。"},
                {"姓名": "💡 解答二",      "職稱": "", "本月應上班": "", "差異原因診斷報告": "【勞基法剛性防護】：為了絕對杜絕「D接N(休息僅8h)」、「連續上班超過5天」、「14天內未休2天」，系統寧可放掉均分，也絕對不產出違法班表。"},
            ]
            st.session_state.explanation_df = pd.concat([pd.DataFrame(explanation_data), pd.DataFrame(faq_data)], ignore_index=True)
            st.rerun()

    # --- 最終展示與警示掃描 ---
    if st.session_state.final_sched is not None:
        # 從 session_state 取回跨 rerun 所需變數
        personal_targets = st.session_state.personal_targets
        st.success("✅ 完整班表已產生！（12-8 中班 + 白班 + 加班線均分）")



        shortages_export = display_safety_radar(st.session_state.final_sched, edited_quota_df, ai_df)
        st.session_state.shortages_export = shortages_export

        # 建立顯示用副本：空白格與預休(O)標示為「休」，特殊假別保留原代碼
        _REST_VALS_D = {"", "O"}
        _day_cols = [str(d) for d in range(1, month_days + 1)]
        _display_df = st.session_state.final_sched.copy()
        for _c in _day_cols:
            _display_df[_c] = _display_df[_c].apply(lambda v: "休" if str(v).strip() in _REST_VALS_D else v)
        # view 模式專用縮寫副本（不影響 edit 模式的 _display_df）
        _view_df = _display_df.copy()
        for _c in _day_cols:
            _view_df[_c] = _view_df[_c].apply(abbrev_display)
        # 預白班：D 格標為 Dx（視覺區分預排白班）
        _view_df = apply_prewhite_dx(_view_df, ai_df, month_days)
        styled_final_df = build_schedule_with_counts(_view_df, st.session_state.final_sched, _day_cols, ai_df)

        stats = []
        for idx, row in ai_df.iterrows():
            _title_s     = str(row["職稱"]).strip()
            is_leader    = _title_s in ADMIN_TITLES
            is_no_hol    = _title_s in NO_HOL_SET   # 護理長/副護理長/組長：不參與假日/夜班分配
            pref_raw     = str(row["包班意願"]).strip()
            pref_norm    = "N" if "大夜" in pref_raw else ("E" if "小夜" in pref_raw else ("12-8" if "中" in pref_raw else ""))
            is_protected = str(row["孕/育嬰免夜班"]).strip() == "是"

            s_vals        = list(st.session_state.final_sched.iloc[idx].values[1:])
            night_count   = s_vals.count("E") + s_vals.count("N") + s_vals.count("12-8")
            holiday_count = 0 if is_no_hol else sum(1 for d_chk in range(1, month_days + 1) if (d_chk in holiday_days_list) and is_work(s_vals[d_chk-1]))

            ot_history_s   = st.session_state.get("ot_history", {})
            ot_days_count_s= st.session_state.get("ot_days_count", {})
            avg_ot_line    = round(sum(ot_history_s[idx]) / len(ot_history_s[idx]), 1) if ot_history_s[idx] else "-"
            worked_count   = sum(1 for v in s_vals if is_work(v))
            target         = personal_targets[idx]

            # 休假統計
            off_count      = s_vals.count("O")                                          # 預休 / 長假
            sp_leave_vals  = [str(v).strip() for v in s_vals if str(v).strip() not in ["","nan","O","上課"] and not is_work(str(v).strip())]  # 公假、喪假等
            sp_leave_count = len(sp_leave_vals)
            sp_leave_types = "、".join(sorted(set(sp_leave_vals))) if sp_leave_vals else "-"

            status_msg = "✅ 達標"
            if worked_count < target: status_msg = f"⚠️ 欠班 {target - worked_count} 天"
            elif worked_count > target: status_msg = f"🟢 超班 {worked_count - target} 天"

            ded_status = "-"
            if pref_norm:
                actual_count = s_vals.count(pref_norm)
                ded_status = f"✅ 達標 ({actual_count}班)" if actual_count >= 15 else f"⚠️ 未達標 ({actual_count}班)"

            # is_no_hol 包含組長，但組長實際有加班線，改用 NO_HOL_ADMIN 排除加班顯示
            _no_ot_display = pref_norm or (_title_s in NO_HOL_ADMIN)
            stats.append({
                "姓名":        row["姓名"],
                "職稱":        "🛡️ 母性保護" if is_protected else (str(row["職稱"]).strip() if is_leader else "護理師"),
                "預休(O)":     off_count,
                "特殊假別":    sp_leave_types,
                "休假合計":    off_count + sp_leave_count,
                "加班天數":    "-" if _no_ot_display else ot_days_count_s[idx],
                "平均線位":    "-" if _no_ot_display else avg_ot_line,
                "12-8班數":   "-" if is_no_hol else s_vals.count("12-8"),
                "夜班數":      "-" if is_no_hol else night_count,
                "包班狀態":    ded_status,
                "假日出勤":    "-" if is_no_hol else holiday_count,
                "目標上班":    target,
                "實際上班":    worked_count,
                "狀態":        status_msg
            })

        stats_df = pd.DataFrame(stats).sort_values(by=["職稱", "加班天數"], ascending=[False, False])

        # ── 四周變形工時審查 ────────────────────────────────────
        per_week_df, violation_df = build_four_week_review(
            st.session_state.final_sched, ai_df, month_days,
            prev_buffer=st.session_state.get("prev_month_buffer")
        )

        # ── 提前產出 Excel（在 tab1 修改 final_sched 之前，確保班表與統計來自同一份資料）──
        _export_sched6 = apply_prewhite_dx(st.session_state.final_sched.copy(), ai_df, month_days)
        output = build_colored_excel(
            _export_sched6,
            stats_df,
            st.session_state.explanation_df,
            shortages_export,
            month_days,
            per_week_df=per_week_df,
            violation_df=violation_df
        )

        tab1, tab2, tab3, tab4 = st.tabs([
            "✨ 全彩排班表預覽",
            "⚖️ 公平性與變形工時結算",
            "🔍 公平性差異診斷報告",
            "📋 勞基法四周變形工時審查"
        ])
        with tab1:
            _edit_final = st.checkbox("🖊️ 開啟手動編輯模式", value=False, key="chk_edit_final_sched")
            if _edit_final:
                st.caption("💡 直接點擊儲存格輸入班別代碼（D/E/N/12-8/O/公假 等），空白格顯示為「休」，修改即時儲存")
                _edited_final = st.data_editor(
                    _display_df,
                    column_config=make_sched_col_config(month_days),
                    use_container_width=True, hide_index=True, key="edit_final_sched"
                )
                # 將顯示用的「休」反轉回空字串存回 final_sched
                _save_final = _edited_final.copy()
                for _c in _day_cols:
                    _save_final[_c] = _save_final[_c].apply(lambda v: "" if str(v).strip() == "休" else v)
                st.session_state.final_sched = _save_final
            else:
                st.dataframe(styled_final_df, use_container_width=True)
            # ── 欠班警示（排在全彩班表預覽下方）────────────────────────
            _s6d = st.session_state.get("s6_deficits", [])
            if _s6d:
                st.error(
                    f"⚠️ **欠班警示（{len(_s6d)} 人）**：以下人員在遵守勞基法及每日人力配額上限的前提下，"
                    f"清尾補班後仍無法達到應上班天數，請至手動編輯模式補排，或調整配額設定後重新計算："
                )
                st.dataframe(pd.DataFrame(_s6d), use_container_width=True, hide_index=True)
            else:
                st.success("✅ 全體人員均已達到應上班天數，無欠班！")
        with tab2:
            st.dataframe(stats_df, use_container_width=True)
        with tab3:
            st.dataframe(st.session_state.explanation_df, use_container_width=True)
        with tab4:
            st.write("#### ⚖️ 勞基法 §30-1 四周變形工時合法性審查")
            st.caption("審查依據：① 任意7天窗格工作天 ≤ 6；② 任意14天窗格工作天 ≤ 12（14休2）；③ 任意28天窗格工作天 ≤ 24（四週制）")
            st.dataframe(per_week_df, use_container_width=True)
            total_violators = len([r for _, r in per_week_df.iterrows() if "🚨" in str(r.get("合法判斷", ""))])
            if total_violators == 0:
                st.success("✅ 全體人員均符合四周變形工時規定，無違規！")
            else:
                st.error(f"🚨 共 {total_violators} 位人員有違規，詳見下方違規明細")
            st.write("##### 違規明細")
            st.dataframe(violation_df, use_container_width=True)

        st.write("---")
        col_btn_back, col_btn_download, col_btn_go7, col_btn_reset = st.columns([1, 2, 2, 1])
        with col_btn_back:
            if st.button("⬅️ 退回重排12-8與加班線", type="secondary"):
                st.session_state.final_sched = None
                st.rerun()

        with col_btn_download:
            st.download_button(
                label="📥 下載班表",
                data=output.getvalue(),
                file_name="Final_Schedule_Colored.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        with col_btn_go7:
            if st.button("🗓️ 前往步驟七：假別分類標示", type="primary"):
                st.session_state.step = 7
                st.session_state.classified_sched = None
                st.rerun()

        with col_btn_reset:
            if st.button("🔄 全部重來", type="secondary"):
                for key in ["step", "base_sched", "pack_sched", "night_sched", "d_sched", "twelve_sched", "final_sched", "classified_sched", "custom_targets", "ai_df"]:
                    if key in st.session_state: del st.session_state[key]
                st.rerun()
# ==========================================
# 📍 第七步：休假假別自動分類（勞基法標示）
# ==========================================
if st.session_state.step >= 7:
    st.divider()
    st.header("7️⃣ 第七步：休假假別自動分類（勞基法標示）")

    ai_df        = st.session_state.ai_df
    month_days   = st.session_state.month_days
    first_wday   = st.session_state.first_wday
    sat_set      = set(st.session_state.saturdays_list)
    sun_set      = set(st.session_state.sundays_list)
    nat_set      = set(st.session_state.nat_holidays_list)

    if st.session_state.classified_sched is None:
        with st.spinner("正在依勞基法自動分類假別…"):
            src = st.session_state.final_sched.copy()
            classified = src.copy()

            # 從 ai_df 重建 pre_type_map：知道哪些 O 是「預假」，哪些是「預長假」，哪些 D 是「預白」
            pre_type_map7 = {}  # (row_i, d) -> "預假" | "預長假" | "預白"
            for row_i7, (idx7, row7) in enumerate(ai_df.iterrows()):
                for d_str in str(row7.get("預休日期", "")).split(","):
                    if d_str.strip().isdigit() and 1 <= int(d_str.strip()) <= month_days:
                        pre_type_map7[(row_i7, int(d_str.strip()))] = "預假"
                for d_str in str(row7.get("預約長假日期", "")).split(","):
                    if d_str.strip().isdigit() and 1 <= int(d_str.strip()) <= month_days:
                        pre_type_map7[(row_i7, int(d_str.strip()))] = "預長假"
                for d_str in str(row7.get("預白日期", "")).split(","):
                    if d_str.strip().isdigit() and 1 <= int(d_str.strip()) <= month_days:
                        pre_type_map7[(row_i7, int(d_str.strip()))] = "預白"

            for row_i in range(len(ai_df)):
                week_has_jiqi = {}
                _title_r7 = str(ai_df.iloc[row_i].get("職稱", "")).strip()
                _is_nohol_r7 = _title_r7 in NO_HOL_ADMIN  # 傷兵/助理：平日不標休假

                # ── 第一回：依日期類型分配假別 ──
                # 規則：
                #   上班班別（D/E/N/12-8）→ 不動
                #   國定假日 → 一律標「國定」（覆蓋 O、特殊假別、空白）
                #   週日（例假日）→ 標「例假」（覆蓋同上）
                #   週六（休息日）→ 標「休假」（覆蓋同上）
                #   平日（非六日非國定）→
                #       O       → 依 pre_type_map7 標「預假」或「預長假」
                #       空白    → 標「休假」（系統排定休息）
                #                 ※ 傷兵/助理（NO_HOL_ADMIN）例外：平日不標休假
                #       其他（公假/喪假等特殊假別）→ 保留原假別名稱
                for d in range(1, month_days + 1):
                    col  = str(d)
                    cidx = classified.columns.get_loc(col)
                    cell = str(classified.iat[row_i, cidx]).strip()

                    if is_work(cell):
                        # 預白班：D 格保留 Dx 特徵顯示
                        if cell == "D" and pre_type_map7.get((row_i, d)) == "預白":
                            classified.iat[row_i, cidx] = "Dx"
                        continue  # 上班班別不動

                    w = (first_wday + d - 1) // 7
                    if d in nat_set:
                        classified.iat[row_i, cidx] = "國定"
                    elif d in sun_set:
                        classified.iat[row_i, cidx] = "例假"
                        week_has_jiqi[w] = True
                    elif d in sat_set:
                        classified.iat[row_i, cidx] = "休假"
                    else:
                        # 平日：依實際假別顯示
                        if cell == "O":
                            classified.iat[row_i, cidx] = pre_type_map7.get(
                                (row_i, d), "預假")
                        elif cell == "":
                            # 傷兵/助理 平日不排休假（排班階段已強制填滿平日）
                            if not _is_nohol_r7:
                                classified.iat[row_i, cidx] = "休假"
                        # 其他（公假、喪假等特殊假別）→ 保留原值不動

                # ── 第二回：確保每週至少一個例假（§36）──
                for d in range(1, month_days + 1):
                    w = (first_wday + d - 1) // 7
                    v = str(classified.iat[row_i, classified.columns.get_loc(str(d))]).strip()
                    if v == "例假":
                        week_has_jiqi[w] = True

                for d in range(1, month_days + 1):
                    w = (first_wday + d - 1) // 7
                    if week_has_jiqi.get(w): continue
                    cidx = classified.columns.get_loc(str(d))
                    v = str(classified.iat[row_i, cidx]).strip()
                    if v == "休假":
                        classified.iat[row_i, cidx] = "例假"
                        week_has_jiqi[w] = True

            st.session_state.classified_sched = classified
            st.rerun()

    if st.session_state.classified_sched is not None:
        classified_df = st.session_state.classified_sched
        _day_cols7    = [str(d) for d in range(1, month_days + 1)]

        st.caption(
            "🟩 上班（D/E/N/12-8）　"
            "🟥 **例**（例假，§36 週日）　"
            "⬜ **休**（休假日，週六/系統排定）　"
            "🟧 **國**（國定假日，§37）　"
            "🔵 **預**（預假/預長假）　"
            "🩷 首字（特殊假別）"
        )

        with st.expander("📄 假別分類班表", expanded=True):
            _edit7 = st.checkbox("🖊️ 開啟手動編輯模式", value=False, key="chk_edit_classified")
            if _edit7:
                _edited7 = st.data_editor(
                    classified_df,
                    column_config=make_sched_col_config(month_days),
                    use_container_width=True, hide_index=True, key="edit_classified_sched"
                )
                st.session_state.classified_sched = _edited7
            else:
                _disp7 = classified_df.copy()
                for _c in _day_cols7:
                    _disp7[_c] = _disp7[_c].apply(abbrev_display)
                st.dataframe(
                    build_schedule_with_counts(_disp7, st.session_state.final_sched, _day_cols7, ai_df),
                    use_container_width=True
                )

        # ── 假別統計 ──
        st.write("#### 📊 各假別統計（每人）")
        leave_stats = []
        for row_i, (idx, row) in enumerate(ai_df.iterrows()):
            s_vals = [
                str(classified_df.iat[row_i, classified_df.columns.get_loc(str(d))]).strip()
                for d in range(1, month_days + 1)
            ]
            special_cnt = sum(
                1 for v in s_vals
                if v not in ["", "例假", "休假", "國定", "O", "預假", "預長假"]
                and not is_work(v)
            )
            sp_types = "、".join(sorted(set(
                v for v in s_vals
                if v not in ["", "例假", "休假", "國定", "O", "預假", "預長假"]
                and not is_work(v)
            ))) or "-"
            leave_stats.append({
                "姓名":        row["姓名"],
                "上班天數":    sum(1 for v in s_vals if is_work(v)),
                "例假":        s_vals.count("例假"),
                "休假日":      s_vals.count("休假"),
                "國定假日":    s_vals.count("國定"),
                "預假":        s_vals.count("預假"),
                "預長假":      s_vals.count("預長假"),
                "特殊假別筆數": special_cnt,
                "特殊假別明細": sp_types,
            })
        leave_df = pd.DataFrame(leave_stats)
        st.dataframe(leave_df, use_container_width=True, hide_index=True)

        # ── 例假合規摘要 ──
        st.write("#### ⚖️ 例假合規性摘要（§36）")
        violation_rows7 = []
        for row_i, (idx, row) in enumerate(ai_df.iterrows()):
            s_vals = [
                str(classified_df.iat[row_i, classified_df.columns.get_loc(str(d))]).strip()
                for d in range(1, month_days + 1)
            ]
            short_weeks = []
            for w_start in range(0, month_days, 7):
                w_days = s_vals[w_start: w_start + 7]
                if not any(v == "例假" for v in w_days):
                    short_weeks.append(f"第 {w_start+1}～{min(w_start+7, month_days)} 日")
            if short_weeks:
                violation_rows7.append({"姓名": row["姓名"], "缺例假週段": "、".join(short_weeks)})
        if violation_rows7:
            st.warning(f"⚠️ {len(violation_rows7)} 位人員有週段缺少例假，請主管確認")
            st.dataframe(pd.DataFrame(violation_rows7), use_container_width=True, hide_index=True)
        else:
            st.success("✅ 全員每週均至少有一天例假，符合 §36 規定。")

        # ── Excel 下載 ──
        _exp_df  = st.session_state.explanation_df if "explanation_df" in st.session_state else pd.DataFrame()
        _short_e = st.session_state.shortages_export if "shortages_export" in st.session_state else pd.DataFrame()
        _cl_output = build_colored_excel(
            classified_df, leave_df, _exp_df, _short_e, month_days
        )

        st.write("---")
        col7_back, col7_dl, col7_reset = st.columns([1, 3, 1])
        with col7_back:
            if st.button("⬅️ 退回重新分類", type="secondary"):
                st.session_state.classified_sched = None
                st.rerun()
        with col7_dl:
            st.download_button(
                label="📥 下載假別分類班表 Excel",
                data=_cl_output.getvalue(),
                file_name="Classified_Schedule.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
        with col7_reset:
            if st.button("🔄 全部重來 ", type="secondary", key="reset_step7"):
                for key in ["step", "base_sched", "pack_sched", "night_sched", "d_sched",
                            "twelve_sched", "final_sched", "classified_sched", "custom_targets", "ai_df"]:
                    if key in st.session_state: del st.session_state[key]
                st.rerun()
