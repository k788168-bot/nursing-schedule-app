"""
護理白班補足系統（獨立版）
============================
輸入：人工班表 Excel（已含 E/N/12-8）+ 人員資料 + 每日配額
輸出：補滿 D 班 + 加班線的完整班表 Excel

使用方式：
  streamlit run fill_d_system.py
"""

import streamlit as st
import pandas as pd
import random
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════
#  常數
# ══════════════════════════════════════════
ILLEGAL_NEXT   = {"D": ["N"], "E": ["D", "N", "12-8"], "12-8": ["N"], "N": []}
NO_HOL_ADMIN   = {"護理長", "副護理長", "助理", "傷兵"}
NO_HOL_SET     = NO_HOL_ADMIN | {"組長"}
ADMIN_TITLES   = {"組長", "護理長", "副護理長", "助理", "傷兵"}
_CORE_SHIFTS   = {"D", "E", "N", "12-8"}
_GROUP_SHIFT_CAP = {"12-8": {"A": 3, "B": 3}, "E": {"A": 2, "B": 2}}

# ══════════════════════════════════════════
#  基本判斷函數
# ══════════════════════════════════════════
def is_work(val):
    v = str(val).strip(); vu = v.upper()
    if vu in ("E★", "E*", "N★", "N*"): return True
    return vu in ["D","E","N","12-8","上課","公差","公"] or vu.startswith("D")

def _shift_norm(v):
    vs = str(v).strip(); vu = vs.upper()
    if not vs or vu in ("O","NAN",""): return None
    if vu in ("E★","E*"): return "E"
    if vu in ("N★","N*"): return "N"
    if vu.startswith("D") or vs in ("上課","公差","公"): return "D"
    if vs in _CORE_SHIFTS: return vs
    return None

def norm_display(v):
    """正規化顯示用班別（用於解析人工班表）"""
    vu = str(v).strip().upper()
    if not vu or vu in ("NAN","","休","O","例","例假","V","預"): return ""
    if vu in ("E★","E*"): return "E"
    if vu in ("N★","N*"): return "N"
    if vu == "12-8": return "12-8"
    if vu.startswith("D") and vu != "DX": return v.strip()  # 保留 D1-D14
    return ""  # 其他（休/O/特殊假）都當空格

def is_night(v):
    vu = str(v).strip().upper()
    return vu in ("E","E★","E*","N","N★","N*","12-8")

def week_variety_ok(sched, idx, s, day, first_wday, month_days):
    wday = (first_wday + day - 1) % 7
    days_since_sun = (wday + 1) % 7
    ws = max(1, day - days_since_sun)
    we = min(month_days, ws + 6)
    s_norm = _shift_norm(s)
    if s_norm is None: return True
    types = {s_norm}
    for dd in range(ws, we + 1):
        if dd == day: continue
        n = _shift_norm(sched[idx][dd])
        if n: types.add(n)
    return len(types) <= 2

def group_cap_ok(idx, s, day, sched, cache_group):
    grp = cache_group.get(idx, "")
    if grp not in ("A","B"): return True
    caps = _GROUP_SHIFT_CAP.get(s)
    if not caps: return True
    ca, cb = caps.get("A",999), caps.get("B",999)
    curr_a = sum(1 for i,sv in sched.items() if cache_group.get(i,"")=="A" and sv[day]==s)
    curr_b = sum(1 for i,sv in sched.items() if cache_group.get(i,"")=="B" and sv[day]==s)
    if grp == "A":
        return curr_a < ca or curr_b < cb
    return curr_b < cb

def can_work_holiday(idx, day, cache_sat, cache_sun, cache_nat, sat_set, sun_set, nat_set):
    is_sat = day in sat_set; is_sun = day in sun_set; is_nat = day in nat_set
    if not (is_sat or is_sun or is_nat): return True
    can_sun_v = cache_sun.get(idx, False); can_nat_v = cache_nat.get(idx, False); can_sat_v = cache_sat.get(idx, False)
    if can_sun_v and can_nat_v: return True
    if can_sat_v: return is_sat
    return False

# ══════════════════════════════════════════
#  目標出勤天數計算
# ══════════════════════════════════════════
def calc_personal_target(row, month_days, hol_set, target_off):
    title = str(row.get("職稱","")).strip()
    if title in NO_HOL_SET:
        weekday_cnt = sum(1 for d in range(1, month_days+1) if d not in hol_set)
        pre_o = str(row.get("預休日期","")).strip()
        pre_o_days = sum(1 for d in pre_o.split(",") if d.strip().isdigit()
                         and 1<=int(d.strip())<=month_days and int(d.strip()) not in hol_set) if pre_o else 0
        sp = str(row.get("特殊假別","")).strip()
        sp_days = sum(1 for item in sp.split(",") if item.strip() and item.strip().split(":")[0].strip().isdigit()
                      and 1<=int(item.strip().split(":")[0])<=month_days) if sp else 0
        return max(0, weekday_cnt - pre_o_days - sp_days)
    else:
        sp = str(row.get("特殊假別","")).strip()
        sp_days = sum(1 for item in sp.split(",") if item.strip() and item.strip().split(":")[0].strip().isdigit()
                      and 1<=int(item.strip().split(":")[0])<=month_days) if sp else 0
        ll = str(row.get("預約長假日期","")).strip()
        ll_days = sum(1 for d in ll.split(",") if d.strip().isdigit()
                      and 1<=int(d.strip())<=month_days and int(d.strip()) not in hol_set) if ll else 0
        return max(0, month_days - target_off - sp_days - ll_days)

# ══════════════════════════════════════════
#  解析人工班表 Excel（支援跨月多列標題格式）
# ══════════════════════════════════════════
def parse_manual_schedule(file, ai_names, month_days):
    raw = pd.read_excel(file, sheet_name=0, header=None)
    hdr_idx = None
    for ri in range(min(5, len(raw))):
        if "姓名" in [str(v).strip() for v in raw.iloc[ri].values]:
            hdr_idx = ri; break
    if hdr_idx is None:
        return None, "找不到「姓名」欄，請確認班表格式"

    hdr = [str(v).strip() for v in raw.iloc[hdr_idx].values]
    name_ci = hdr.index("姓名")

    date_cols = []
    for ci, v in enumerate(hdr):
        try: date_cols.append((ci, int(float(v))))
        except: pass

    dates = [d for _, d in date_cols]
    split = next((i for i in range(1, len(dates)) if dates[i] < dates[i-1] and dates[i] <= 5), None)
    this_month = [(ci, d) for ci, d in (date_cols[split:] if split else date_cols) if 1 <= d <= month_days]

    # 結果：{nurse_idx: {day: shift_value}}
    result = {i: {} for i in range(len(ai_names))}

    for ri in range(hdr_idx + 1, len(raw)):
        name = str(raw.iloc[ri, name_ci]).strip()
        if not name or name in ("nan","NaN","None","姓名"): continue
        try: float(name); continue
        except: pass
        if name not in ai_names: continue
        idx = ai_names.index(name)
        for ci, d in this_month:
            v = str(raw.iloc[ri, ci]).strip()
            if v in ("nan","NaN","","V"): continue
            result[idx][d] = v

    return result, None

# ══════════════════════════════════════════
#  D 班補足主算法
# ══════════════════════════════════════════
def fill_d_shifts(sched, ai_df, quota_df, month_days,
                  sat_set, sun_set, nat_set, hol_set,
                  first_wday, target_off):
    cache_title  = {i: str(row.get("職稱","")).strip() for i, row in ai_df.iterrows()}
    cache_pref   = {i: str(row.get("包班意願","")).strip() for i, row in ai_df.iterrows()}
    cache_sat    = {i: str(row.get("能上週六","")).strip()=="是" for i, row in ai_df.iterrows()}
    cache_sun    = {i: str(row.get("能上週日","")).strip()=="是" for i, row in ai_df.iterrows()}
    cache_nat    = {i: str(row.get("能上國定假日","")).strip()=="是" for i, row in ai_df.iterrows()}
    cache_group  = {i: str(row.get("組別","")).strip().upper() for i, row in ai_df.iterrows()}
    cache_class  = {i: [s.strip() for s in str(row.get("上課日期","")).split(",") if s.strip().isdigit()]
                    for i, row in ai_df.iterrows()}

    personal_targets = {i: calc_personal_target(row, month_days, hol_set, target_off)
                        for i, row in ai_df.iterrows()}

    def can_work(idx, day):
        """D 班排班資格綜合判斷"""
        if sched[idx][day] not in ("", "上課"): return False
        t = cache_title[idx]
        if t in NO_HOL_SET and day in hol_set: return False
        if cache_pref[idx] == "" and not can_work_holiday(idx, day, cache_sat, cache_sun, cache_nat,
                                                           sat_set, sun_set, nat_set): return False
        if sum(1 for v in sched[idx] if is_work(v)) >= personal_targets[idx]: return False
        # §34 接續
        y = (sched[idx][day-1] or "") if day > 1 else ""
        t_next = (sched[idx][day+1] or "") if day < month_days else ""
        y_base = "D" if y.startswith("D") or y in ("上課","公差") else y
        t_base = "D" if t_next.startswith("D") or t_next in ("上課","公差") else t_next
        if is_work(y) and "D" in ILLEGAL_NEXT.get(y_base, []): return False
        if is_work(t_next) and t_base in ILLEGAL_NEXT.get("D", []): return False
        # 連五
        sc = 1
        for bd in range(day-1, 0, -1):
            if is_work(sched[idx][bd]): sc += 1
            else: break
        for fd in range(day+1, month_days+1):
            if is_work(sched[idx][fd]): sc += 1
            else: break
        if sc > 5: return False
        # 14日窗口
        w_min = max(1, day-13)
        for sd in range(w_min, min(day, month_days-13)+1):
            ed = min(month_days, sd+13)
            ww = sum(1 for cd in range(sd, ed+1) if cd!=day and is_work(sched[idx][cd])) + 1
            if ww > 12: return False
        # week_variety
        if not week_variety_ok(sched, idx, "D", day, first_wday, month_days): return False
        return True

    def d_quota_ok(idx, day, buffer=3):
        if cache_title[idx] in NO_HOL_ADMIN: return True
        row_q = quota_df[quota_df["日期"] == str(day)]
        if row_q.empty: return True
        try:
            req = int(row_q.iloc[0]["D班"])
            curr = sum(1 for i in ai_df.index
                       if isinstance(sched[i][day], str) and sched[i][day].startswith("D")
                       and cache_title[i] not in NO_HOL_ADMIN)
            limit = req if day in hol_set else req + buffer
            return curr < limit
        except: return True

    def day_priority(idx, day):
        y = sched[idx][day-1] if day > 1 else ""
        t = sched[idx][day+1] if day < month_days else ""
        if is_work(y) and is_work(t): return 0
        if is_work(y) or is_work(t): return 1
        return 2

    # ── Pass 1：嚴守配額，填連段 ──
    for idx in sorted(ai_df.index, key=lambda i: sum(1 for v in sched[i] if is_work(v)) - personal_targets[i]):
        if sum(1 for v in sched[idx] if is_work(v)) >= personal_targets[idx]: continue
        for day in sorted(range(1, month_days+1), key=lambda d: day_priority(idx, d)):
            if sum(1 for v in sched[idx] if is_work(v)) >= personal_targets[idx]: break
            if can_work(idx, day) and d_quota_ok(idx, day):
                sched[idx][day] = "D"

    # ── Pass 2：放寬週多樣性，補足仍欠班者 ──
    def can_work_relaxed(idx, day):
        if sched[idx][day] not in ("", "上課"): return False
        t = cache_title[idx]
        if t in NO_HOL_SET and day in hol_set: return False
        if cache_pref[idx] == "" and not can_work_holiday(idx, day, cache_sat, cache_sun, cache_nat,
                                                           sat_set, sun_set, nat_set): return False
        if sum(1 for v in sched[idx] if is_work(v)) >= personal_targets[idx]: return False
        y = (sched[idx][day-1] or "") if day > 1 else ""
        t_next = (sched[idx][day+1] or "") if day < month_days else ""
        y_base = "D" if y.startswith("D") or y in ("上課","公差") else y
        t_base = "D" if t_next.startswith("D") or t_next in ("上課","公差") else t_next
        if is_work(y) and "D" in ILLEGAL_NEXT.get(y_base, []): return False
        if is_work(t_next) and t_base in ILLEGAL_NEXT.get("D", []): return False
        sc = 1
        for bd in range(day-1, 0, -1):
            if is_work(sched[idx][bd]): sc += 1
            else: break
        for fd in range(day+1, month_days+1):
            if is_work(sched[idx][fd]): sc += 1
            else: break
        if sc > 5: return False
        for sd in range(max(1, day-13), min(day, month_days-13)+1):
            ed = min(month_days, sd+13)
            if sum(1 for cd in range(sd, ed+1) if cd!=day and is_work(sched[idx][cd])) + 1 > 12: return False
        return True

    for idx in sorted(ai_df.index, key=lambda i: sum(1 for v in sched[i] if is_work(v)) - personal_targets[i]):
        if sum(1 for v in sched[idx] if is_work(v)) >= personal_targets[idx]: continue
        for day in sorted(range(1, month_days+1), key=lambda d: day_priority(idx, d)):
            if sum(1 for v in sched[idx] if is_work(v)) >= personal_targets[idx]: break
            if can_work_relaxed(idx, day) and d_quota_ok(idx, day, buffer=5):
                sched[idx][day] = "D"

    return sched, personal_targets

# ══════════════════════════════════════════
#  加班線分配
# ══════════════════════════════════════════
def assign_ot_lines(sched, ai_df, month_days, hol_set):
    cache_pref   = {i: str(row.get("包班意願","")).strip() for i, row in ai_df.iterrows()}
    cache_title  = {i: str(row.get("職稱","")).strip() for i, row in ai_df.iterrows()}
    cache_leader = {i: str(row.get("控台資格","")).strip() for i, row in ai_df.iterrows()}
    cache_class  = {i: [s.strip() for s in str(row.get("上課日期","")).split(",") if s.strip().isdigit()]
                    for i, row in ai_df.iterrows()}

    ot_count = {i: 0 for i in ai_df.index}
    ot_hist  = {i: [] for i in ai_df.index}

    for day in range(1, month_days+1):
        if day in hol_set: continue
        d_workers = [i for i in ai_df.index if sched[i][day] == "D"]
        if not d_workers: continue

        elig = [i for i in d_workers
                if cache_pref[i] == ""
                and cache_title[i] not in NO_HOL_ADMIN
                and (sched[i][day-1] if day > 1 else "") != "12-8"]
        elig.sort(key=lambda x: (ot_count[x], random.random()))
        num = min(14, len(elig))
        sel = elig[:num]

        def avg_line(i): return sum(ot_hist[i])/len(ot_hist[i]) if ot_hist[i] else 15.0

        leaders = [x for x in sel if cache_leader[x]]
        cls_lst = [x for x in sel if x not in leaders and str(day) in cache_class.get(x, [])]
        regs    = [x for x in sel if x not in leaders and x not in cls_lst]
        leaders.sort(key=avg_line, reverse=True)
        regs.sort(key=avg_line, reverse=True)

        slots = list(range(1, num+1))
        assign = {}
        for p in leaders:
            vs = [s for s in slots if s >= 6]
            ch = min(vs) if vs else (max(slots) if slots else None)
            if ch: slots.remove(ch); assign[p] = ch
        for p in cls_lst:
            vs = [s for s in slots if s >= 8]
            if vs: ch = min(vs); slots.remove(ch); assign[p] = ch
        for p in regs:
            if not slots: break
            ch = min(slots); slots.remove(ch); assign[p] = ch

        for i in d_workers:
            if cache_pref[i] != "" or cache_title[i] in NO_HOL_ADMIN: continue
            if i in assign:
                sched[i][day] = f"D{assign[i]}"
                ot_count[i] += 1
                ot_hist[i].append(assign[i])
            else:
                ot_hist[i].append(15)

    # OT 均等後處理
    elig_set = {i for i in ai_df.index if cache_pref[i]=="" and cache_title[i] not in NO_HOL_ADMIN}
    for _ in range(500):
        cts = {i: ot_count[i] for i in elig_set}
        if not cts or max(cts.values()) - min(cts.values()) <= 1: break
        ov_l = [i for i,c in cts.items() if c==max(cts.values())]
        un_l = [i for i,c in cts.items() if c==min(cts.values())]
        sw = False
        for ov in ov_l:
            if sw: break
            for un in un_l:
                if sw: break
                for day in range(1, month_days+1):
                    if day in hol_set: continue
                    vo = sched[ov][day]; vu = sched[un][day]
                    if not (isinstance(vo,str) and vo.startswith("D") and len(vo)>1): continue
                    if vu != "D": continue
                    if (sched[un][day-1] if day>1 else "") == "12-8": continue
                    ln = int(vo[1:])
                    sched[un][day] = vo; sched[ov][day] = "D"
                    ot_count[un] += 1; ot_count[ov] -= 1
                    if ln in ot_hist[ov]: ot_hist[ov].remove(ln)
                    ot_hist[un].append(ln)
                    sw = True; break
        if not sw: break

    return sched, ot_count

# ══════════════════════════════════════════
#  Excel 彩色輸出
# ══════════════════════════════════════════
COLOR_MAP = {
    "D":    ("d4edda","155724"), "E": ("fff3cd","856404"),
    "N":    ("e2d9f3","4a148c"), "12-8": ("cce5ff","004085"),
    "O":    ("dee2e6","495057"), "休": ("f1f3f5","adb5bd"),
    "上課": ("ffeeba","856404"), "公差": ("fad7a0","784212"),
    "公":   ("fad7a0","784212"),
}

def get_cell_color(v):
    vu = str(v).strip().upper()
    if vu.startswith("D") and vu != "DX": return COLOR_MAP["D"]
    if vu in ("E","E★","E*"): return COLOR_MAP["E"]
    if vu in ("N","N★","N*"): return COLOR_MAP["N"]
    if vu == "12-8": return COLOR_MAP["12-8"]
    if vu == "O": return COLOR_MAP["O"]
    if vu in ("休",""): return COLOR_MAP["休"]
    return None

def build_excel(sched, ai_df, month_days, hol_set, personal_targets, ot_count):
    wb = Workbook()

    # ── 工作表1：全彩班表 ──
    ws = wb.active; ws.title = "班表"
    day_cols = [str(d) for d in range(1, month_days+1)]
    headers = ["姓名"] + day_cols
    ws.append(headers)
    ws.row_dimensions[1].height = 20

    for i in ai_df.index:
        name = str(ai_df.at[i, "姓名"]).strip()
        row_data = [name] + [sched[i][d] or "" for d in range(1, month_days+1)]
        ws.append(row_data)
        row_num = ws.max_row
        ws.row_dimensions[row_num].height = 18
        for col_idx, val in enumerate(row_data[1:], start=2):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = val or ""
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=9)
            colors = get_cell_color(val)
            if colors:
                cell.fill = PatternFill("solid", fgColor=colors[0])
                cell.font = Font(size=9, color=colors[1])

    # 每日人數統計列
    for label, shift_key in [("D班人數","D"),("E班人數","E"),("N班人數","N"),("12-8人數","12-8")]:
        stat_row = [label]
        for d in range(1, month_days+1):
            if shift_key == "D":
                cnt = sum(1 for idx in ai_df.index if isinstance(sched[idx][d],str) and sched[idx][d].upper().startswith("D") and sched[idx][d].upper()!="DX")
            else:
                cnt = sum(1 for idx in ai_df.index if sched[idx][d]==shift_key)
            stat_row.append(cnt if cnt else "")
        ws.append(stat_row)
        stat_rn = ws.max_row
        for col_idx in range(2, month_days+2):
            cell = ws.cell(row=stat_rn, column=col_idx)
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(size=8, bold=True)

    # 欄寬
    ws.column_dimensions["A"].width = 8
    for col_idx in range(2, month_days+2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 4.5
        # 假日標頭
        d = col_idx - 1
        cell = ws.cell(row=1, column=col_idx)
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(size=8, bold=True)
        if d in hol_set:
            cell.fill = PatternFill("solid", fgColor="FFE0E0")
        else:
            cell.fill = PatternFill("solid", fgColor="E8F4F8")

    # ── 工作表2：公平性結算 ──
    ws2 = wb.create_sheet("公平性結算")
    ws2.append(["姓名","職稱","目標出勤","實際出勤","狀態","加班天數","假日e��勤"])
    for i in ai_df.index:
        name = str(ai_df.at[i,"姓名"]).strip()
        title = str(ai_df.at[i,"職稱"]).strip() if "職稱" in ai_df.columns else ""
        target = personal_targets.get(i, 0)
        worked = sum(1 for d in sched[i] if is_work(v))
        hol_worked = sum(1 for d in range(1,month_days+1) if d in hol_set and is_work(sched[i][d]))
        ot = ot_count.get(i, 0)
        if worked < target: status = f"⚠️ 欠班 {target-worked} 天"
        elif worked > target: status = f"🟢 超班 {worked-target} 天"
        else: status = "✅ 達標"
        ws2.append([name, title, target, worked, status, ot, hol_worked])
    for col in ["A","B","E"]:
        ws2.column_dimensions[col].width = 12
    for col in ["C","D","F","G"]:
        ws2.column_dimensions[col].width = 8

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ══════════════════════════════════════════
#  Streamlit UI
# ══════════════════════════════════════════
st.set_page_config(page_title="護理白班補足系統", page_icon="🏥", layout="wide")
st.title("🏥 護理白班補足系統")
st.caption("上傳人工班表（已排好 E/N/12-8）→ 自動補足 D 班 + 加班線，輸出完整班表 Excel")

with st.sidebar:
    st.header("⚙️ 系統設定")
    target_off = st.number_input("每人每月應休天數（target_off）", min_value=4, max_value=14, value=10)
    d_buffer   = st.number_input("平日D班配額緩衝（+N人）", min_value=0, max_value=10, value=3)
    first_wday = st.selectbox("本月1日是星期幾（0=週一…6=週日）",
                               options=list(range(7)),
                               format_func=lambda x: ["週一","週二","週三","週四","週五","週六","週日"][x],
                               index=3)  # 預設週四（5月）
    st.divider()
    st.subheader("假日設定")
    sat_input  = st.text_input("週六日期（逗號分隔）", "2,9,16,23,30")
    sun_input  = st.text_input("週日日期（逗號分隔）", "3,10,17,24,31")
    nat_input  = st.text_input("國定假日（逗號分隔）", "1")
    month_days = st.number_input("本月天數", min_value=28, max_value=31, value=31)

st.header("📂 上傳檔案")
col1, col2, col3 = st.columns(3)
with col1:
    manual_file = st.file_uploader("① $��工班表 Excel（含E/N/12-8）", type=["xlsx"], key="manual")
    st.caption("支援跨月格式（前月+本月）")
with col2:
    staff_file  = st.file_uploader("② $��員資料 Excel", type=["xlsx"], key="staff")
    st.caption("需含：姓名、職稱、包班意願、組別、能上週六/週日/國定假日等")
with col3:
    quota_file  = st.file_uploader("③ 每日配額 Excel（選填）", type=["xlsx"], key="quota")
    st.caption("若未上傳，D班配額使用人工班表的每日D班均值")

if manual_file and staff_file:
    if st.button("🚀 開始補足 D 班 + 加班線", type="primary", use_container_width=True):
        with st.spinner("正在執行..."):

            # 解析假日
            def parse_days(s):
                return set(int(x.strip()) for x in s.split(",") if x.strip().isdigit())
            sat_set = parse_days(sat_input)
            sun_set = parse_days(sun_input)
            nat_set = parse_days(nat_input)
            hol_set = sat_set | sun_set | nat_set

            # 讀人員資料（自動偵測標題行與姓名欄）
            _raw_staff = pd.read_excel(staff_file, header=None)
            _staff_hdr_idx = 0
            for _ri in range(min(5, len(_raw_staff))):
                _row_vals = [str(v).strip() for v in _raw_staff.iloc[_ri].values]
                if "姓名" in _row_vals:
                    _staff_hdr_idx = _ri; break
            ai_df = pd.read_excel(staff_file, header=_staff_hdr_idx)
            ai_df.columns = [str(c).strip() for c in ai_df.columns]
            _name_col_s = next((c for c in ai_df.columns if "姓名" in str(c)), None)
            if _name_col_s is None:
                st.error(f"人員資料找不到「姓名」欄，目前欄位：{list(ai_df.columns)[:8]}")
                st.stop()
            ai_df = ai_df.rename(columns={_name_col_s: "姓名"})
            ai_df["姓名"] = ai_df["姓名"].astype(str).str.strip()
            ai_df = ai_df[ai_df["姓名"].notna() & (ai_df["姓名"] != "") & (ai_df["姓名"] != "nan")].reset_index(drop=True)
            ai_names = ai_df["姓名"].tolist()

            # 解析人工班表
            manual_sched, err = parse_manual_schedule(manual_file, ai_names, month_days)
            if err:
                st.error(err)
                st.stop()

            # 建立 sched 字典：index → [""(dummy), day1, day2, ...]
            # 從人工班表的所有班別出發（E/N/12-8/D 都保留；休/O/空格 留空）
            sched = {i: [""] + [""] * month_days for i in ai_df.index}
            for i in ai_df.index:
                for d, v in manual_sched.get(i, {}).items():
                    vu = str(v).strip().upper()
                    # 保留 E/N/12-8（夜班，固定不動）
                    if vu in ("E","E★","E*"): sched[i][d] = "E"
                    elif vu in ("N","N★","N*"): sched[i][d] = "N"
                    elif vu == "12-8": sched[i][d] = "12-8"
                    # D班也保留（包班護師的固定D班）
                    elif vu.startswith("D") and vu != "DX": sched[i][d] = v.strip()
                    # 其他特殊假別（O/公/上課/喪/產檢）也保留
                    elif vu in ("O","公","公差","上課","喪","產檢"): sched[i][d] = v.strip()
                    # 休/空格 留空讓系統補D班

            # 配額處理
            if quota_file:
                quota_df = pd.read_excel(quota_file)
                quota_df["日期"] = quota_df["日期"].astype(str).str.strip()
            else:
                # 從人工班表推算每日D班配額（取已有D班人數）
                quota_rows = []
                for d in range(1, month_days+1):
                    d_cnt = sum(1 for i in ai_df.index
                                if str(manual_sched.get(i,{}).get(d,"")).upper().startswith("D"))
                    e_cnt = sum(1 for i in ai_df.index
                                if str(manual_sched.get(i,{}).get(d,"")).upper() in ("E","E★","E*"))
                    n_cnt = sum(1 for i in ai_df.index
                                if str(manual_sched.get(i,{}).get(d,"")).upper() in ("N","N★","N*"))
                    b_cnt = sum(1 for i in ai_df.index
                                if str(manual_sched.get(i,{}).get(d,"")).upper() == "12-8")
                    quota_rows.append({"日期":str(d),"D班":max(d_cnt,1),"E班":max(e_cnt,1),
                                       "N班":max(n_cnt,1),"12-8":max(b_cnt,1)})
                quota_df = pd.DataFrame(quota_rows)

            # D 班補足
            sched, personal_targets = fill_d_shifts(
                sched, ai_df, quota_df, month_days,
                sat_set, sun_set, nat_set, hol_set, first_wday, target_off
            )

            # 加班線分配
            sched, ot_count = assign_ot_lines(sched, ai_df, month_days, hol_set)

            # 結果統計
            total_worked = sum(sum(1 for v in sched[i] if is_work(v)) for i in ai_df.index)
            total_target = sum(personal_targets.values())
            deficit_nurses = [(str(ai_df.at[i,"姓名"]).strip(),
                               personal_targets[i],
                               sum(1 for v in sched[i] if is_work(v)))
                              for i in ai_df.index
                              if sum(1 for v in sched[i] if is_work(v)) < personal_targets[i]]

        st.success(f"✅ 完成！總出勤 {total_worked} 天 / 目標 {total_target} 天")

        col_a, col_b, col_c = st.columns(3)
        col_a.metric("總出勤天數", total_worked)
        col_b.metric("總目標天數", total_target)
        col_c.metric("欠班人數", len(deficit_nurses))

        if deficit_nurses:
            with st.expander(f"⚠️ {len(deficit_nurses)} 人未達目標出勤", expanded=True):
                deficit_df = pd.DataFrame(deficit_nurses, columns=["姓名","目標","實際"])
                deficit_df["欠班"] = deficit_df["目標"] - deficit_df["實際"]
                st.dataframe(deficit_df.sort_values("欠班", ascending=False), hide_index=True)

        # week_variety 違規掃描
        viol_list = []
        for nm_i in ai_df.index:
            checked = set()
            for d in range(1, month_days+1):
                wd = (first_wday+d-1)%7; dsun = (wd+1)%7; ws = max(1,d-dsun)
                if ws in checked: continue
                checked.add(ws); we = min(month_days, ws+6)
                wtypes = set()
                for dd in range(ws, we+1):
                    t = _shift_norm(sched[nm_i][dd])
                    if t: wtypes.add(t)
                if len(wtypes) >= 3:
                    viol_list.append({"護師": str(ai_df.at[nm_i,"姓名"]).strip(),
                                       "週次": f"{ws}~{we}日", "種類": "/".join(sorted(wtypes))})
        if viol_list:
            with st.expander(f"🌙 week_variety 違規 {len(viol_list)} 件", expanded=False):
                st.dataframe(pd.DataFrame(viol_list), hide_index=True)
        else:
            st.info("✅ 無 week_variety 違規")

        # 輸出 Excel
        excel_bytes = build_excel(sched, ai_df, month_days, hol_set, personal_targets, ot_count)
        st.download_button(
            "📥 下載完整班表 Excel",
            data=excel_bytes,
            file_name="Schedule_Filled.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
else:
    st.info("請上傳①人工班表 和 ②人員資料 後按「開始補足」")
    with st.expander("📋 人員資料欄位說明"):
        st.markdown("""
| 欄位 | 說明 | 必填 |
|------|------|------|
| 姓名 | 護師姓名（需與班表一致）| ✅ |
| 職稱 | 護理長/副護理長/組長/護理師 等 | ✅ |
| 包班意願 | 大夜/小夜/中班（包班護師填寫）| |
| 組別 | A 或 B | |
| 能上週六 | 是/否 | |
| 能上週日 | 是/否 | |
| 能上國定假日 | 是/否 | |
| 預休日期 | 逗號分隔的日期（如 5,12,19）| |
| 特殊假別 | 格式 日期:假別（如 3:病假）| |
| 預約長假日期 | 逗號分隔 | |
| 上課日期 | 逗號分隔 | |
""")
